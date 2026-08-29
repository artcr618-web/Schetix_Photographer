#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Пересчитывает временные копии чистой книги и сверяет сценарии с calc.html."""
from pathlib import Path
from tempfile import TemporaryDirectory
import json,shutil,subprocess
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]; BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа.xlsx'
LO=shutil.which('libreoffice') or shutil.which('soffice')
if not LO: raise SystemExit('Нужен LibreOffice Calc')
regimes=[
 ('НПД 4%',{'C104':'Самозанятый (НПД), Физ. лица'},{'поля':{'regime':'npd','npd_who':'phys'}}),
 ('НПД 5%',{'C104':'Самозанятый (НПД), средняя ставка'},{'поля':{'regime':'npd','npd_who':'mix'}}),
 ('НПД 6%',{'C104':'Самозанятый (НПД), Юр. лица'},{'поля':{'regime':'npd','npd_who':'jur'}}),
 ('УСН 6%',{'C104':'УСН доходы'},{'поля':{'regime':'usn6'}}),
 ('УСН 15%',{'C104':'УСН доходы минус расходы'},{'поля':{'regime':'usn15'}}),
 ('АУСН 8%',{'C104':'АУСН доходы'},{'поля':{'regime':'ausn8'}}),
 ('АУСН 20%',{'C104':'АУСН доходы минус расходы'},{'поля':{'regime':'ausn20'}}),
]
scenarios=regimes+[
 ('Налоги выключены',{'C154':1},{'поля':{'tax_off':True}}),
 ('Резервы включены',{'C155':1,'C157':1},{'поля':{'fund_on':True,'disc_on':True}}),
 ('Сайт самостоятельно',{'C81':'Делал (а) самостоятельно'},{'радио':{'site_mode':'self'}}),
 ('Дополнительная комиссия',{'C159':.015},{'допКомиссии':1.5}),
 ('Собственное жильё',{'C64':'да'},{'поля':{'own_home':True}}),
 ('Рабочее место исключено',{'C219':1},{'EXC_ВНЕШ':{'Form009b':True}}),
 ('Обучение исключено',{'C220':1},{'EXC_ВНЕШ':{'Form006':True}}),
 ('Сайт исключён',{'C221':1},{'EXC_ВНЕШ':{'Form014':True}}),
]
checks={'C164':'R','C165':'taxAll','C166':'aq','C168':'Rb','C169':'taxB','C170':'aqB','C171':'rateZero','C180':'taxC','C186':'currentResult','C196':'costHour','C197':'markup'}
for idx,(title,cells,override) in enumerate(scenarios):
    with TemporaryDirectory() as td:
        td=Path(td); inp=td/f's{idx}.xlsx'; out=td/'out'; prof=td/'profile'; out.mkdir();prof.mkdir()
        wb=load_workbook(BOOK,data_only=False); ws=wb['calc']
        for cell,value in cells.items(): ws[cell]=value
        wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'; wb.save(inp)
        p=subprocess.run([LO,f'-env:UserInstallation=file://{prof}','--headless','--convert-to','xlsx','--outdir',str(out),str(inp)],capture_output=True,text=True)
        if p.returncode or not (out/inp.name).exists(): raise SystemExit(f'{title}: LibreOffice не пересчитал')
        values=load_workbook(out/inp.name,data_only=True)
        d=json.loads(subprocess.check_output(['node',str(ROOT/'Инструменты'/'харнесс.js'),str(ROOT),json.dumps(override,ensure_ascii=False)],text=True))
        bad=[]
        for cell,key in checks.items():
            val=values['calc'][cell].value; exp=d[key]
            if not isinstance(val,(int,float)) or abs(val-exp)>1e-5: bad.append(f'{cell}/{key}: {val} ≠ {exp}')
        if bad: raise SystemExit(title+': '+'; '.join(bad))
        print('✓',title)
print(f'Сценариев сверено: {len(scenarios)}; расхождений 0')
