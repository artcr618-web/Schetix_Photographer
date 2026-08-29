#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Собирает чистый лист Программа_лояльности по фактической логике report.html."""
from pathlib import Path
import json, math, subprocess
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT=Path(__file__).resolve().parents[1]
BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа.xlsx'
HARNESS=ROOT/'Инструменты'/'харнесс.js'
wb=load_workbook(BOOK,data_only=False); ws=wb['Программа_лояльности']
ws.delete_rows(1,ws.max_row)
HEAD=['Раздел','ID','Наименование','Значение','Единица','Формула / правило','Зависит от','Поле d','Источник','Статус']
ws.append(HEAD)

def add(section,ident,title,value,unit='',rule='',depends='',field='',source='Веб/report.html',status='синхронизировано'):
    ws.append([section,ident,title,value,unit,rule,depends,field,source,status])
    row=ws.max_row
    if ident:
        if ident in wb.defined_names: del wb.defined_names[ident]
        wb.defined_names.add(DefinedName(ident,attr_text=f"'Программа_лояльности'!$D${row}"))
    return row

# Входы и утверждённые ограничения.
add('Вход','loyalty_on','Резерв на программу лояльности / включение','=disc_on','0/1','Выключено → вся модель получает ноль','disc_on','d.discP')
add('Вход','loyalty_fund','Фонд программы лояльности','=ROUND(discount_year_model,0)','₽/год','Финальная сумма фонда округляется для пользовательского распределения','discount_year_model','d.discY')
add('Вход','loyalty_shoot_duration','Продолжительность съёмки / для программы лояльности','=MAX(ROUND(shoot_duration,0),1)','ч','Не меньше одного целого часа','shoot_duration','d.S')
add('Вход','loyalty_base_hour','Желаемая ставка / округлённая для Прайса','=IF(project_pool=0,0,CEILING(((loyalty_shoot_duration*profile_multiplier+client_time)*(revenue_target/project_pool)/loyalty_shoot_duration),100))','₽/ч','Округление вверх до 100 ₽','project_pool, profile_multiplier, client_time, revenue_target','d.R, d.pool')
add('Вход','loyalty_check','Средний чек съёмки','=loyalty_base_hour*loyalty_shoot_duration','₽','Ставка Прайса × Продолжительность съёмки','loyalty_base_hour, loyalty_shoot_duration','','Веб/report.html · чек')

# Сетка Скидки за объём строится от минимальной Продолжительности съёмки
# пользователя; следующие карточки идут через +2, +4 и +6 часов.
add('Скидка за объём','volume_base_price','Съёмочный проект / базовая стоимость','=loyalty_check','₽','Без скидки','loyalty_check','','Веб/report.html · dmap')
add('Скидка за объём','volume_base_discount','Скидка за объём / базовая',0,'%','Базовая карточка')
for key,delta in [('step1',2),('step2',4),('step3',6)]:
    add('Скидка за объём',f'volume_{key}_duration',f'Продолжительность съёмки / ступень +{delta}',f'=loyalty_shoot_duration+{delta}','ч')
    add('Скидка за объём',f'volume_{key}_hour',f'Желаемая ставка / ступень +{delta}',f'=IF(project_pool=0,0,CEILING(((volume_{key}_duration*profile_multiplier+client_time)*(revenue_target/project_pool)/volume_{key}_duration),100))','₽/ч','Округление вверх до 100 ₽',f'volume_{key}_duration, profile_multiplier, client_time, revenue_target, project_pool')
    add('Скидка за объём',f'volume_{key}_price',f'Съёмочный проект / стоимость / ступень +{delta}',f'=volume_{key}_hour*volume_{key}_duration','₽')
    add('Скидка за объём',f'volume_{key}_discount',f'Скидка за объём / ступень +{delta}',f'=IF(loyalty_base_hour=0,0,1-volume_{key}_hour/loyalty_base_hour)','%')

add('Правило','repeat_fund_limit','Программа лояльности / предел фонда для постоянных клиентов',0.45,'%','Не более 45% Фонда программы лояльности','','','Паспорт Резервного фонда')
add('Правило','repeat_orders_limit','Программа лояльности / предел заказов постоянных клиентов',0.30,'%','Не более 30% годового потока Съёмочных проектов','','','Паспорт Резервного фонда')
add('Расчёт','repeat_budget','Программа лояльности / бюджет постоянных клиентов','=loyalty_fund*repeat_fund_limit','₽/год','','loyalty_fund, repeat_fund_limit')
add('Расчёт','repeat_orders','Программа лояльности / целевой поток постоянных клиентов','=projects_year*repeat_orders_limit','заказов/год','','projects_year, repeat_orders_limit','d.py')

# Ступени: ставка, визиты и доля внутреннего бюджета 40/40/20.
for key,title,rate,visits,weight in [('5','Раз в полгода',.05,2,.40),('7','Раз в квартал',.07,4,.40),('10','Раз в месяц',.10,12,.20)]:
    add('Ступень '+key,f'tier{key}_rate',f'Скидка постоянному клиенту / {title} / ставка',rate,'%')
    add('Ступень '+key,f'tier{key}_visits',f'Скидка постоянному клиенту / {title} / съёмок',visits,'шт/год')
    add('Ступень '+key,f'tier{key}_weight',f'Скидка постоянному клиенту / {title} / доля бюджета',weight,'%')

# Последовательное распределение: остатки предыдущей ступени ограничивают следующую.
def raw_formula(k,rem_b,rem_o):
    return f'=IFERROR(IF(loyalty_on=0,0,FLOOR(MIN(MIN(repeat_budget*tier{k}_weight,{rem_b})/(tier{k}_visits*loyalty_check*tier{k}_rate),MIN(repeat_orders*tier{k}_weight,{rem_o})/tier{k}_visits),1)),0)'
add('Расчёт ступеней','tier5_raw_clients','Ступень 5% / клиентов до резервного правила',raw_formula('5','repeat_budget','repeat_orders'),'чел.')
add('Расчёт ступеней','tier5_raw_orders','Ступень 5% / заказов','=tier5_raw_clients*tier5_visits','шт')
add('Расчёт ступеней','tier5_raw_sum','Ступень 5% / сумма скидок','=ROUND(tier5_raw_orders*loyalty_check*tier5_rate,0)','₽')
add('Расчёт ступеней','tier7_raw_clients','Ступень 7% / клиентов',raw_formula('7','repeat_budget-tier5_raw_sum','repeat_orders-tier5_raw_orders'),'чел.')
add('Расчёт ступеней','tier7_orders','Ступень 7% / заказов','=tier7_raw_clients*tier7_visits','шт')
add('Расчёт ступеней','tier7_sum','Ступень 7% / сумма скидок','=ROUND(tier7_orders*loyalty_check*tier7_rate,0)','₽')
add('Расчёт ступеней','tier10_raw_clients','Ступень 10% / клиентов',raw_formula('10','repeat_budget-tier5_raw_sum-tier7_sum','repeat_orders-tier5_raw_orders-tier7_orders'),'чел.')
add('Расчёт ступеней','tier10_orders','Ступень 10% / заказов','=tier10_raw_clients*tier10_visits','шт')
add('Расчёт ступеней','tier10_sum','Ступень 10% / сумма скидок','=ROUND(tier10_orders*loyalty_check*tier10_rate,0)','₽')
add('Расчёт ступеней','tier_raw_clients_total','Постоянные клиенты / всего до резервного правила','=tier5_raw_clients+tier7_raw_clients+tier10_raw_clients','чел.')
fallback='AND(tier_raw_clients_total=0,ROUND(tier5_visits*loyalty_check*tier5_rate,0)<=repeat_budget,tier5_visits<=repeat_orders)'
add('Расчёт ступеней','tier5_clients','Ступень 5% / итог клиентов',f'=IF({fallback},1,tier5_raw_clients)','чел.')
add('Расчёт ступеней','repeat_discount_sum','Постоянные клиенты / сумма скидок',f'=IF(loyalty_on=0,0,IF({fallback},ROUND(tier5_visits*loyalty_check*tier5_rate,0),tier5_raw_sum+tier7_sum+tier10_sum))','₽/год')
add('Расчёт ступеней','repeat_orders_total','Постоянные клиенты / заказов',f'=IF(loyalty_on=0,0,IF({fallback},tier5_visits,tier5_raw_orders+tier7_orders+tier10_orders))','шт/год')
add('Проверка','repeat_orders_share','Постоянные клиенты / доля заказов','=IF(projects_year=0,0,repeat_orders_total/projects_year)','%','','repeat_orders_total, projects_year')

# Сертификаты: максимум 30% фонда, номиналы 85/40/25% чека, шаг 500 ₽.
add('Сертификаты','cert_budget_limit','Подарочные сертификаты / предел фонда',.30,'%')
add('Сертификаты','cert_budget','Подарочные сертификаты / бюджет','=MIN(loyalty_fund*cert_budget_limit,MAX(loyalty_fund-repeat_discount_sum,0))','₽/год')
for key,share in [('high',.85),('mid',.40),('low',.25)]:
    add('Сертификаты',f'cert_{key}_share',f'Подарочный сертификат / номинал {share*100:.0f}% чека / доля',share,'%')
    add('Сертификаты',f'cert_{key}_value',f'Подарочный сертификат / номинал {share*100:.0f}% чека',f'=MAX(ROUND(loyalty_check*cert_{key}_share/500,0)*500,500)','₽')
add('Сертификаты','cert_high_count','Подарочный сертификат / большой / количество','=IF(cert_budget>=cert_high_value,1,0)','шт')
add('Сертификаты','cert_after_high','Подарочные сертификаты / остаток после большого','=cert_budget-cert_high_value*cert_high_count','₽')
add('Сертификаты','cert_mid_count','Подарочный сертификат / средний / количество','=MAX(0,FLOOR(cert_after_high*0.45/cert_mid_value,1))','шт')
add('Сертификаты','cert_after_mid','Подарочные сертификаты / остаток после среднего','=cert_after_high-cert_mid_value*cert_mid_count','₽')
add('Сертификаты','cert_low_count','Подарочный сертификат / малый / количество','=MAX(0,FLOOR(cert_after_mid/cert_low_value,1))','шт')
add('Сертификаты','certificates_sum','Подарочные сертификаты / сумма','=cert_high_value*cert_high_count+cert_mid_value*cert_mid_count+cert_low_value*cert_low_count','₽/год')
add('Распределение','volume_discount_sum','Скидка за объём / остаток фонда','=MAX(loyalty_fund-repeat_discount_sum-certificates_sum,0)','₽/год')
add('Проверка','loyalty_total','Программа лояльности / распределено','=repeat_discount_sum+certificates_sum+volume_discount_sum','₽/год')
add('Проверка','loyalty_balance','Программа лояльности / расхождение','=loyalty_total-loyalty_fund','₽','Должно быть 0')

# Независимые контрольные снимки настоящего calc() + алгоритма report.html.
def model(d):
    rRate=lambda n: math.ceil((n or 0)/100)*100
    S=max(round(d['S'] or 2),1); hp=d['R']/d['pool'] if d['pool'] else 0
    check=rRate((S*(1+d['K'])+d['cl'])*hp/S)*S; fund=round(d['discY'] or 0)
    post=orders=people=0; budget=fund*.45; target=d['py']*.30; rb,ro=budget,target
    for p,v,w in ((.05,2,.40),(.07,4,.40),(.10,12,.20)):
        pc=v*check*p; n=max(0,math.floor(min(min(budget*w,rb)/pc if pc else 0,min(target*w,ro)/v)))
        sm=round(n*v*check*p); rb-=sm; ro-=n*v; post+=sm; orders+=n*v; people+=n
    if people==0:
        sm=round(2*check*.05)
        if sm<=budget and 2<=target: post=sm; orders=2
    r500=lambda n:max(round(n/500)*500,500); cb=min(fund*.30,max(fund-post,0))
    hi,mid,low=r500(check*.85),r500(check*.40),r500(check*.25); rest=cb
    n1=1 if rest>=hi else 0; rest-=hi*n1; n2=max(0,math.floor(rest*.45/mid)); rest-=mid*n2; n3=max(0,math.floor(rest/low))
    cert=hi*n1+mid*n2+low*n3; volume=max(fund-post-cert,0)
    return fund,post,cert,volume,orders

for pct in (5,10,15):
    raw=subprocess.check_output(['node',str(HARNESS),str(ROOT),json.dumps({'поля':{'disc_on':True,'disc_pct':str(pct)}})],text=True)
    vals=model(json.loads(raw)); names=('Фонд программы лояльности','Постоянные клиенты','Подарочные сертификаты','Скидка за объём','Заказы постоянных клиентов')
    units=('₽/год','₽/год','₽/год','₽/год','шт/год')
    for name,value,unit in zip(names,vals,units):
        add(f'Контроль {pct}%', '', name, value, unit, 'Контрольный снимок настоящего сценария','','','calc.html + report.html','контроль')

# Оформление.
for c in ws[1]: c.fill=PatternFill('solid',fgColor='1B9331'); c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(vertical='center',wrap_text=True)
ws.row_dimensions[1].height=38; border=Border(bottom=Side(style='thin',color='E5E7EB'))
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment=Alignment(vertical='top',wrap_text=True); c.border=border
ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:J{ws.max_row}'; ws.sheet_view.showGridLines=False
widths=[24,28,58,24,14,72,42,22,36,20]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
wb.save(BOOK)
print(f'Программа_лояльности: {ws.max_row-1} строк, формул {sum(c.data_type=="f" for row in ws.iter_rows() for c in row)}')
