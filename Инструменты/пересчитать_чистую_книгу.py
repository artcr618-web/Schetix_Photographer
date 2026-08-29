#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Пересчитывает чистую книгу совместимым LibreOffice и сверяет с calc.html."""
from pathlib import Path
from tempfile import TemporaryDirectory
import json,shutil,subprocess,sys
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа.xlsx'
LO=shutil.which('libreoffice') or shutil.which('soffice')
if not LO:
    # Пакеты среды не сохраняются между запусками Agent Mode. Устанавливаем
    # движок автоматически в том же процессе, чтобы `карта.py` была одной
    # воспроизводимой командой.
    subprocess.check_call(['sudo','apt-get','update','-qq'])
    subprocess.check_call(['sudo','apt-get','install','-y','-qq','--no-install-recommends','libreoffice-calc'])
    LO=shutil.which('libreoffice') or shutil.which('soffice')
if not LO: raise SystemExit('Не удалось установить LibreOffice Calc')

with TemporaryDirectory() as td:
    td=Path(td); inp=td/'Schetix_clean_recalc.xlsx'; out=td/'out'; profile=td/'profile'
    out.mkdir(); profile.mkdir(); shutil.copy2(BOOK,inp)
    p=subprocess.run([LO,f'-env:UserInstallation=file://{profile}','--headless','--convert-to','xlsx','--outdir',str(out),str(inp)],capture_output=True,text=True)
    if p.returncode or not (out/inp.name).exists():
        raise SystemExit('LibreOffice не пересчитал книгу: '+(p.stderr or p.stdout))
    result=out/inp.name
    original=load_workbook(BOOK,data_only=False); formulas=load_workbook(result,data_only=False); values=load_workbook(result,data_only=True)
    if formulas.sheetnames!=original.sheetnames or len(formulas.defined_names)!=len(original.defined_names):
        raise SystemExit('После пересчёта изменилась структура книги или определённые имена')
    errors=[]
    for ws in formulas.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.data_type=='f':
                    val=values[ws.title][c.coordinate].value
                    if val is None or (isinstance(val,str) and val.startswith('#')):
                        errors.append(f'{ws.title}!{c.coordinate}={val}')
    if errors: raise SystemExit('Ошибки формул: '+', '.join(errors[:12]))
    d=json.loads(subprocess.check_output(['node',str(ROOT/'Инструменты'/'харнесс.js'),str(ROOT)],text=True))
    checks={'C164':'R','C165':'taxAll','C167':'totalExpenses','C168':'Rb','C171':'rateZero','C175':'rateHour','C180':'taxC','C186':'currentResult','C196':'costHour','C197':'markup'}
    bad=[]
    for cell,key in checks.items():
        val=values['calc'][cell].value
        if not isinstance(val,(int,float)) or abs(val-d[key])>1e-6: bad.append(f'{cell}: книга {val} ≠ calc.{key} {d[key]}')
    if bad: raise SystemExit('Расхождение с calc.html: '+'; '.join(bad))
    shutil.copy2(result,BOOK)
print('Чистая книга пересчитана: ошибок формул 0; ключевые значения совпадают с calc.html')
