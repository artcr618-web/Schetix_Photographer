# -*- coding: utf-8 -*-
"""Сверка наименований книги и документа группировок со СПРАВОЧНИКОМ.

Справочник (Инструменты/таблицы_прототип.py) — единственный источник названий.
Скрипт показывает, каким наименованиям определение не дано.

    python3 сверить_глоссарий.py /home/user/schetix
"""
import sys, re, importlib.util, unicodedata

корень = sys.argv[1] if len(sys.argv) > 1 else '/home/user/schetix'

spec = importlib.util.spec_from_file_location('тп', корень + '/Инструменты/таблицы_прототип.py')
м = importlib.util.module_from_spec(spec)
try: spec.loader.exec_module(м)
except SystemExit: pass
СПР = м.СПРАВОЧНИК

def ключ(с):
    с = unicodedata.normalize('NFKC', str(с)).lower().replace('ё', 'е')
    с = re.sub(r'\(.*?\)', ' ', с)
    return re.sub(r'[^а-яa-z0-9]+', ' ', с).strip()

ИНДЕКС = {ключ(k): k for k in СПР}

def есть(имя):
    к = ключ(имя)
    if к in ИНДЕКС: return ИНДЕКС[к]
    for эк, ор in ИНДЕКС.items():
        if к and (к == эк or к.startswith(эк + ' ') or эк.startswith(к + ' ')): return ор
    return None

import openpyxl
кн = openpyxl.load_workbook(корень + '/Книга/Калькулятор_ставки_часа.xlsx', data_only=True)

СЛУЖЕБНОЕ = re.compile(r'^(наименование|термин|было|стало|прочие|\d+\s*·|[A-Z]\s*·|ПЕРЕКЛЮЧАТЕЛЬ)', re.I)

def имена_листа(лист, столбец=1):
    из = []
    for стр in кн[лист].iter_rows(min_col=столбец, max_col=столбец, values_only=True):
        з = стр[0]
        if not isinstance(з, str): continue
        з = з.strip()
        if not з or СЛУЖЕБНОЕ.match(з) or len(з) > 60: continue
        из.append(з)
    return из

for лист in ['01_Глоссарий', '09i_Справочник', '09f_Группы_затрат']:
    if лист not in кн.sheetnames: continue
    имена = имена_листа(лист)
    нет = [и for и in имена if not есть(и)]
    print('=' * 70)
    print(f'{лист}: {len(имена)} наименований · без понятия в справочнике — {len(нет)}')
    for и in нет: print('   ·', и)

print('=' * 70)
print(f'В справочнике {len(СПР)} записей')
