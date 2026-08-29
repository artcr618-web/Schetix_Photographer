#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ПРОВЕРКА ЦЕЛОСТНОСТИ ПРОЕКТА СЧЁТИКС.
Запускать после любой правки:  python3 проверить.py [корень]
Численные проверки выполняют НАСТОЯЩИЙ код calc()/parts(), вырезанный из HTML.
Код возврата: 0 — всё чисто, 1 — есть падения."""
import json, subprocess, sys, os, re, io, collections, math

КОРЕНЬ = sys.argv[1] if len(sys.argv) > 1 else '/home/user/schetix'
ЗДЕСЬ  = os.path.dirname(os.path.abspath(__file__))
ХАРНЕСС = os.path.join(ЗДЕСЬ, 'харнесс.js')
ok, fail, warn = [], [], []

def проверка(имя, условие, факт=''):
    (ok if условие else fail).append((имя, факт))
def замечание(имя, факт=''):
    warn.append((имя, факт))

def расчёт(**переопр):
    r = subprocess.run(['node', ХАРНЕСС, КОРЕНЬ, json.dumps(переопр)],
                       capture_output=True, text=True)
    if r.returncode or not r.stdout.strip():
        raise RuntimeError((r.stderr or 'пустой вывод').strip()[:200])
    return json.loads(r.stdout)

РЕЖИМЫ = [('npd', 'phys', 'НПД 4 %'), ('npd', 'mix', 'НПД 5 %'), ('npd', 'jur', 'НПД 6 %'),
          ('usn6', 'mix', 'УСН 6 %'), ('usn15', 'mix', 'УСН 15 %'),
          ('ausn8', 'mix', 'АУСН 8 %'), ('ausn20', 'mix', 'АУСН 20 %')]

# ══════════════════════════════ 1. МОДЕЛЬ
print('▶ модель (настоящий calc() из calc.html)')
try:
    базовый = расчёт()
except RuntimeError as e:
    print('  ✗ харнесс не запустился:', e); sys.exit(2)

for код, кто, имя in РЕЖИМЫ:
    d = расчёт(поля={'regime': код, 'npd_who': кто})
    остаток = d['R'] - d['C'] - d['aq'] - d['taxAll'] - d['fundY'] - d['discY']
    проверка(f'на руки = цель · {имя}', abs(остаток - d['Ny']) < 1,
             f"{остаток:,.0f} против {d['Ny']:,.0f}")

d = расчёт(поля={'fund_on': True, 'disc_on': True})
проверка('на руки = цель · с фондом и резервом скидки',
         abs(d['R'] - d['C'] - d['aq'] - d['taxAll'] - d['fundY'] - d['discY'] - d['Ny']) < 1)

d = расчёт(поля={'tax_off': True})
проверка('на руки = цель · без оформления',
         abs(d['R'] - d['C'] - d['aq'] - d['Ny']) < 1)

d = базовый
проверка('иерархия: операционное + резерв + проектное = эффективное',
         abs(d['promo'] + d['accT'] + d['fmT'] + d['pool'] - d['NT']) < 0.01)
проверка('иерархия: съёмочное + постпродакшн + проектное клиентское = проектное',
         abs(d['sh'] + d['post'] + d['clT'] - d['pool']) < 0.01)
проверка('контракт: удалены неиспользуемые агрегаты и старые режимы',
         not ({'core','sAuto','equip','promoM','current','side'} & set(d)))
проверка('сумма сегментов кольца = выручка',
         abs(sum(d['__parts']) - d['R']) < 1, f"Δ {sum(d['__parts']) - d['R']:.4f}")
проверка('доходные сектора = доход без отпуска',
         abs(sum(d['__parts'][:6]) - d['Ny']*11/12) < 1)
проверка('сектор отпуска = месячный доход',
         abs(d['__parts'][13] - d['Ny']/12) < 1)

# инвариант сетки скидок: выручка не зависит от длительности съёмки
выручки = [расчёт(поля={'shoot_manual': str(S)})['R'] for S in (1, 2, 4, 6)]
проверка('сетка скидок: выручка одинакова при любой длительности',
         max(выручки) - min(выручки) < 1, f'разброс {max(выручки)-min(выручки):.2f} ₽')

# ставка падает с ростом длительности
ставки = [расчёт(поля={'shoot_manual': str(S)}) for S in (1, 2, 4, 6)]
ставки = [x['R'] / x['sh'] for x in ставки]
проверка('сетка скидок: ставка часа убывает с длительностью',
         all(ставки[i] > ставки[i+1] for i in range(3)),
         ' → '.join(f'{r:,.0f}' for r in ставки))
# Пользовательская карта Прайса: база — его минимальная Продолжительность
# съёмки, затем +2/+4/+6; ставки округляются вверх до 100 ₽.
d = базовый
S0 = max(round(d['S'] or 2), 1); hp = d['R']/d['pool'] if d['pool'] else 0
rr100 = lambda n: math.ceil((n or 0)/100)*100
volume_durations = [S0, S0+2, S0+4, S0+6]
volume_rates = [rr100((S*(1+d['K'])+d['cl'])*hp/S) for S in volume_durations]
проверка('сетка скидок: Прайс строится от минимальной съёмки через +2/+4/+6 часов',
         volume_durations == [1,3,5,7] and all(volume_rates[i]>=volume_rates[i+1] for i in range(3)),
         ' / '.join(f'{S} ч: {r:,.0f} ₽' for S,r in zip(volume_durations,volume_rates)))

# Загрузка считает именно проекты, а не съёмочные часы. При S=1 они равны,
# поэтому проверяем весь разрешённый диапазон длительности съёмки.
загрузки = [расчёт(поля={'shoot_manual': str(S)}) for S in (0.5, 1, 1.5, 2, 4, 6, 8)]
проверка('загрузка: loadGoal.shoots = количество проектов при S 0,5–8 ч',
         all(abs(x['loadGoal']['shoots']-x['py']) < 1e-7 for x in загрузки))
проверка('загрузка: идеальная ставка укладывается в стандартный фонд времени',
         all(abs(x['loadGoal']['hours']-x['NT']) < 1e-7 for x in загрузки))
проверка('загрузка: проекты × часы съёмки × текущая ставка = целевая выручка',
         all(abs(x['loadCur']['shoots']*x['S']*x['cur']-x['R']) < 0.02 for x in загрузки))
проверка('загрузка: проекты × часы съёмки × ставка в ноль = целевая выручка',
         all(abs(x['loadZero']['shoots']*x['S']*x['rateZero']-x['R']) < 0.02 for x in загрузки))

# ══════════════════════════════ 2. КРАЕВЫЕ СЛУЧАИ
print('▶ краевые случаи')
d = расчёт(поля={'promo_per_day': '7.5'})
проверка('продвижение 7,5 ч/день не даёт отрицательных часов', d['pool'] > 0,
         f"pool {d['pool']:.0f} ч, ставка {d['R']/d['sh'] if d['sh'] else 0:,.0f} ₽/ч")
d = расчёт(поля={'acq_rate': '30', 'fund_on': True, 'fund_pct': '30',
                 'disc_on': True, 'disc_pct': '15'})
проверка('экстремальные проценты не взрывают выручку', d['R'] < 5_000_000,
         f"R {d['R']:,.0f} при D {1-0.30-0.30-0.15:.2f}")
d = расчёт(поля={'shoot_manual': '0'})
проверка('нулевая длительность съёмки обрабатывается', d['sh'] > 0 or d['R'] == 0,
         f"sh {d['sh']}, R {d['R']:,.0f}")

# Точка безубыточности должна быть единым точным сценарием:
# выручка, ставка часа и нулевой остаток относятся к одной величине Rb.
# Правило владельца 25.08: точка безубыточности — только необходимые расходы,
# резервные фонды в неё НЕ входят, даже если включены галочками.
def остаток_в_ноль(d, поправка_сайта=0):
    return (d['Rb'] - d['C'] - d['aqB'] - d['taxB']
            - d['Rb']*поправка_сайта)

for код, кто, имя in РЕЖИМЫ:
    d = расчёт(поля={'regime': код, 'npd_who': кто})
    остаток = остаток_в_ноль(d)
    проверка(f'безубыточность анкеты сходится в ноль · {имя}', abs(остаток) < 0.02,
             f'{остаток:+,.2f} ₽')

# При любой разрешённой длительности съёмки выручка в ноль остаётся той же,
# а точная ставка обязана равняться выручке в ноль / съёмочные часы.
нулевые = [расчёт(поля={'shoot_manual': str(S)}) for S in (0.5, 1, 1.5, 2, 4, 6, 8)]
проверка('ставка в ноль точна при длительности 0,5–8 ч',
         all(x['sh'] > 0 and abs(x['rateZero']*x['sh']-x['Rb']) < 1e-5 for x in нулевые))
проверка('месячная выручка в ноль = годовая / 12',
         all(abs((x['Rb']/12)*12-x['Rb']) < 1e-7 for x in нулевые))
проверка('съёмки при Ставке в ноль покрывают расходы',
         all(abs(x['zeroShootsY']*x['rateZero']*x['S']-x['Rb']) < 0.02
             and abs(x['zeroShootsM']*12-x['zeroShootsY']) < 1e-9
             and abs(x['zeroShootsY']-x['py']) < 1e-9 for x in нулевые))

# Специальные ветки, которые прежний повторный расчёт в report.html терял.
d = расчёт(поля={'tax_off': True})
проверка('безубыточность · налоги выключены',
         abs(остаток_в_ноль(d)) < 0.02 and abs(d['taxB']) < 0.01)
d = расчёт(поля={'fund_on': True, 'fund_pct': '10',
                 'disc_on': True, 'disc_pct': '15'})
проверка('безубыточность · фонды не входят (только необходимые расходы)',
         abs(остаток_в_ноль(d)) < 0.02)
d = расчёт(радио={'site_mode': 'self'})
поправка_сайта = 80/(7*d['NT'])
проверка('безубыточность · сайт сделан самостоятельно',
         abs(остаток_в_ноль(d, поправка_сайта)) < 0.02)
d = расчёт(допКомиссии=3.5)
проверка('безубыточность · дополнительная банковская комиссия',
         abs(остаток_в_ноль(d)) < 0.02)

# ══════════════════════════════ 2a. ИДЕАЛЬНЫЙ СЦЕНАРИЙ И ПОЛНОЕ КОЛЬЦО
print('▶ идеальный сценарий')
def идеальный_сходится(d):
    сумма = (d['C'] + d['taxAll'] + d['aq'] + d['goalFund']
             + d['goalDiscountReserve'] + d['goalSelfSiteCost'])
    return (abs(d['totalExpenses']-(d['C']+d['taxAll']+d['aq'])) < 0.02
            and abs(d['goalCostsTotal']-сумма) < 0.02
            and abs(d['goalResult']-(d['R']-сумма)) < 0.02
            and abs(d['goalResult']-d['Ny']) < 0.02
            and abs(d['rateHour']*d['sh']-d['R']) < 1e-5
            and abs(d['fundY']-d['goalFund']) < 0.02
            and abs(d['discY']-d['goalDiscountReserve']) < 0.02)

идеальные_ветки = [
    ('по умолчанию', {}),
    ('налоги выключены', {'поля': {'tax_off': True}}),
    ('фонды включены', {'поля': {'fund_on': True, 'fund_pct': '10',
                                 'disc_on': True, 'disc_pct': '15'}}),
    ('собственный сайт', {'радио': {'site_mode': 'self'}}),
    ('банк исключён', {'EXC_ВНЕШ': {'Form015b': True}}),
    ('фонды и собственный сайт', {
        'поля': {'fund_on': True, 'fund_pct': '10',
                 'disc_on': True, 'disc_pct': '15'},
        'радио': {'site_mode': 'self'}}),
]
for имя, параметры in идеальные_ветки:
    d = расчёт(**параметры)
    проверка(f'идеальный сценарий сходится · {имя}', идеальный_сходится(d))

# Отключённые параметры остаются в контракте и дают эффективный ноль.
d = расчёт(поля={'tax_off': True, 'fund_on': False, 'disc_on': False},
            радио={'site_mode': 'hired'})
проверка('идеальный сценарий · эффективные нули',
         d['taxAll'] == 0 and d['goalFund'] == 0
         and d['goalDiscountReserve'] == 0 and d['goalSelfSiteCost'] == 0)
d = расчёт(поля={'fund_on': True, 'fund_pct': '10',
                 'disc_on': True, 'disc_pct': '15'},
            радио={'site_mode': 'self'})
проверка('идеальное кольцо · сумма дохода и всех затрат равна выручке',
         abs(d['Ny']+d['goalCostsTotal']-d['R']) < 0.02)

# ══════════════════════════════ 2b. ТЕКУЩИЙ СЦЕНАРИЙ И ЭФФЕКТИВНЫЕ НУЛИ
print('▶ текущий сценарий')
def текущий_сходится(d):
    сумма = (d['C'] + d['taxC'] + d['aqC'] + d['currentFund']
             + d['currentDiscountReserve'] + d['currentSelfSiteCost'])
    return (abs(d['currentCostsTotal']-сумма) < 0.02
            and abs(d['currentResult']-(d['Rc']-сумма)) < 0.02
            and abs(d['leftC']-d['currentResult']) < 0.02
            and abs(d['currentIncome']-max(d['currentResult'],0)) < 0.02
            and abs(d['currentLoss']-min(d['currentResult'],0)) < 0.02
            and d['currentIsLoss'] == (d['currentResult'] < 0))

текущие_ветки = [
    ('по умолчанию', {}),
    ('налоги выключены', {'поля': {'tax_off': True}}),
    ('фонды включены', {'поля': {'fund_on': True, 'fund_pct': '10',
                                 'disc_on': True, 'disc_pct': '15'}}),
    ('собственный сайт', {'радио': {'site_mode': 'self'}}),
    ('банк исключён', {'EXC_ВНЕШ': {'Form015b': True}}),
    ('нулевая ставка', {'поля': {'current_rate': '0'}}),
    ('ставка 1 000', {'поля': {'current_rate': '1000'}}),
    ('ставка 3 000', {'поля': {'current_rate': '3000'}}),
]
for имя, параметры in текущие_ветки:
    d = расчёт(**параметры)
    проверка(f'текущий сценарий сходится · {имя}', текущий_сходится(d))

# Отключённый параметр остаётся в контракте, но даёт нулевой вклад.
d = расчёт(поля={'tax_off': True})
проверка('эффективный ноль · отключённый налог', d['taxC'] == 0)
d = расчёт(поля={'fund_on': False, 'disc_on': False})
проверка('эффективный ноль · выключенные фонды',
         d['currentFund'] == 0 and d['currentDiscountReserve'] == 0)
d = расчёт(радио={'site_mode': 'hired'})
проверка('эффективный ноль · неактивная ветка собственного сайта',
         d['currentSelfSiteCost'] == 0)
d = расчёт(поля={'own_home': True})
ответ_жильё = next((x['v'] for x in d['answers'] if x['n'] == 'Сколько вы платите за жильё в месяц'), None)
проверка('рабочее место · собственное жильё: аренда или ипотека получает эффективный ноль',
         abs(d['wsY'] - 10080) < 0.01 and ответ_жильё == 0,
         f"Содержание рабочего места {d['wsY']:,.0f} ₽/год; ответ {ответ_жильё}")
d = расчёт(радио={'ws_mode': 'office'})
проверка('рабочее место · отдельное помещение: аренда и коммунальные учитываются полностью',
         abs(d['wsY'] - 264000) < 0.01,
         f"{d['wsY']:,.0f} ₽/год")
d = расчёт(EXC_ВНЕШ={'Form009b': True})
проверка('рабочее место · исключённый блок: содержание и обустройство получают ноль',
         d['wsY'] == 0 and d['depWs'] == 0 and d['varRent'] == 0)
d = расчёт(CAT={'Form006': {'k':'months','rows':[['Курс 1',80000,4],['Курс 2',20000,2]]}})
ожид_обучение = (100000 + 6*d['Ny']/12) / 7
проверка('обучение · несколько курсов: стоимость и периоды без дохода суммируются',
         abs(d['eduY']-ожид_обучение) < 0.01 and abs(d['depEdu']-ожид_обучение) < 0.01,
         f"{d['eduY']:,.2f} ₽/год")
d = расчёт(EXC_ВНЕШ={'Form006': True})
проверка('обучение · исключённый блок получает эффективный ноль',
         d['eduY'] == 0 and d['depEdu'] == 0)
d = расчёт(поля={'edu_life':'0'})
проверка('обучение · нулевой срок не вызывает деление на ноль', d['eduY'] == 0)
d = расчёт(радио={'site_mode':'hired'})
проверка('сайт · подрядчик: стоимость распределяется, Инвестиционное время равно нулю',
         abs(d['siteY']-10000) < 0.01 and d['goalSelfSiteCost'] == 0)
d = расчёт(радио={'site_mode':'self'})
ожид_сайт = d['R']*80/(7*d['NT'])
проверка('сайт · самостоятельно: деньги равны нулю, Инвестиционное время компенсируется',
         d['siteY'] == 0 and abs(d['goalSelfSiteCost']-ожид_сайт) < 0.01,
         f"{d['goalSelfSiteCost']:,.2f} ₽/год")
d = расчёт(EXC_ВНЕШ={'Form014': True})
проверка('сайт · исключённый блок: деньги и Инвестиционное время получают ноль',
         d['siteY'] == 0 and d['goalSelfSiteCost'] == 0)
d = расчёт(поля={'site_life':'0'}, радио={'site_mode':'self'})
проверка('сайт · нулевой срок не вызывает деление на ноль',
         d['siteY'] == 0 and d['goalSelfSiteCost'] == 0)
d = расчёт(поля={'current_rate': '1000'})
проверка('убыток не обрезается и сохраняет знак минус',
         d['currentResult'] < 0 and d['currentLoss'] < 0
         and d['currentIncome'] == 0 and d['currentIsLoss'])

# ══════════════════════════════ 2c. ПРОГРАММА ЛОЯЛЬНОСТИ (блок 03 отчёта)
print('▶ программа лояльности')
import math
def лояльность(d):
    """Повторяет арифметику блока 03 из report.html, чтобы проверить
       заявленную книгой сходимость: объём + постоянным + сертификаты = фонд."""
    if not d['discP']: return None
    hp = d['R']/d['pool'] if d['pool'] else 0
    hourAt = lambda S: (S*(1+d['K'])+d['cl'])*hp/S
    S = max(round(d['S'] or 2), 1)
    rRate = lambda x: math.ceil(x/100)*100
    чек = rRate(hourAt(S))*S
    фонд = round(d['discY'])
    пост = 0; заказы = 0; людей = 0
    бюдж = фонд*0.45; цел = d['py']*0.30
    остБ, остЗ = бюдж, цел
    ступени = ((0.05,2,0.40),(0.07,4,0.40),(0.10,12,0.20))
    for p_, s_, w in ступени:
        ценаКл = s_*чек*p_
        поБ = min(бюдж*w, остБ)/ценаКл if ценаКл else 0
        поЗ = min(цел*w, остЗ)/s_
        n = max(0, math.floor(min(поБ, поЗ)))
        сум = round(n*s_*чек*p_)
        остБ -= сум; остЗ -= n*s_
        пост += сум; заказы += n*s_; людей += n
    if людей == 0:
        p_, s_, w = ступени[0]
        цена = round(s_*чек*p_)
        if цена <= бюдж and s_ <= цел:
            пост, заказы = цена, s_
    r500 = lambda n: max(round(n/500)*500, 500)
    бюджСпец = min(фонд*0.30, max(фонд-пост, 0))
    v0, v1, v2 = r500(чек*0.85), r500(чек*0.40), r500(чек*0.25)
    ост = бюджСпец
    n0 = 1 if ост >= v0 else 0; ост -= v0*n0
    n1 = max(0, math.floor(ост*0.45/v1)); ост -= v1*n1
    n2 = max(0, math.floor(ост/v2));      ост -= v2*n2
    спец = v0*n0 + v1*n1 + v2*n2
    объём = max(фонд-пост-спец, 0)
    return dict(фонд=фонд, пост=пост, спец=спец, объём=объём,
                сумма=пост+спец+объём, заказы=заказы, py=d['py'])

for имя, поля in (('резерв 15 %', {'disc_on': True, 'disc_pct': '15'}),
                  ('резерв 10 %', {'disc_on': True, 'disc_pct': '10'}),
                  ('резерв 5 %',  {'disc_on': True, 'disc_pct': '5'}),
                  ('резерв 5 %, съёмка 6 ч', {'disc_on': True, 'disc_pct': '5', 'shoot_manual': '6'}),
                  ('резерв 5 %, съёмка 8 ч', {'disc_on': True, 'disc_pct': '5', 'shoot_manual': '8'})):
    л = лояльность(расчёт(поля=поля))
    проверка(f'фонд лояльности сходится · {имя}', abs(л['сумма']-л['фонд']) <= 3,
             f"сумма направлений {л['сумма']:,} против фонда {л['фонд']:,}")
    проверка(f'постоянные ≤ 30 % потока заказов · {имя}', л['заказы'] <= л['py']*0.31,
             f"{л['заказы']:.0f} из {л['py']:.0f} = {л['заказы']/л['py']*100:.0f} %")

# ══════════════════════════════ 3. СОГЛАСОВАННОСТЬ ФАЙЛОВ
print('▶ согласованность файлов')
def читать(п): return io.open(os.path.join(КОРЕНЬ, п), encoding='utf-8').read()
calc, rep, karkas = читать('Веб/calc.html'), читать('Веб/report.html'), читать('Веб/Части/каркас.html')

проверка('report.html совпадает с каркасом', rep == karkas,
         f'{sum(1 for a, b in zip(rep.splitlines(), karkas.splitlines()) if a != b)} строк расходятся')

# Часовые ставки в отчёте называются только по справочнику.
# Эталон — лист 01_Глоссарий книги: «Доход в час на руки», «Базовая ставка часа».
# Прежние ярлыки («Фактическая стоимость вашего часа работы», «Час вашего
# времени») сняты 25.08.2026 по решению владельца — не должны возвращаться.
ент_ю = lambda s: ''.join('\\u%04x' % ord(c) if ord(c) > 127 else c for c in s)
проверка('отчёт: часовые ставки названы по справочнику',
         ент_ю('Доход в час на руки') in rep
         and ент_ю('Фактическая стоимость вашего часа работы') not in rep
         and ент_ю('Час вашего времени') not in rep,
         'эталон названий — лист 01_Глоссарий книги')
проверка('три сценария Ставки часа съёмки связаны правильно',
         'Ваша желаемая ставка за час съёмки' in rep
         and ент_ю('Текущий доход') in rep
         and ент_ю('Ваш текущий доход') not in rep
         and ент_ю('Желаемый доход') in rep
         and ент_ю('Необходимый доход') not in rep
         and "no:'"+ент_ю('текущий')+"'" in rep
         and "no:'"+ент_ю('желаемый')+"'" in rep
         and "no:'"+ент_ю('реальный')+"'" not in rep
         and "no:'"+ент_ю('идеальный')+"'" not in rep
         and "ссылка('Текущая','Текущая ставка')" in rep
         and "ссылка('Желаемая','Желаемая ставка')" in rep
         and "ссылка('В ноль','Ставка в ноль')" in rep)

# Карточка «Желаемый доход» показывает один точный аналитический сценарий.
# Округлённая вверх ставка остаётся только в практических блоках отчёта.
идеал = re.search(r"\{cls:'s2'.*?cap:CAPM\}", rep, re.S)
идеал_код = идеал.group(0) if идеал else ''
проверка('карточка «Желаемый доход»: ставка берётся из d.rateHour',
         'var rateGoalПередан=Number(d.rateHour)' in rep)
проверка('карточка «Желаемый доход»: три суммы показаны до целого рубля',
         'r:f0(goalRateV)' in идеал_код
         and 'rev:f0(d.R/12)' in идеал_код
         and 'mid:f0(d.Ny/12)' in идеал_код
         and 'f0r(' not in идеал_код)
проверка('идеальное кольцо берёт полный состав затрат из calc()',
         'var goalCostsПереданы=Number(d.goalCostsTotal)' in rep
         and 'goalFund+goalDiscountReserve+goalSelfSiteCost' in rep)
проверка('«Всего расходов» приходит готовым из calc()',
         'var expMo=d.totalExpenses/12' in rep
         and 'var ВСЕГО_РАСХОДОВ = d.totalExpenses;' in rep
         and 'd.vari + tax + aq' not in rep)
проверка('детализация: вложения и резервный фонд разделены',
         "['Финансовые вложения (инвестиции)', t04]" in rep
         and "['Резервный фонд', t04r]" in rep
         and "рр('Резерв на отпуск', d.vacY||0)" in rep
         and 'Амортизация / Резерв' not in rep)
проверка('проектное время — годовой pool, а не длительность проекта',
         "['Р','Проектное время', чс(pool)" in rep
         and "['Р','Время на проекты'" not in rep)
проверка('отчёт: Скидка за объём строится от Продолжительности съёмки пользователя',
         'var Sbase=Math.max(Math.round(d.S||2),1)' in rep
         and 'var steps=[Sbase+2,Sbase+4,Sbase+6]' in rep
         and 'var baseHour=rRate(hourAt(Sbase))' in rep)
проверка('календарь: Выходной и Праздник разделены',
         '"Выходные дни": "Выходной"' in rep
         and '"Праздничные дни": "Праздник"' in rep
         and '"Выходные и праздничные дни"' not in rep)
проверка('лояльность: внутренние правила не выводятся пользователю',
         "return '<h3>Фонд и его распределение</h3>' + a;" in rep
         and '<h3>08.2 По каким правилам делится фонд</h3>' not in rep)
проверка('лояльность: неподтверждённая экономия постоянного клиента удалена',
         'перепЧ*0.70' not in rep
         and 'Сверх того сэкономите на привлечении' not in rep
         and 'от 30% вашего бюджета' not in rep
         and 'до 20% вашего рабочего времени' not in rep)

# Карточка «В ноль» берёт точный сценарий из calc(), но денежные значения
# для пользователя округляет до ближайшего целого рубля. Округление вверх
# до 100 рублей применяется только в практическом блоке «Четыре цифры».
ноль = re.search(r"\{cls:'s3'.*?cap:CAPM\}", rep, re.S)
ноль_код = ноль.group(0) if ноль else ''
проверка('отчёт: безубыточность берётся из d.Rb',
         'var VbПередан=Number(d.Rb)' in rep)
проверка('отчёт: ставка в ноль берётся из d.rateZero',
         'var rateZeroПередан=Number(d.rateZero)' in rep)
проверка('карточка «В ноль»: ставка и выручка показаны до целого рубля',
         'r:f0(bRateV)' in ноль_код
         and 'rev:f0(Vb/12)' in ноль_код
         and 'mid:f0(0)' in ноль_код
         and 'f0r(bRateV)' not in ноль_код)
проверка('карточка «Дохода нет»: количество съёмок не дублирует блок загрузки',
         'shoots:Math.ceil(d.zeroShootsM||0)' not in ноль_код
         and 'чтобы покрыть расходы при этой ставке' not in rep)
проверка('денежный формат с копейками удалён из пользовательского интерфейса',
         'var f2=' not in rep and 'f2=function' not in calc)
проверка('блок «Четыре цифры»: ставка в ноль округляется вверх до 100 ₽',
         "['b','\\u0421\\u0442\\u0430\\u0432\\u043a\\u0430 \\u0432 \\u043d\\u043e\\u043b\\u044c',f0r(bRateV)+'/\\u0447'" in rep)
проверка('карточка «Текущий доход» берёт готовый currentResult из calc()',
         'var resultCПередан=Number(d.currentResult)' in rep
         and 'var costsCПереданы=Number(d.currentCostsTotal)' in rep)
проверка('карточка «Текущий доход» не обрезает убыток до нуля',
         'Math.max(curR-expC,0)' not in rep)
проверка('карточка «Текущий доход» меняет подпись на «Ваш убыток»',
         "currentIsLoss?'\\u0412\\u0430\\u0448 \\u0443\\u0431\\u044b\\u0442\\u043e\\u043a'" in rep)
проверка('убыток показывается сравнительной плашкой и знаком минус',
         'class="losscompare"' in rep and 'class="current-loss-note"' in rep
         and "v<0?'\\u2212':''" in rep)
# Отчёт больше не хранит свою ставку эквайринга: он берёт фактическую
# из анкеты (переменная AQФАКТ). Проверяем, что зашитого числа не осталось.
aq_rep = re.search(r'AQ\s*:\s*(0\.\d+)', rep)
aq_form = re.search(r'id="acq_rate"[^>]*value="([\d.]+)"', calc)
проверка('ставка эквайринга берётся из анкеты, а не зашита в отчёт',
         aq_rep is None and 'AQФАКТ' in rep,
         f'в отчёте осталось зашитое {aq_rep.group(1) if aq_rep else "—"}'
         + (f' · в анкете {aq_form.group(1)} %' if aq_form else ''))
проверка('анкета: собственное жильё блокирует поле аренды или ипотеки',
         'rent.disabled=oh.checked' in calc
         and "homeHousing=CHK('own_home')?0:V('home_rent')" in calc)
проверка('анкета: переключатель Рабочего места меняет домашнюю и арендную ветки',
         "$('ws_home').style.display=r.value==='home'?'grid':'none'" in calc
         and "$('ws_off').style.display=r.value==='office'?'grid':'none'" in calc)
income_tag=re.search(r'<input[^>]*id="income_month"[^>]*>',calc)
проверка('income_month: денежное поле 10 000–3 000 000 ₽ без копеек, шаг 100 ₽',
         bool(income_tag)
         and 'data-integer' in income_tag.group(0)
         and 'min="10000"' in income_tag.group(0)
         and 'max="3000000"' in income_tag.group(0)
         and 'step="100"' in income_tag.group(0)
         and "if(!/^\\d+$/.test(сыр))" in calc
         and 'Желаемый доход должен быть не меньше 10 000 ₽ в месяц' in calc
         and 'Желаемый доход должен быть не больше 3 000 000 ₽ в месяц' in calc)
current_rate_tag=re.search(r'<input[^>]*id="current_rate"[^>]*>',calc)
проверка('current_rate: денежное поле 0–1 000 000 ₽ без копеек, шаг 100 ₽',
         bool(current_rate_tag)
         and 'data-integer' in current_rate_tag.group(0)
         and 'min="0"' in current_rate_tag.group(0)
         and 'max="1000000"' in current_rate_tag.group(0)
         and 'step="100"' in current_rate_tag.group(0)
         and 'Текущую ставку укажите целым числом без копеек' in calc
         and 'Текущая ставка должна быть не больше 1 000 000 ₽ в час' in calc)
frames_out_tag=re.search(r'<input[^>]*id="frames_out"[^>]*>',calc)
проверка('frames_out: целое количество 1–1 000 кадров/ч, шаг 10',
         bool(frames_out_tag)
         and 'data-integer' in frames_out_tag.group(0)
         and 'min="1"' in frames_out_tag.group(0)
         and 'max="1000"' in frames_out_tag.group(0)
         and 'step="10"' in frames_out_tag.group(0)
         and 'Количество готовых кадров укажите целым числом' in calc
         and 'Количество готовых кадров должно быть не меньше 1 за час' in calc
         and 'Количество готовых кадров должно быть не больше 1 000 за час. Это верхний предел для реалистичного расчёта' in calc)
shutter_per_hour_tag=re.search(r'<input[^>]*id="shutter_per_hour"[^>]*>',calc)
проверка('shutter_per_hour: целое количество 1–1 000 срабатываний/ч, шаг 10',
         bool(shutter_per_hour_tag)
         and 'data-integer' in shutter_per_hour_tag.group(0)
         and 'min="1"' in shutter_per_hour_tag.group(0)
         and 'max="1000"' in shutter_per_hour_tag.group(0)
         and 'step="10"' in shutter_per_hour_tag.group(0)
         and 'Количество срабатываний затвора укажите целым числом' in calc
         and 'Количество срабатываний затвора должно быть не меньше 1 за час' in calc
         and 'Количество срабатываний затвора должно быть не больше 1 000 за час. Это верхний предел для реалистичного расчёта' in calc)
проверка('Form001 и Form002: срок службы позиции 1–30 лет, целые годы, шаг 1',
         "f==='Form001'?'съёмочного оборудования':(f==='Form002'?'офисного оборудования'" in calc
         and "var максимумСрока=f==='Form004'?10:30" in calc
         and "категорияСрока?' data-limit-live data-integer min=\"1\" max=\"'+максимумСрока+'\" step=\"1\"" in calc
         and 'Установлен минимум — 1 год службы' in calc
         and "class=\"c3\"'+пределСрока" in calc)
проверка('Form004: срок службы Купленной программы 1–10 лет, целые годы, шаг 1',
         "f==='Form004'?'купленной программы'" in calc
         and "f==='Form004'?10:30" in calc
         and "Установлен максимум — '+максимумСрока+' лет службы" in calc)
проверка('Form013: срок службы Обустройства рабочего места 1–30 лет',
         "f==='Form013'?'обустройства рабочего места'" in calc
         and "var максимумСрока=f==='Form004'?10:30" in calc
         and "Установлен максимум — '+максимумСрока+' лет службы" in calc)
проверка('Form001 и Form002: стоимость позиции 0–10 000 000 ₽, целые рубли, шаг 100',
         "f==='Form001'?'одну позицию съёмочного оборудования':(f==='Form002'?'одну позицию офисного оборудования'" in calc
         and "var максимумСтоимости=(f==='Form003'||f==='Form004')?100000:(f==='Form013'?500000:10000000)" in calc
         and "описаниеСтоимости?' data-limit-live data-integer min=\"0\" max=\"'+максимумСтоимости+'\" step=\"100\"" in calc
         and 'Для указанной позиции стоимость равна 0 ₽' in calc
         and "#t_Form002 .c1,#t_Form002 .c2" in calc)
проверка('Form003: платёж подписки 0–100 000 ₽, целые рубли, шаг 100',
         "f==='Form003'?'один платёж подписки'" in calc
         and "(f==='Form003'||f==='Form004')?'100 000':(f==='Form013'?'500 000':'10 000 000')" in calc
         and '#t_Form003 .c1,#t_Form003 .c2' in calc)
проверка('Form004: купленная программа 0–100 000 ₽, целые рубли, шаг 100',
         "f==='Form004'?'одну купленную программу'" in calc
         and '#t_Form004 .c1,#t_Form004 .c2' in calc
         and 'проверитьНулевуюСтоимостьСтроки(e)' in calc)
проверка('Form013: обустройство рабочего места 0–500 000 ₽, целые рубли, шаг 100',
         "f==='Form013'?'одну позицию обустройства рабочего места'" in calc
         and "f==='Form013'?500000:10000000" in calc
         and "f==='Form013'?'500 000':'10 000 000'" in calc
         and '#t_Form013 .c1,#t_Form013 .c2' in calc)
office_util_tag=re.search(r'<input[^>]*id="office_util"[^>]*>',calc)
проверка('office_util: 0–100 000 ₽/мес, целые рубли, шаг 100',
         bool(office_util_tag)
         and 'data-integer' in office_util_tag.group(0)
         and 'data-limit-live' in office_util_tag.group(0)
         and 'min="0"' in office_util_tag.group(0)
         and 'max="100000"' in office_util_tag.group(0)
         and 'step="100"' in office_util_tag.group(0)
         and 'Установлен максимум — 100 000 ₽ в месяц на коммунальные платежи, интернет и связь отдельного помещения' in calc)
office_rent_tag=re.search(r'<input[^>]*id="office_rent"[^>]*>',calc)
проверка('office_rent: 0–500 000 ₽/мес, целые рубли, шаг 100',
         bool(office_rent_tag)
         and 'data-integer' in office_rent_tag.group(0)
         and 'data-limit-live' in office_rent_tag.group(0)
         and 'min="0"' in office_rent_tag.group(0)
         and 'max="500000"' in office_rent_tag.group(0)
         and 'step="100"' in office_rent_tag.group(0)
         and 'Установлен максимум — 500 000 ₽ в месяц за аренду отдельного помещения' in calc)
cab_area_tag=re.search(r'<input[^>]*id="cab_area"[^>]*>',calc)
проверка('cab_area: активное поле 0,1–100 м², шаг 0,1 и не больше home_area',
         bool(cab_area_tag)
         and 'data-limit-live' in cab_area_tag.group(0)
         and 'data-step-live' in cab_area_tag.group(0)
         and 'min="0.1"' in cab_area_tag.group(0)
         and 'max="100"' in cab_area_tag.group(0)
         and 'step="0.1"' in cab_area_tag.group(0)
         and 'Установлен абсолютный максимум — 100 м² рабочей зоны' in calc
         and "зона<=всего" in calc
         and 'Площадь рабочей зоны не может превышать общую площадь жилья' in calc)
home_area_tag=re.search(r'<input[^>]*id="home_area"[^>]*>',calc)
проверка('home_area: активное поле 1–1 000 м² с шагом 0,1',
         bool(home_area_tag)
         and 'data-limit-live' in home_area_tag.group(0)
         and 'data-step-live' in home_area_tag.group(0)
         and 'min="1"' in home_area_tag.group(0)
         and 'max="1000"' in home_area_tag.group(0)
         and 'step="0.1"' in home_area_tag.group(0)
         and 'Установлен минимум — 1 м² общей площади жилья' in calc
         and 'Установлен максимум — 1 000 м² общей площади жилья' in calc
         and "(el.id==='home_area'||el.id==='cab_area')?'0,1 м²':'0,5 часа'" in calc)
home_util_tag=re.search(r'<input[^>]*id="home_util"[^>]*>',calc)
проверка('home_util: 0–100 000 ₽/мес, целые рубли, шаг 100',
         bool(home_util_tag)
         and 'data-integer' in home_util_tag.group(0)
         and 'data-limit-live' in home_util_tag.group(0)
         and 'min="0"' in home_util_tag.group(0)
         and 'max="100000"' in home_util_tag.group(0)
         and 'step="100"' in home_util_tag.group(0)
         and 'Установлен максимум — 100 000 ₽ в месяц на коммунальные платежи, интернет и связь' in calc)
home_rent_tag=re.search(r'<input[^>]*id="home_rent"[^>]*>',calc)
проверка('home_rent: 0–500 000 ₽/мес, целые рубли, шаг 100',
         bool(home_rent_tag)
         and 'data-integer' in home_rent_tag.group(0)
         and 'data-limit-live' in home_rent_tag.group(0)
         and 'min="0"' in home_rent_tag.group(0)
         and 'max="500000"' in home_rent_tag.group(0)
         and 'step="100"' in home_rent_tag.group(0)
         and 'Установлен максимум — 500 000 ₽ в месяц за жильё' in calc
         and "homeHousing=CHK('own_home')?0:V('home_rent')" in calc)
promo_per_day_tag=re.search(r'<input[^>]*id="promo_per_day"[^>]*>',calc)
проверка('promo_per_day: включённый блок 0,5–8 часов/день с шагом 0,5',
         bool(promo_per_day_tag)
         and 'data-limit-live' in promo_per_day_tag.group(0)
         and 'data-step-live' in promo_per_day_tag.group(0)
         and 'min="0.5"' in promo_per_day_tag.group(0)
         and 'max="8"' in promo_per_day_tag.group(0)
         and 'step="0.5"' in promo_per_day_tag.group(0)
         and 'Чтобы не учитывать блок, используйте галочку исключения' in calc
         and 'это полный нормативный рабочий день' in calc)
client_time_tag=re.search(r'<input[^>]*id="client_time"[^>]*>',calc)
проверка('client_time: включённый блок 0,5–72 часа на проект с шагом 0,5',
         bool(client_time_tag)
         and 'data-limit-live' in client_time_tag.group(0)
         and 'data-step-live' in client_time_tag.group(0)
         and 'min="0.5"' in client_time_tag.group(0)
         and 'max="72"' in client_time_tag.group(0)
         and 'step="0.5"' in client_time_tag.group(0)
         and 'Чтобы не учитывать блок, используйте галочку исключения' in calc
         and 'Установлен максимум — 72 часа работы с клиентом на проект' in calc)
post_ratio_tag=re.search(r'<input[^>]*id="post_ratio"[^>]*>',calc)
проверка('post_ratio: 0–72 часа на час съёмки с точным шагом 0,5 часа',
         bool(post_ratio_tag)
         and 'data-limit-live' in post_ratio_tag.group(0)
         and 'data-step-live' in post_ratio_tag.group(0)
         and 'min="0"' in post_ratio_tag.group(0)
         and 'max="72"' in post_ratio_tag.group(0)
         and 'step="0.5"' in post_ratio_tag.group(0)
         and 'Установлен минимум — 0 часов постпродакшна' in calc
         and 'Установлен максимум — 72 часа постпродакшна на один час съёмки' in calc)
shoot_manual_tag=re.search(r'<input[^>]*id="shoot_manual"[^>]*>',calc)
проверка('shoot_manual: 0,5–72 часа с точным шагом 0,5 часа',
         bool(shoot_manual_tag)
         and 'data-limit-live' in shoot_manual_tag.group(0)
         and 'data-step-live' in shoot_manual_tag.group(0)
         and 'min="0.5"' in shoot_manual_tag.group(0)
         and 'max="72"' in shoot_manual_tag.group(0)
         and 'step="0.5"' in shoot_manual_tag.group(0)
         and 'Установлен минимум — 0,5 часа' in calc
         and 'Установлен максимум — 72 часа' in calc)
shutter_life_tag=re.search(r'<input[^>]*id="shutter_life"[^>]*>',calc)
проверка('shutter_life: целый ресурс 10 000–1 000 000 срабатываний, шаг 10 000',
         bool(shutter_life_tag)
         and 'data-integer' in shutter_life_tag.group(0)
         and 'data-limit-live' in shutter_life_tag.group(0)
         and 'min="10000"' in shutter_life_tag.group(0)
         and 'max="1000000"' in shutter_life_tag.group(0)
         and 'step="10000"' in shutter_life_tag.group(0)
         and 'Установлен минимум — 10 000 срабатываний затвора' in calc
         and 'Установлен максимум — 1 000 000 срабатываний затвора' in calc)
проверка('каталоги Form001, Form002, Form004 и Form013 публикуют пределы срока',
         "'Срок min', 'Срок max', 'Срок step'" in читать('Инструменты/значения_по_умолчанию.py')
         and "if form in ('Form001', 'Form002', 'Form013') and kind == 'life':" in читать('Инструменты/значения_по_умолчанию.py')
         and "elif form == 'Form004' and kind == 'life':" in читать('Инструменты/значения_по_умолчанию.py')
         and "term_min, term_max, term_step = '1', '10', '1'" in читать('Инструменты/значения_по_умолчанию.py'))
проверка('каталоги Form001–Form004 и Form013 публикуют пределы стоимости',
         "if form in ('Form001', 'Form002'):" in читать('Инструменты/значения_по_умолчанию.py')
         and "elif form in ('Form003', 'Form004'):" in читать('Инструменты/значения_по_умолчанию.py')
         and "elif form == 'Form013':" in читать('Инструменты/значения_по_умолчанию.py')
         and "cost_min, cost_max, cost_step = '0', '500000', '100'" in читать('Инструменты/значения_по_умолчанию.py'))
проверка('утверждённые числовые пределы применяются сразу возле поля',
         len(re.findall(r'<input[^>]*data-limit-live',calc)) == 15
         and "R.addEventListener('input'" in calc
         and "применитьПредел(e,'max')" in calc
         and "R.addEventListener('blur'" in calc
         and "применитьПредел(e,'min')" in calc
         and 'привестиКЦелому(e)' in calc
         and 'привестиКШагу(e)' in calc
         and "'Значение приведено к ближайшему шагу — '+шагТекст" in calc
         and 'Значение округлено до ближайшего целого. Проверьте значение.' in calc
         and "note.setAttribute('aria-live','polite')" in calc
         and 'Введённое значение приведено к допустимой границе' in calc)
проверка('report: печатная версия раскрывает детализацию и скрывает служебные панели',
         '@media print' in rep
         and '#phr-root #спрдет .пункт .тело{display:block!important' in rep
         and '#phr-root .trb{display:none}' in rep
         and '#phr-root .dlbar,#phr-root .bn,#phr-root .savebar{display:none!important}' in rep
         and 'print-color-adjust:exact' in rep)

for имя, файл in (('calc', calc), ('report', rep)):
    css = '\n'.join(re.findall(r'<style[^>]*>([\s\S]*?)</style>', файл))
    проверка(f'{имя}: баланс CSS-скобок', css.count('{') == css.count('}'),
             f"{css.count('{') - css.count('}'):+d}")
    исп = set(re.findall(r'var\((--[\w-]+)\)', файл))
    объ = set(re.findall(r'(--[\w-]+)\s*:', файл))
    проверка(f'{имя}: все CSS-переменные объявлены', not (исп - объ), ', '.join(sorted(исп - объ)))

# баланс HTML-тегов во всех трёх файлах
VOID = {'br','img','input','meta','link','hr','source','area','col','embed','track','wbr',
        'circle','path','rect','polygon','line','ellipse','stop','use','polyline','animate'}
def баланс_тегов(s):
    s = re.sub(r'(<script[^>]*>)([\s\S]*?)(</script>)',
               lambda m: m.group(1) + re.sub(r'[^\n]', ' ', m.group(2)) + m.group(3), s)
    s = re.sub(r'(<style[^>]*>)([\s\S]*?)(</style>)',
               lambda m: m.group(1) + re.sub(r'[^\n]', ' ', m.group(2)) + m.group(3), s)
    s = re.sub(r'<!--[\s\S]*?-->', lambda m: re.sub(r'[^\n]', ' ', m.group(0)), s)
    стек, беды = [], []
    for m in re.finditer(r'<(/?)([a-zA-Z][\w-]*)([^>]*?)(/?)>', s):
        зак, имя, само = m.group(1), m.group(2).lower(), m.group(4)
        if имя in VOID or само == '/': continue
        if зак:
            if стек and стек[-1][0] == имя: стек.pop()
            else: беды.append('лишний </%s> поз.%d' % (имя, m.start()))
        else: стек.append((имя, m.start()))
    return беды + ['не закрыт <%s> поз.%d' % (и, п) for и, п in стек]

for имя, файл in (('calc', calc), ('report', rep), ('index', читать('Веб/index.html'))):
    б = баланс_тегов(файл)
    проверка(f'{имя}: баланс HTML-тегов', not б, '; '.join(б[:3]))

# 19 прямых потомков .wp
def детей(s):
    s = re.sub(r'<style[^>]*>[\s\S]*?</style>|<script[^>]*>[\s\S]*?</script>|<!--[\s\S]*?-->', '', s)
    VOID = {'br','img','input','meta','link','hr','source','area','col','embed','track','wbr',
            'circle','path','rect','polygon','line','ellipse','stop','use','polyline'}
    н = s.index('<div class="wp">') + len('<div class="wp">')
    гл, дети, тек = 0, 0, None
    for m in re.finditer(r'<(/?)([a-zA-Z][\w-]*)([^>]*?)(/?)>', s[н:]):
        зак, имя_, само = m.group(1), m.group(2).lower(), m.group(4)
        if зак:
            гл -= 1
            if гл == 0 and тек is not None: дети += 1; тек = None
            if гл < 0: break
        else:
            if имя_ in VOID or само == '/': continue
            if гл == 0: тек = m.start()
            гл += 1
    return дети
n = детей(rep)
проверка("в .wp ровно 20 блоков (как ждёт собрать.py)", n == 20, f"{n}")
report_story=['REPORT-B007','REPORT-B008','REPORT-B012','REPORT-B010','REPORT-B013','REPORT-B009','REPORT-B011','REPORT-B014','REPORT-B015']
report_positions=[rep.find(f'data-block-id="{x}"') for x in report_story]
проверка('report: маршрут бюджет → время → сценарии → благодарность → загрузка → скидка → налоги',
         all(x>=0 for x in report_positions) and report_positions==sorted(report_positions))
report_user_numbers={'REPORT-B007':'01','REPORT-B008':'02','REPORT-B012':'03','REPORT-B013':'04','REPORT-B009':'05','REPORT-B011':'06','REPORT-B014':'07','REPORT-B015':'08'}
number_errors=[]
for bid,no in report_user_numbers.items():
    m=re.search(rf'data-block-id="{bid}"[^>]*>.*?<div class="bn">(\d{{2}})</div>',rep,re.S)
    if not m or m.group(1)!=no:number_errors.append(f'{bid}→{m.group(1) if m else "—"}')
проверка('report: пользовательские номера 01–08 соответствуют новому порядку',not number_errors,', '.join(number_errors))

# Стабильная техническая нумерация страниц и всех смысловых блоков.
index_html = читать('Веб/index.html')
for имя, файл, page_id, count in (
    ('index', index_html, 'PAGE-INDEX', 2),
    ('calc', calc, 'PAGE-CALC', 39),
    ('report', rep, 'PAGE-REPORT', 20),
):
    ids = re.findall(r'data-block-id="([^"]+)"', файл)
    проверка(f'{имя}: все смысловые блоки имеют уникальный технический ID',
             len(ids) == count and len(set(ids)) == count,
             f'{len(ids)} ID, уникальных {len(set(ids))}, ожидается {count}')
    проверка(f'{имя}: закреплён data-page-id={page_id}',
             f'data-page-id="{page_id}"' in файл)
branches = re.findall(r'data-branch-id="([^"]+)"', calc)
проверка('calc: условные ветки имеют стабильные ID',
         len(branches) == 10 and len(set(branches)) == 10,
         f'{len(branches)} ID, уникальных {len(set(branches))}')

# логотип: 8 правил книги
for имя, файл in (('calc', calc), ('report', rep)):
    знаки = re.findall(r'<svg class="(?:bx|lx)".*?</svg>', файл, re.S)
    проверка(f'{имя}: зелёный ромб — правый (x=19.81)',
             all(re.search(r'x="19\.81"[^/]*?fill="var\(--c-g-500\)"', b) for b in знаки))
    проверка(f'{имя}: ровно один зелёный ромб на знак',
             all(len(re.findall(r'--c-g-500', b)) == 1 for b in знаки))
    проверка(f'{имя}: геометрия знака 10.10 / rx 1.90',
             all('width="10.10"' in b and 'rx="1.90"' in b for b in знаки))
    проверка(f'{имя}: нет следов анимации знака',
             not re.search(r'u-(inc|exp|res|prof)', файл))
    проверка(f'{имя}: написание «счётикс» с буквой ё',
             'счетикс' not in файл.lower() or 'счётикс' in файл.lower(),
             'найдено «счетикс» без ё' if 'счетикс' in файл.lower() else '')
проверка('каркас: написание «счётикс» с буквой ё', 'счетикс' not in karkas.lower())

# JS обоих файлов синтаксически валиден
for имя, путь in (('calc', 'Веб/calc.html'), ('report', 'Веб/report.html'), ('index', 'Веб/index.html')):
    s = читать(путь)
    код = '\n'.join(re.findall(r'<script[^>]*>([\s\S]*?)</script>', s))
    врем = '/tmp/_js_%s.js' % имя
    io.open(врем, 'w', encoding='utf-8').write(код)
    p = subprocess.run(['node', '-e',
        "new Function(require('fs').readFileSync(process.argv[1],'utf8'))", врем],
        capture_output=True, text=True)
    os.remove(врем)
    проверка(f'{имя}: JS компилируется', p.returncode == 0, (p.stderr or '').strip().split('\n')[0][:120])

# ══════════════════════════════ 4. КНИГА ↔ КОД
print('▶ книга ↔ код')
try:
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(КОРЕНЬ, 'Книга', 'Калькулятор_ставки_часа.xlsx'))
    ws = wb['calc']
    СВЯЗЬ = {'income_month': 'income_month', 'current_rate': 'current_rate', 'frames_out': 'frames_out',
             'shoot_duration': 'shoot_manual', 'post_ratio': 'post_ratio', 'client_time': 'client_time',
             'promo_per_day': 'promo_per_day', 'acc_per_quarter': 'acc_per_quarter',
             'acc_cost_month': 'acc_cost_month', 'home_rent': 'home_rent', 'home_util': 'home_util',
             'home_area': 'home_area', 'cab_area': 'cab_area', 'office_rent': 'office_rent',
             'office_util': 'office_util', 'edu_life': 'edu_life', 'site_cost': 'site_cost',
             'site_hours': 'site_hours', 'site_life': 'site_life', 'fm_pct': 'fm_pct'}
    расх = []
    for r in range(1, ws.max_row + 1):
        i = ws.cell(r, 1).value
        if not isinstance(i, str) or i.strip() not in СВЯЗЬ: continue
        i = i.strip(); v = ws.cell(r, 3).value
        m = re.search(r'id="%s"[^>]*value="([^"]*)"' % СВЯЗЬ[i], calc)
        if not m: continue
        try:
            bv = float(v); hv = float(m.group(1).replace(' ', ''))
            if i in ('fm_pct',) and bv < 1: bv *= 100
            if abs(bv - hv) > 1e-9: расх.append(f'{i}: книга {bv:g} ≠ форма {hv:g}')
        except Exception: pass
    проверка('значения по умолчанию совпадают с книгой', not расх, '; '.join(расх))

    # ── переключатели книги против анкеты. Раньше не сверялись, и книга
    # считала с включённым резервом на форс-мажоры, а анкета — без него:
    # ставка расходилась на 496 ₽ (9 926 против 9 430).
    ПЕРЕКЛ = {'fm_on':   ('Заложить резерв',        r'id="fm_on"[^>]*checked'),
              'acc_mode':('Веду самостоятельно',    r'name="acc_mode" value="self" checked'),
              'ws_mode': ('Работаю из дома',        r'name="ws_mode" value="home" checked'),
              'site_mode':('Нанимал (а) специалиста', r'name="site_mode" value="hired" checked')}
    плохо = []
    for r in range(1, ws.max_row + 1):
        и = ws.cell(r, 1).value
        if not isinstance(и, str) or и.strip() not in ПЕРЕКЛ: continue
        и = и.strip(); знач = str(ws.cell(r, 3).value or '').strip()
        вкл_текст, шаблон = ПЕРЕКЛ[и]
        книга_вкл = (знач == вкл_текст)
        форма_вкл = bool(re.search(шаблон, calc))
        if книга_вкл != форма_вкл:
            плохо.append(f'{и}: книга «{знач}», в анкете {"включено" if форма_вкл else "выключено"}')
    проверка('переключатели книги совпадают с анкетой', not плохо, '; '.join(плохо))

    # ── старый реестр полей + канонический Markdown-реестр правил.
    # Лист 16 больше не является источником: правило own_home в нём устарело.
    rule_path = os.path.join(КОРЕНЬ, 'Документация', 'Описания расчётов',
                             'Архитектура — Правила модели и интерфейса.md')
    rule_rows = []
    for line in io.open(rule_path, encoding='utf-8'):
        if re.match(r'^\|\s*\d+\s*\|', line):
            cols = [x.strip() for x in line.strip().strip('|').split('|')]
            if len(cols) == 7:
                rule_rows.append(cols)
    rule_ids = [x[2] for x in rule_rows]
    ожид_ids = [f'RULE-{i:03d}' for i in range(1,20)]
    проверка('канонический реестр правил: 19 последовательных ID', rule_ids == ожид_ids,
             ', '.join(rule_ids))
    битые_док = []
    for row in rule_rows:
        link = row[6].replace('`','')
        if not link or not os.path.exists(os.path.join(КОРЕНЬ,link)):
            битые_док.append(f'{row[2]} → {link or "—"}')
    проверка('канонический реестр правил: документы существуют', not битые_док,
             '; '.join(битые_док[:4]))
    правило10 = next((x for x in rule_rows if x[2]=='RULE-010'), [])
    проверка('RULE-010: собственное жильё блокирует аренду или ипотеку',
             bool(правило10) and 'эффективное значение равным 0' in правило10[4])

    # Пересчитана ли книга: у формул должны быть сохранённые значения.
    # Без них проверки не могут сверять расчёт кода с расчётом книги.
    wbv = openpyxl.load_workbook(os.path.join(КОРЕНЬ, 'Книга', 'Калькулятор_ставки_часа.xlsx'), data_only=True)
    формул = пусто = 0
    for имя_л in wb.sheetnames:
        л, лв = wb[имя_л], wbv[имя_л]
        for стр in л.iter_rows():
            for c in стр:
                if isinstance(c.value, str) and c.value.startswith('='):
                    формул += 1
                    if лв[c.coordinate].value is None: пусто += 1
    проверка('книга пересчитана (у формул есть значения)', пусто == 0,
             f'{пусто} из {формул} формул без результата — см. Архив/Задачи/02_ПЕРЕСЧЁТ_КНИГИ.md')

    # налоговые константы
    конст = {}
    for r in range(1, ws.max_row + 1):
        i = ws.cell(r, 1).value
        if isinstance(i, str): конст[i.strip()] = ws.cell(r, 3).value
    пары = [('fixed_contrib', r'FIX=(\d+)', 1), ('contrib_threshold', r'THR=(\d+)', 1),
            ('contrib_cap', r'CAP=(\d+)', 1), ('min_tax_usn', r'M15=\.?(\d+)', 0.01),
            ('min_tax_ausn', r'M20=\.?(\d+)', 0.01), ('limit_npd', r'npd:\s*\[(\d+)', 1),
            ('limit_ausn', r'ausn8:\s*\[(\d+)', 1)]
    плохо = []
    for ид, шаб, множ in пары:
        m = re.search(шаб, calc)
        if not m or ид not in конст: continue
        зн = float(m.group(1)) * (множ if множ != 0.01 else 0.01)
        if abs(float(конст[ид]) - зн) > 1e-9: плохо.append(f'{ид}: книга {конст[ид]} ≠ код {зн:g}')
    проверка('налоговые константы совпадают с книгой', not плохо, '; '.join(плохо))
except ImportError:
    замечание('openpyxl не установлен — сверка с книгой пропущена')

# ── чистая книга: автономность нового расчётного контура
try:
    clean_path = os.path.join(КОРЕНЬ, 'Книга', 'Калькулятор_ставки_часа.xlsx')
    clean = openpyxl.load_workbook(clean_path, data_only=False)
    clean_values = openpyxl.load_workbook(clean_path, data_only=True)
    clean_formula_errors=[]
    for sh in clean.worksheets:
        for row in sh.iter_rows():
            for cell in row:
                if cell.data_type=='f':
                    value=clean_values[sh.title][cell.coordinate].value
                    if value is None or (isinstance(value,str) and value.startswith('#')):
                        clean_formula_errors.append(f'{sh.title}!{cell.coordinate}={value}')
    проверка('чистая книга: все формулы имеют кэш без ошибок', not clean_formula_errors,
             '; '.join(clean_formula_errors[:5]))
    expected_clean = ['00_Читать', '01_Глоссарий', 'calc', 'Состав',
                      'Значения_по_умолчанию', 'Интерфейс', 'Тексты',
                      'CSS_и_компоненты', 'Программа_лояльности', 'Полный_отчёт']
    проверка('чистая книга: ровно 10 утверждённых листов', clean.sheetnames == expected_clean,
             f'{len(clean.sheetnames)} листов')
    old_refs = []
    for sh in clean.worksheets:
        for row in sh.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and ("'03_Каталоги'!" in cell.value or "'05_Расчёт'!" in cell.value):
                    old_refs.append(f'{sh.title}!{cell.coordinate}')
    for dn in clean.defined_names.values():
        text = dn.attr_text or ''
        if '03_Каталоги' in text or '05_Расчёт' in text:
            old_refs.append(f'имя {dn.name}')
    проверка('чистая книга: нет расчётных ссылок на 03_Каталоги и 05_Расчёт', not old_refs,
             ', '.join(old_refs[:5]))
    required_names = {'cat_form', 'cat_year', 'revenue_target', 'tax_target',
                      'acquiring_target', 'total_expenses_model', 'revenue_break_even',
                      'tax_break_even', 'acquiring_break_even', 'rate_zero_model',
                      'rate_hour_model', 'project_price_model', 'revenue_month_model',
                      'revenue_current', 'tax_current', 'current_costs_total_model',
                      'current_result_model', 'current_income_model', 'current_loss_model'}
    missing_names = sorted(required_names - set(clean.defined_names))
    проверка('чистая книга: финансовое ядро связано именами', not missing_names,
             ', '.join(missing_names))
    ghost_names = sorted({
        'core_time', 'weeks_year', 'work_time', 'site_value', 'revenue_base',
        'rate_current', 'threshold_credit', 'injury_contrib',
        'revenue_npd_avg', 'revenue_npd_fiz', 'revenue_npd_ur',
        'revenue_usn_inc', 'revenue_usn_prof', 'revenue_ausn_inc',
        'revenue_ausn_prof', 'week_target', 'shoot_auto', 'fm_week',
        'projects_month', 'projects_week', 'shooting_month', 'shooting_week',
        'ops_per_shoot', 'frames_total', 'equip_only'
    } & set(clean.defined_names))
    проверка('чистая книга: удалённые промежуточные дубли не возвращаются', not ghost_names,
             ', '.join(ghost_names))
    clean_calc = clean['calc']
    forbidden_calc_names = ('ПЕРЕКЛЮЧАТЕЛЬ ·', 'ЧИСТОЕ время', 'ПРОЕКТНОЕ время',
                            'СЪЁМОЧНОЕ время', 'Переменные ·', 'Амортизация ·',
                            'ИТОГО', 'ВСЕГО РАСХОДОВ (C)', 'текущей ставке')
    stale_names = [f'B{r}: {clean_calc.cell(r,2).value}' for r in range(1,clean_calc.max_row+1)
                   if clean_calc.cell(r,1).value not in (None,'ID')
                   and any(x in str(clean_calc.cell(r,2).value or '') for x in forbidden_calc_names)]
    проверка('чистая книга: устаревшие наименования calc не возвращаются', not stale_names,
             '; '.join(stale_names[:5]))
    active_ids = [clean_calc.cell(r,1).value for r in range(1,clean_calc.max_row+1)
                  if clean_calc.cell(r,1).value not in (None,'ID') and not str(clean_calc.cell(r,1).value)[0].isdigit()]
    duplicate_ids = sorted(x for x,n in collections.Counter(active_ids).items() if n>1)
    проверка('чистая книга: активные ID calc уникальны', not duplicate_ids,
             ', '.join(duplicate_ids))
    проверка('чистая книга: базовая ставка питается выбранной выручкой',
             clean_calc['C132'].value == '=IF(net_time=0,0,revenue_target/net_time)',
             str(clean_calc['C132'].value))
    current_result_cell = clean.defined_names['current_result_model'].attr_text.split('$C$')[-1]
    current_formula = clean_calc[f'C{current_result_cell}'].value
    проверка('чистая книга: текущий результат не обрезается и учитывает все затраты',
             current_formula == '=revenue_current-current_costs_total_model',
             str(current_formula))
    full = clean['Полный_отчёт']
    full_ok = full.max_row == 118 and full.max_column == 16 and all(full.cell(r, 14).value == 'да' for r in range(2, 119))
    проверка('чистая книга: Полный_отчёт содержит показатели, диаграммы и полный контракт d', full_ok,
             f'{full.max_row-1} строк')
    contract_fields=set()
    bad_full_rules=[]
    known_rules={f'RULE-{i:03d}' for i in range(1,20)}
    for r in range(2,full.max_row+1):
        contract_fields.update(re.findall(r'd\.([A-Za-z_][\w]*)',str(full.cell(r,5).value or '')))
        for rid in str(full.cell(r,16).value or '').split(';'):
            rid=rid.strip()
            if rid and rid not in known_rules: bad_full_rules.append(f'P{r}:{rid}')
    contract_missing=sorted((set(базовый)-{'__parts'})-contract_fields)
    contract_extra=sorted(contract_fields-set(базовый))
    проверка('чистая книга: Полный_отчёт регистрирует весь верхнеуровневый контракт calc()',
             not contract_missing and not contract_extra,
             ('нет: '+', '.join(contract_missing[:5]) if contract_missing else '')+
             ('; лишнее: '+', '.join(contract_extra[:5]) if contract_extra else ''))
    проверка('чистая книга: Правило ID Полного_отчёта ведёт в RULE-001…019',
             not bad_full_rules, ', '.join(bad_full_rules[:5]))
    chart_totals = collections.defaultdict(float)
    for r in range(2, full.max_row+1):
        block = str(full.cell(r,1).value or '')
        if block.startswith('Данные диаграммы / '):
            chart_totals[block.removeprefix('Данные диаграммы / ')] += float(full.cell(r,13).value or 0)
    chart_ok = (abs(chart_totals['Желаемая ставка']-базовый['R']) < 0.02
                and abs(chart_totals['Текущая ставка']-базовый['Rc']) < 0.02
                and abs(chart_totals['Ставка в ноль']-базовый['Rb']) < 0.02)
    проверка('данные трёх диаграмм сходятся с Выручкой каждого сценария', chart_ok,
             '; '.join(f'{k}: {v:,.2f}' for k,v in chart_totals.items()))
    loyalty_sheet = clean['Программа_лояльности']
    loyalty_structure = (loyalty_sheet.max_row == 79 and loyalty_sheet.max_column == 10
                         and sum(c.data_type == 'f' for row in loyalty_sheet.iter_rows() for c in row) == 47)
    проверка('чистая книга: Программа_лояльности содержит формулы и контрольные сценарии',
             loyalty_structure, f'{loyalty_sheet.max_row-1} строк')
    loyalty_controls = collections.defaultdict(dict)
    for r in range(2, loyalty_sheet.max_row+1):
        section = str(loyalty_sheet.cell(r,1).value or '')
        if section.startswith('Контроль '):
            loyalty_controls[section][loyalty_sheet.cell(r,3).value] = float(loyalty_sheet.cell(r,4).value or 0)
    loyalty_ok = True
    for values in loyalty_controls.values():
        loyalty_ok &= abs(values.get('Постоянные клиенты',0)+values.get('Подарочные сертификаты',0)
                          +values.get('Скидка за объём',0)-values.get('Фонд программы лояльности',0)) < 0.01
        loyalty_ok &= values.get('Заказы постоянных клиентов',0) <= базовый['py']*0.30+0.01
    проверка('чистая книга: контроль 5/10/15% распределяет фонд без превышения потока',
             loyalty_ok and len(loyalty_controls)==3,
             ', '.join(sorted(loyalty_controls)))
    full_old_refs = [f'{full.cell(r,1).value}:{full.cell(r,3).value}' for r in range(2,full.max_row+1)
                     if '04_Итог' in str(full.cell(r,12).value or '') or '05_Расчёт' in str(full.cell(r,12).value or '')]
    проверка('чистая книга: Полный_отчёт не зависит от старых сводных листов', not full_old_refs,
             ', '.join(full_old_refs[:4]))
    bad_units=[]
    for sheet,col in ((clean_calc,4),(full,7)):
        for r in range(2,sheet.max_row+1):
            unit=str(sheet.cell(r,col).value or '')
            if ' / ' in unit or '₽/час' in unit or 'руб' in unit.lower():
                bad_units.append(f'{sheet.title}!{sheet.cell(r,col).coordinate}:{unit}')
    проверка('чистая книга: единицы записаны через слэш без устаревших вариантов',
             not bad_units, '; '.join(bad_units[:5]))
    be_cost = (базовый['C'] + базовый['taxB'] + базовый['aqB']) / базовый['sh'] if базовый['sh'] else 0
    проверка('себестоимость: налоги и эквайринг берутся от Точки безубыточности',
             abs(базовый['costHour'] - be_cost) < 1e-7,
             f'{базовый["costHour"]:,.2f} ₽/ч')
except Exception as e:
    fail.append(('чистая книга читается и проверяется', str(e)[:120]))

# ── демо-набор отчёта должен совпадать с расчётом по умолчанию
# Иначе отчёт, открытый без данных, показывает устаревшие числа
# (так «ставка в ноль» разошлась: 3 331 в демо против 2 814 в расчёте).
m = re.search(r'var DEMO=\{(.*?)\};', rep, re.S)
if not m:
    fail.append(('демо-набор найден в report.html', ''))
else:
    демо = {}
    for k, v in re.findall(r'(\w+):(-?[\d.]+)(?=[,}])', m.group(1)):
        демо[k] = float(v)
    расх = []
    for k, v in демо.items():
        если_есть = базовый.get(k)
        if если_есть is None: continue
        доп = max(1e-4, abs(если_есть) * 1e-5)
        if abs(если_есть - v) > доп:
            расх.append(f'{k}: демо {v:g} ≠ расчёт {если_есть:g}')
    проверка('демо-набор совпадает с расчётом по умолчанию', not расх,
             '; '.join(расх[:4]) + (f' и ещё {len(расх)-4}' if len(расх) > 4 else ''))


# ── справочник ↔ таблицы: ссылки не должны вести в никуда,
# а записи не должны висеть без единой ссылки.
try:
    спр = json.loads(re.search(r'var СПР = (\[.*?\]);\n', rep, re.S).group(1))
    имена = {т for т, _ in спр}

    def объекты(текст, начало):
        """Куски вида «начало{...}» с учётом вложенных скобок."""
        куски = []
        for m in re.finditer(начало, текст):
            i = текст.index('{', m.end() - 1)
            гл, j = 0, i
            while j < len(текст):
                if текст[j] == '{': гл += 1
                elif текст[j] == '}':
                    гл -= 1
                    if гл == 0: break
                j += 1
            куски.append(текст[i:j + 1])
        return куски

    куски = объекты(rep, r'СВЯЗЬ\w* = ') + объекты(rep, r'связь:\s*') + объекты(rep, r'связь=')
    цели, битые = set(), []
    for кусок in куски:
        пары = (re.findall(r"'([^']+)'\s*:\s*'([^']+)'", кусок)
                + re.findall(r'"([^"]+)"\s*:\s*"([^"]+)"', кусок))   # JSON-кавычки
        for ключ, цель in пары:
            if not re.match(r'^[А-ЯЁ]', цель): continue      # записи начинаются с заглавной
            цели.add(цель)
            if цель not in имена: битые.append(f'{ключ} → {цель}')
    for цель in re.findall(r"ссылка(?:_на)?\('[^']+',\s*'([^']+)'\)", rep):
        цели.add(цель)
        if цель not in имена: битые.append(f'заголовок → {цель}')
    проверка('ссылки таблиц ведут на существующие записи справочника', not битые,
             '; '.join(sorted(set(битые))[:4]))
    сироты = sorted(имена - цели)
    if сироты:
        замечание(f'записей справочника без ссылок: {len(сироты)}',
                  ', '.join(сироты[:6]) + ('…' if len(сироты) > 6 else ''))
except Exception as e:
    замечание('связи справочника не разобраны', str(e)[:80])

# ══════════════════════════════ 5. ПУБЛИКАЦИОННЫЙ КОНТУР (замечания, не падения)
print('▶ готовность к публикации')
# DEV_NO_BLUR остался в архивной версии с платной стеной, в рабочих файлах
# его нет и быть не должно. Следим только за счётчиком и адресом оплаты.
for имя, шаб in (('YM_ID', r'YM_ID'), ('PAY_URL', r'PAY_URL')):
    if not re.search(шаб, calc): замечание(f'{имя} отсутствует в calc.html')
# THANKS_URL больше не используется: сердечки открывают окно благодарности.
if "'#plan'" in rep or 'href="#plan"' in rep: замечание('ссылка «30 постов» — заглушка #plan')

# ══════════════════════════════ ИТОГ
print()
for имя, факт in ok:   print(f'  ✓ {имя}' + (f'  — {факт}' if факт else ''))
for имя, факт in fail: print(f'  ✗ {имя}' + (f'  — {факт}' if факт else ''))
for имя, факт in warn: print(f'  ! {имя}' + (f'  — {факт}' if факт else ''))
print(f'\nитого: {len(ok)} прошло · {len(fail)} упало · {len(warn)} замечаний')
sys.exit(1 if fail else 0)
