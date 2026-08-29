#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Собирает чистый лист «Интерфейс» из технической разметки HTML.

Источники: index.html, calc.html, report.html. Все смысловые блоки обязаны иметь
невидимый data-block-id. Пользовательский номер, технический ID и порядок —
разные поля. Условные части имеют data-branch-id.
"""
from pathlib import Path
from collections import defaultdict
from bs4 import BeautifulSoup, Tag
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
PAGES = [
    ('PAGE-INDEX', 1, ROOT / 'Веб' / 'index.html', 'phw-root'),
    ('PAGE-CALC', 2, ROOT / 'Веб' / 'calc.html', 'phc-root'),
    ('PAGE-REPORT', 3, ROOT / 'Веб' / 'report.html', 'phr-root'),
]
HEADERS = [
    'Страница ID', 'Порядок страницы', 'Блок ID', 'Технический №',
    'Пользовательский №', 'Название блока', 'Тип блока', 'Родительский блок',
    'Порядок блока', 'Порядок элемента', 'Тип элемента', 'DOM-ID', 'name',
    'data-t', 'Наименование / текст', 'Параметр calc', 'Поле d',
    'Ветка ID', 'Название ветки', 'Условие показа', 'Источник', 'Статус'
]


def clean(s): return re.sub(r'\s+', ' ', str(s or '')).strip()
def short(s,n=500):
    s=clean(s); return s if len(s)<=n else s[:n-1]+'…'


def direct_bn(block):
    for child in block.find_all(recursive=False):
        if isinstance(child, Tag) and 'bn' in child.get('class',[]):
            return clean(child.get_text(' ',strip=True))
    return ''


def block_type(block, user_no):
    cls=set(block.get('class',[])); name=(block.get('data-block-name') or '').lower()
    if user_no: return 'вопрос анкеты' if block.find_parent(id='frm') else 'нумерованный блок отчёта'
    if {'shm','thxm'} & cls: return 'модальное окно'
    if 'limbar' in cls or 'logi' in cls or 'trustbar' in cls: return 'уведомление'
    if 'tbar' in cls or 'savebar' in cls or 'thxbar' in cls: return 'панель действий'
    if 'foot' in cls: return 'подвал'
    if 'hdr' in cls or 'hwrap' in cls or 'hero-cover' in cls: return 'главный блок'
    if 'form-section' in cls or 'core-inputs' in cls: return 'раздел / контейнер'
    if 'card' in cls: return 'карточка отчёта'
    if 'trial' in cls or block.get('id')=='trial': return 'информационный баннер'
    return 'смысловой блок'


def element_type(el):
    if el.name=='input': return 'поле:'+el.get('type','text')
    return {'select':'выпадающий список','textarea':'текстовое поле','button':'кнопка',
            'a':'ссылка','table':'таблица'}.get(el.name,'текст / вывод')


def label(el):
    if el.name in {'button','a'}: return short(el.get_text(' ',strip=True),300)
    if el.get('data-t'): return short(el.get_text(' ',strip=True),500)
    if el.name=='input':
        lab=el.find_parent('label')
        if lab:return short(lab.get_text(' ',strip=True),300)
    own=short(el.get_text(' ',strip=True),300)
    if own and own not in {'—','×'}:return own
    return el.get('aria-label') or el.get('placeholder') or el.get('id') or el.get('name') or ''


def visibility(el, branch_name):
    r=[]
    if branch_name:r.append('активна ветка «'+branch_name+'»')
    if el.has_attr('hidden'):r.append('hidden')
    if 'display:none' in el.get('style','').replace(' ',''):r.append('display:none')
    if el.has_attr('disabled'):r.append('disabled')
    return '; '.join(r)

rows=[]
for page_id,page_order,path,root_id in PAGES:
    soup=BeautifulSoup(path.read_text(encoding='utf-8'),'html.parser')
    root=soup.select_one('#'+root_id)
    if not root or root.get('data-page-id')!=page_id:
        raise SystemExit(f'{path}: отсутствует data-page-id={page_id}')
    blocks=root.select('[data-block-id]')
    block_order={id(b):i for i,b in enumerate(blocks,1)}
    # Строка страницы.
    rows.append([page_id,page_order,page_id.replace('PAGE-','')+'-PAGE','—','—',path.name,
                 'страница','','0','0','страница',root_id,'','','', '', '', '', '', '',
                 str(path.relative_to(ROOT)),'фактически существует'])
    # Строки блоков.
    for b in blocks:
        bid=b.get('data-block-id'); user_no=direct_bn(b)
        tech=(re.search(r'-B(\d+)',bid).group(1) if re.search(r'-B(\d+)',bid) else '')
        parent=b.find_parent(attrs={'data-block-id':True})
        name=b.get('data-block-name') or short((b.find(['h1','h2','h3','h4']) or b).get_text(' ',strip=True),180)
        rows.append([page_id,page_order,bid,tech,user_no or '—',name,block_type(b,user_no),
                     parent.get('data-block-id') if parent else '',block_order[id(b)],0,'блок',
                     b.get('id',''),' ',b.get('data-t',''),name,'','','','','',
                     str(path.relative_to(ROOT)),'фактически существует'])
    # Элементы назначаются ближайшему техническому блоку.
    element_order=defaultdict(int); seen=set()
    selector='input,select,textarea,button,a,table,[id],[data-t]'
    for el in root.select(selector):
        if id(el) in seen or el.get('data-block-id'):continue
        if el.name not in {'input','select','textarea','button','a','table'} and not (el.get('id') or el.get('data-t')):continue
        block=el.find_parent(attrs={'data-block-id':True})
        if not block:continue
        seen.add(id(el));bid=block.get('data-block-id');element_order[bid]+=1
        branch=el.find_parent(attrs={'data-branch-id':True})
        branch_id=branch.get('data-branch-id','') if branch else ''
        branch_name=branch.get('data-branch-name','') if branch else ''
        dom=el.get('id','');ename=el.get('name','');param=''
        if page_id=='PAGE-CALC' and el.name in {'input','select','textarea'}:
            param=ename if el.name=='input' and el.get('type')=='radio' else dom
        bno=direct_bn(block);tech=(re.search(r'-B(\d+)',bid).group(1) if re.search(r'-B(\d+)',bid) else '')
        rows.append([page_id,page_order,bid,tech,bno or '—',block.get('data-block-name',''),
                     block_type(block,bno),block.find_parent(attrs={'data-block-id':True}).get('data-block-id') if block.find_parent(attrs={'data-block-id':True}) else '',
                     block_order[id(block)],element_order[bid],element_type(el),dom,ename,
                     el.get('data-t',''),label(el),param,'',branch_id,branch_name,
                     visibility(el,branch_name),str(path.relative_to(ROOT)),'фактически существует'])

wb=load_workbook(BOOK);ws=wb['Интерфейс'];ws.delete_rows(1,ws.max_row);ws.append(HEADERS)
for r in rows:ws.append(r)
for c in ws[1]:c.fill=PatternFill('solid',fgColor='1B9331');c.font=Font(color='FFFFFF',bold=True);c.alignment=Alignment(vertical='center',wrap_text=True)
ws.row_dimensions[1].height=38;border=Border(bottom=Side(style='thin',color='E5E7EB'))
for row in ws.iter_rows(min_row=2):
    fill=PatternFill('solid',fgColor='D9EAD3') if row[10].value=='страница' else PatternFill('solid',fgColor='EAF3FA') if row[10].value=='блок' else None
    for c in row:
        c.alignment=Alignment(vertical='top',wrap_text=True);c.border=border
        if fill:c.fill=fill;c.font=Font(bold=True)
ws.freeze_panes='A2';ws.auto_filter.ref=f'A1:V{ws.max_row}';ws.sheet_view.showGridLines=False
widths=[16,10,22,12,14,42,24,22,12,12,22,24,20,18,65,22,18,22,34,38,28,24]
for i,w in enumerate(widths,1):ws.column_dimensions[get_column_letter(i)].width=w
readme=wb['00_Читать']
for row in readme.iter_rows():
    if row[0].value=='Интерфейс':row[2].value=f'собран из технических ID HTML: {len(rows)} строк';break
wb.save(BOOK)
print(f'Интерфейс: {len(rows)} строк')
for page in ['PAGE-INDEX','PAGE-CALC','PAGE-REPORT']:
    b=sum(r[0]==page and r[10]=='блок' for r in rows);e=sum(r[0]==page and r[10] not in {'блок','страница'} for r in rows)
    print(f'  {page}: {b} блоков, {e} элементов')
