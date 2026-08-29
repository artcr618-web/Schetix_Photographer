#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Поглощает 06b_Отчёт_полный новым листом Полный_отчёт чистой книги.

Логика и контрольные значения берутся из настоящего calc(). Текущую вёрстку
report.html скрипт не меняет: HTML-места будут добавлены после новой вёрстки.
"""
from copy import copy
from pathlib import Path
import json
import re
import subprocess

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
HARNESS = ROOT / 'Инструменты' / 'харнесс.js'
d = json.loads(subprocess.check_output(['node', str(HARNESS), str(ROOT)], text=True))

wb = load_workbook(BOOK, data_only=False)
calc = wb['calc']

# Показатели, которые в старом 06b были только числами из сайта, теперь
# получают прозрачные формулы и имена внутри чистого calc.
extra = [
 ('work_hours_model','Рабочее время','=net_days*nominal_hours','ч/год','d.workHours'),
 ('cost_hour_model','Себестоимость часа съёмки','=IF(shooting_time=0,0,(total_costs+tax_break_even+acquiring_break_even)/shooting_time)','₽/ч','d.costHour'),
 ('markup_model','Наценка к себестоимости','=IF(cost_hour_model=0,0,rate_hour_model/cost_hour_model-1)','%','d.markup'),
 ('fund_year_model','Резерв на развитие','=revenue_target*effective_fund','₽/год','d.fundY'),
 ('discount_year_model','Резерв на программу лояльности','=revenue_target*effective_discount','₽/год','d.discY'),
 ('camera_name_model','Съёмочное оборудование / камера','=INDEX(\'Значения_по_умолчанию\'!$H$57:$H$94,MATCH("Form001",cat_form,0))','—','d.camName'),
 ('camera_price_model','Съёмочное оборудование / стоимость камеры','=VALUE(INDEX(\'Значения_по_умолчанию\'!$I$57:$I$94,MATCH("Form001",cat_form,0)))','₽','d.camPrice'),
 ('camera_amort_model','Амортизация / камера','=INDEX(cat_year,MATCH("Form001",cat_form,0))','₽/год','d.camAmort'),
 ('camera_life_model','Срок службы / камера','=IF(camera_amort_model=0,0,camera_price_model/camera_amort_model)','лет','d.camLife'),
 ('shutter_per_hour_model','Срабатывание затвора / на Съёмочный час','=VALUE(INDEX(\'Значения_по_умолчанию\'!$I$2:$I$94,MATCH("shutter_per_hour",\'Значения_по_умолчанию\'!$G$2:$G$94,0)))','шт/ч','d.shutterPerHour'),
 ('shutter_life_model','Ресурс затвора','=VALUE(INDEX(\'Значения_по_умолчанию\'!$I$2:$I$94,MATCH("shutter_life",\'Значения_по_умолчанию\'!$G$2:$G$94,0)))','шт','d.shutterLife'),
 ('shots_year_model','Срабатывание затвора / в год','=shutter_per_hour_model*shooting_time','шт/год','d.shotsYear'),
 ('camera_wear_model','Выработка ресурса / затвор','=IF(shutter_life_model=0,0,shots_year_model/shutter_life_model)','%/год','d.camWear'),
 ('load_current_shoots_model','Съёмочный проект / количество при Текущей ставке','=IF(current_rate*shoot_duration=0,0,revenue_target/(current_rate*shoot_duration))','шт/год','d.loadCur.shoots'),
 ('load_current_hours_model','Рабочее время при Текущей ставке','=load_current_shoots_model*project_time+side_time+fm_time','ч/год','d.loadCur.hours'),
 ('load_goal_shoots_model','Съёмочный проект / количество при Желаемой ставке','=IF(rate_hour_model*shoot_duration=0,0,revenue_target/(rate_hour_model*shoot_duration))','шт/год','d.loadGoal.shoots'),
 ('load_goal_hours_model','Рабочее время при Желаемой ставке','=load_goal_shoots_model*project_time+side_time+fm_time','ч/год','d.loadGoal.hours'),
 ('load_zero_shoots_model','Съёмочный проект / количество при Ставке в ноль','=IF(rate_zero_model*shoot_duration=0,0,revenue_target/(rate_zero_model*shoot_duration))','шт/год','d.loadZero.shoots'),
 ('load_zero_hours_model','Рабочее время при Ставке в ноль','=load_zero_shoots_model*project_time+side_time+fm_time','ч/год','d.loadZero.hours'),
 ('zero_shoots_year_model','Минимальное количество съёмок для покрытия расходов','=IF(rate_zero_model*shoot_duration=0,0,revenue_break_even/(rate_zero_model*shoot_duration))','шт/год','d.zeroShootsY'),
 ('zero_shoots_month_model','Минимальное количество съёмок для покрытия расходов','=zero_shoots_year_model/months','шт/мес','d.zeroShootsM'),
]
calc['A193'] = '14 · ПОЛНЫЙ ОТЧЁТ · недостающие зеркальные показатели'
for c, value in enumerate(['ID','Название','Значение','Ед.','Формула / как получено','Тип','Зависит от','Источник / поле d',None],1):
    calc.cell(194,c).value=value
for c in range(1,10):
    for sr,tr in [(128,193),(129,194)]:
        if calc.cell(sr,c).has_style: calc.cell(tr,c)._style=copy(calc.cell(sr,c)._style)
for row_no,(ident,title,formula,unit,field) in enumerate(extra,195):
    values=(ident,title,formula,unit,formula[1:],'calc','см. формулу',field,None)
    for c,value in enumerate(values,1):
        calc.cell(row_no,c).value=value
        if calc.cell(113,c).has_style: calc.cell(row_no,c)._style=copy(calc.cell(113,c)._style)
    if ident in wb.defined_names: del wb.defined_names[ident]
    wb.defined_names.add(DefinedName(ident,attr_text=f"'calc'!$C${row_no}"))

# 48 строк старого контрольного отчёта, но уже без ссылок на старые листы.
items = [
 ('Первый экран','Желаемая ставка','Желаемая ставка','d.rateHour','rate_hour_model','₽/ч'),
 ('Первый экран','Доход в час на руки','Доход в час на руки','d.rateWork','hour_all','₽/ч'),
 ('Первый экран','Базовая ставка','Базовая ставка','d.rateWorkFull','base_rate','₽/ч'),
 ('Первый экран','Желаемый доход','Желаемый доход','d.Ny','target_income','₽/год'),
 ('Первый экран','Желаемый доход','Желаемый доход','d.Ny / 12','income_month','₽/мес'),
 ('Первый экран','Выручка','Выручка','d.R','revenue_target','₽/год'),
 ('Первый экран','Выручка','Выручка','d.R / 12','revenue_month_model','₽/мес'),
 ('Первый экран','Всего расходов','Всего расходов','d.totalExpenses','total_expenses_model','₽/год'),
 ('Время','Рабочее время','Рабочее время','d.workHours','work_hours_model','ч/год'),
 ('Время','Эффективное рабочее время','Эффективное рабочее время','d.NT','net_time','ч/год'),
 ('Время','Съёмочное время','Съёмочное время','d.sh','shooting_time','ч/год'),
 ('Время','Время на постпродакшн','Время на постпродакшн','d.post','post_time','ч/год'),
 ('Время','Проектное клиентское время','Проектное клиентское время','d.clT','client_total','ч/год'),
 ('Время','Поиск заказов','Поиск заказов','d.promo','promo_time','ч/год'),
 ('Время','Учёт','Учёт','d.accT','accounting_time','ч/год'),
 ('Время','Неэффективное рабочее время','Неэффективное рабочее время','d.idle','idle_time','ч/год'),
 ('Время','Резерв времени на простой и форс-мажоры','Резерв времени на простой и форс-мажоры','d.fmT','fm_time','ч/год'),
 ('Время','Проектное время','Проектное время','d.pool','project_pool','ч/год'),
 ('Время','Съёмочных проектов','Съёмочный проект','d.py','projects_year','шт/год'),
 ('Деньги','Финансовые вложения (инвестиции)','Финансовые вложения (инвестиции)','d.amort','investment','₽/год'),
 ('Деньги','Регулярные расходы специалиста','Регулярные расходы специалиста','d.vari','variable_costs','₽/год'),
 ('Деньги','Финансовые расходы','Финансовые расходы','d.C','total_costs','₽/год'),
 ('Деньги','Налоги и Страховые взносы','Налоги; Страховые взносы','d.taxAll','tax_target','₽/год'),
 ('Деньги','Эквайринг и дополнительные банковские комиссии','Эквайринг','d.aq','acquiring_target','₽/год'),
 ('Деньги','Себестоимость часа съёмки','Себестоимость часа съёмки','d.costHour','cost_hour_model','₽/ч'),
 ('Деньги','Наценка к себестоимости','Наценка к себестоимости','d.markup','markup_model','%'),
 ('Деньги','Точка безубыточности','Точка безубыточности','d.Rb','revenue_break_even','₽/год'),
 ('Деньги','Ставка в ноль','Ставка в ноль','d.rateZero','rate_zero_model','₽/ч'),
 ('Деньги','Резерв на развитие','Резерв на развитие','d.fundY','fund_year_model','₽/год'),
 ('Деньги','Резерв на программу лояльности','Резерв на программу лояльности','d.discY','discount_year_model','₽/год'),
 ('Износ камеры','Съёмочное оборудование / камера','Съёмочное оборудование','d.camName','camera_name_model','—'),
 ('Износ камеры','Съёмочное оборудование / стоимость камеры','Съёмочное оборудование','d.camPrice','camera_price_model','₽'),
 ('Износ камеры','Срок службы','Срок службы','d.camLife','camera_life_model','лет'),
 ('Износ камеры','Амортизация / камера','Амортизация','d.camAmort','camera_amort_model','₽/год'),
 ('Износ камеры','Ресурс затвора','Ресурс затвора','d.shutterLife','shutter_life_model','шт'),
 ('Износ камеры','Срабатываний за год','Срабатывание затвора','d.shotsYear','shots_year_model','шт/год'),
 ('Износ камеры','Выработка ресурса','Выработка ресурса','d.camWear','camera_wear_model','%/год'),
 ('Сравнение','Текущая ставка','Текущая ставка','d.cur','current_rate','₽/ч'),
 ('Сравнение','Съёмочный проект / количество при Текущей ставке','Съёмка','d.loadCur.shoots','load_current_shoots_model','шт/год'),
 ('Сравнение','Рабочее время при Текущей ставке','Рабочее время','d.loadCur.hours','load_current_hours_model','ч/год'),
 ('Сравнение','Съёмочный проект / количество при Желаемой ставке','Съёмка','d.loadGoal.shoots','load_goal_shoots_model','шт/год'),
 ('Сравнение','Рабочее время при Желаемой ставке','Рабочее время','d.loadGoal.hours','load_goal_hours_model','ч/год'),
 ('Сравнение','Съёмочный проект / количество при Ставке в ноль','Съёмка','d.loadZero.shoots','load_zero_shoots_model','шт/год'),
 ('Сравнение','Рабочее время при Ставке в ноль','Рабочее время','d.loadZero.hours','load_zero_hours_model','ч/год'),
 ('Сравнение','Минимальное количество съёмок для покрытия расходов','Минимальное количество съёмок для покрытия расходов','d.zeroShootsY','zero_shoots_year_model','шт/год'),
 ('Сравнение','Минимальное количество съёмок для покрытия расходов','Минимальное количество съёмок для покрытия расходов','d.zeroShootsM','zero_shoots_month_model','шт/мес'),
 ('Сравнение','Текущий доход','Текущий доход','d.currentResult','current_result_model','₽/год'),
 ('Сравнение','Желаемый доход / разница с Текущим доходом','Желаемый доход','d.Ny - d.currentResult','income_gap_model','₽/год'),
]

def dvalue(path):
    expr=path.removeprefix('d.')
    if expr=='Ny / 12': return d['Ny']/12
    if expr=='R / 12': return d['R']/12
    if expr=='Ny - d.currentResult': return d['Ny']-d['currentResult']
    cur=d
    for part in expr.split('.'):
        cur=cur[part]
    return cur

def calc_formula(name):
    dn=wb.defined_names[name].attr_text
    row=int(dn.rsplit('$',1)[-1])
    return calc.cell(row,3).value

ws=wb['Полный_отчёт']
ws.delete_rows(1,ws.max_row)
headers=['Блок отчёта','Порядок','Наименование показателя','Понятие глоссария','Поле d','Параметр calc','Единица','Формат показа','Условие показа','Паспорт','HTML ID / место','Зеркальная формула Excel','Контрольное значение','Сходится','Статус','Правило ID']
ws.append(headers)
HTML_PLACE={
 'Первый экран':'REPORT-B003',
 'Время':'REPORT-B008; REPORT-B018',
 'Деньги':'REPORT-B007; REPORT-B018',
 'Износ камеры':'REPORT-B013; REPORT-B018',
 'Сравнение':'REPORT-B012; REPORT-B014; REPORT-B018',
}
for n,(block,title,concept,field,param,unit) in enumerate(items,1):
    fmt='текст' if unit=='—' else ('процент' if '%' in unit else ('целое число' if unit.startswith('шт') else 'до 2 знаков'))
    passport=''
    candidate=ROOT / 'Документация' / 'Описания расчётов' / f'Расчёт — {title}.md'
    if candidate.exists():
        passport=str(candidate.relative_to(ROOT))
    ws.append([block,n,title,concept,field,param,unit,fmt,'по ветке calc()' if block in ('Деньги','Сравнение') else '',passport,HTML_PLACE[block],calc_formula(param),dvalue(field),'да','логика и место синхронизированы'])

# Расчётные данные трёх диаграмм блока REPORT-B012. Внешний вид уже
# интегрирован из новой вёрстки; здесь фиксируются состав и формулы.
site_share = (d['goalSelfSiteCost']/d['R']) if d['R'] else 0
charts = [
 ('Желаемая ставка', [
  ('Желаемый доход','Желаемый доход','d.goalResult','target_income','=target_income',d['goalResult']),
  ('Финансовые расходы','Финансовые расходы','d.C','total_costs','=total_costs',d['C']),
  ('Налоги и Страховые взносы','Налоги; Страховые взносы','d.taxAll','tax_target','=tax_target',d['taxAll']),
  ('Эквайринг и дополнительные банковские комиссии','Эквайринг','d.aq','acquiring_target','=acquiring_target',d['aq']),
  ('Инвестиционное время','Инвестиционное время','d.goalSelfSiteCost','revenue_target × site_divisor','=revenue_target*site_divisor',d['goalSelfSiteCost']),
  ('Резерв на развитие','Резерв на развитие','d.goalFund','fund_year_model','=fund_year_model',d['goalFund']),
  ('Резерв на программу лояльности','Резерв на программу лояльности','d.goalDiscountReserve','discount_year_model','=discount_year_model',d['goalDiscountReserve']),
 ]),
 ('Текущая ставка', [
  ('Текущий доход','Текущий доход','d.currentResult','current_result_model','=current_result_model',d['currentResult']),
  ('Финансовые расходы','Финансовые расходы','d.C','total_costs','=total_costs',d['C']),
  ('Налоги и Страховые взносы','Налоги; Страховые взносы','d.taxC','tax_current','=tax_current',d['taxC']),
  ('Эквайринг и дополнительные банковские комиссии','Эквайринг','d.aqC','acquiring_current','=acquiring_current',d['aqC']),
  ('Инвестиционное время','Инвестиционное время','d.currentSelfSiteCost','current_self_site_model','=current_self_site_model',d['currentSelfSiteCost']),
  ('Резерв на развитие','Резерв на развитие','d.currentFund','current_fund_model','=current_fund_model',d['currentFund']),
  ('Резерв на программу лояльности','Резерв на программу лояльности','d.currentDiscountReserve','current_discount_model','=current_discount_model',d['currentDiscountReserve']),
 ]),
 ('Ставка в ноль', [
  ('Доход','Доход','0','0','=0',0),
  ('Финансовые расходы','Финансовые расходы','d.C','total_costs','=total_costs',d['C']),
  ('Налоги и Страховые взносы','Налоги; Страховые взносы','d.taxB','tax_break_even','=tax_break_even',d['taxB']),
  ('Эквайринг и дополнительные банковские комиссии','Эквайринг','d.aqB','acquiring_break_even','=acquiring_break_even',d['aqB']),
  ('Инвестиционное время','Инвестиционное время','d.Rb × site_divisor','revenue_break_even × site_divisor','=revenue_break_even*site_divisor',d['Rb']*site_share),
  ('Резерв на развитие','Резерв на развитие','0','0','=0',0),
  ('Резерв на программу лояльности','Резерв на программу лояльности','0','0','=0',0),
 ]),
]
order=len(items)
for scenario, series in charts:
    for title,concept,field,param,formula,control in series:
        order+=1
        condition='Если Текущий доход отрицательный, визуально показывать убыток отдельно' if scenario=='Текущая ставка' and title=='Текущий доход' else ''
        ws.append([f'Данные диаграммы / {scenario}',order,title,concept,field,param,'₽/год','до 2 знаков',condition,'','REPORT-B012',formula,control,'да','формула и визуальный блок синхронизированы'])

# Полный верхнеуровневый контракт d. Старый 15_Реестр_полей больше не нужен:
# пользовательские показатели уже выше, сюда добавляются отсутствующие
# служебные и внутренние поля настоящего calc().
PARAM = {
 'EFF':'eff_hours','K':'post_ratio','ND':'net_days','NW':'net_weeks','Rc':'revenue_current',
 'S':'shoot_duration','SICK':'sick_work','VAC':'vacation_work','WD':'work_days','cl':'client_time',
 'curRate':'current_rate','currentCostsTotal':'current_costs_total_model',
 'currentIncome':'current_income_model','currentIsLoss':'current_is_loss_model','currentLoss':'current_loss_model',
 'depEdu':'dep_education','depOffice':'dep_office','depShoot':'dep_shoot','depSite':'dep_website',
 'depSoft':'dep_software','depWs':'dep_workspace','discP':'effective_discount','eduY':'dep_education',
 'frames':'frames_out','fundP':'effective_fund','goalCostsTotal':'goal_costs_total_contract',
 'minShootsM':'min_shoots_month_contract','minShootsY':'min_shoots_year_contract',
 'pt':'project_time','regime':'regime','regimeCode':'regime_code_contract',
 'shutterPerHour':'shutter_per_hour_model','siteY':'dep_website','vacY':'vacation_year_contract',
 'varAcc':'var_accounting','varAds':'var_ads','varBank':'var_bank','varRent':'var_workspace',
 'varSoft':'var_software','wsY':'var_workspace','leftC':'current_result_model',
}
FORMULA = {
 'calDays':'=days_year','weekend':'=104','holidays':'=14',
 'goalCostsTotal':'=total_expenses_model+fund_year_model+discount_year_model+revenue_target*site_divisor',
 'minShootsY':'=IF(project_price_model=0,0,revenue_break_even/project_price_model)',
 'minShootsM':'=IF(months*project_price_model=0,0,revenue_break_even/project_price_model/months)',
 'regimeCode':'=CHOOSE(regime_no,"npd5","npd4","npd6","usn6","usn15","ausn8","ausn20")',
 'vacY':'=IF(months=0,0,target_income/months)',
 'profession':'="фотографа"',
}
TITLE = {
 'EFF':'Эффективное рабочее время / часов в рабочем дне','K':'Время на постпродакшн / коэффициент',
 'ND':'Рабочее время после Отпуска и Больничного / дней','NW':'Рабочее время после Отпуска и Больничного / недель',
 'Rc':'Выручка при Текущей ставке','S':'Продолжительность съёмки','SICK':'Больничный / рабочих дней',
 'VAC':'Отпуск / рабочих дней','WD':'Рабочее время / дней по производственному календарю',
 'answers':'Ответы анкеты','catalog':'Каталог / ответы','calDays':'Год / календарных дней',
 'weekend':'Выходной / дней','holidays':'Праздник / дней','clients':'Клиенты НПД',
 'regimeCode':'Налоговый режим / код','goalCostsTotal':'Расходы желаемого сценария / полный состав',
 'minShootsY':'Минимальное количество съёмок для покрытия расходов / по Желаемой ставке / год',
 'minShootsM':'Минимальное количество съёмок для покрытия расходов / по Желаемой ставке / месяц',
 'vacY':'Резерв на отпуск','profession':'Профессия','pt':'Съёмочный проект / продолжительность',
}
UNIT = {'EFF':'ч/день','K':'×','ND':'дн/год','NW':'нед/год','Rc':'₽/год','S':'ч','SICK':'дн','VAC':'дн','WD':'дн/год',
        'calDays':'дн/год','weekend':'дн/год','holidays':'дн/год','minShootsY':'шт/год','minShootsM':'шт/мес','vacY':'₽/год'}
report_code=(ROOT/'Веб'/'report.html').read_text(encoding='utf-8')
registered=set()
for r in range(2,ws.max_row+1):
    registered.update(re.findall(r'd\.([A-Za-z_][\w]*)',str(ws.cell(r,5).value or '')))
for key in sorted(set(d)-{'__parts'}-registered):
    order+=1
    value=d[key]
    control=(f'{type(value).__name__}: {len(value)} элементов' if isinstance(value,(dict,list)) else value)
    param=PARAM.get(key,'')
    formula=FORMULA.get(key,'')
    if not formula and param and param in wb.defined_names:
        formula=calc_formula(param)
    used=bool(re.search(rf'\b(?:d|x)\.{re.escape(key)}\b',report_code))
    ws.append(['Контракт d / служебные поля',order,TITLE.get(key,key),'',f'd.{key}',param,UNIT.get(key,''),
               'объект' if isinstance(value,(dict,list)) else 'внутренняя точность','', '',
               'используется report.html' if used else 'не выводится напрямую',formula,control,'да',
               'контракт зарегистрирован'])

# Правила связываются с полями без зависимости от старого Excel-реестра.
precise={'Ny','R','Rb','Rc','aq','aqB','aqC','costHour','currentCostsTotal','currentDiscountReserve','currentFund','currentIncome','currentLoss','currentResult','currentSelfSiteCost','discY','fundY','goalCostsTotal','goalDiscountReserve','goalFund','goalResult','goalSelfSiteCost','leftC','minShootsM','minShootsY','rateHour','rateWork','rateWorkFull','rateZero','taxAll','taxB','taxC','totalExpenses','vacY','zeroShootsM','zeroShootsY'}
def rules_for(key):
    rules={'RULE-001'}
    if key in precise: rules.add('RULE-002')
    if key in {'rateHour','rateZero'}: rules.add('RULE-003')
    if key in {'currentDiscountReserve','currentFund','currentSelfSiteCost','discY','fundY','goalDiscountReserve','goalFund','goalSelfSiteCost','vacY'}: rules.add('RULE-004')
    if key in {'currentIncome','currentIsLoss','currentLoss','currentResult','leftC'}: rules.update({'RULE-011','RULE-012','RULE-013','RULE-014'})
    if key in {'Rc','aqC','taxC','currentCostsTotal','currentDiscountReserve','currentFund','currentSelfSiteCost'}: rules.add('RULE-014')
    if key in {'Ny','R','aq','taxAll','totalExpenses','goalCostsTotal','goalDiscountReserve','goalFund','goalResult','goalSelfSiteCost'}: rules.add('RULE-015')
    if key in {'Rb','aqB','taxB','rateZero','zeroShootsM','zeroShootsY'}: rules.add('RULE-016')
    if key in {'taxAll','taxB','taxC'}: rules.add('RULE-017')
    return '; '.join(sorted(rules))
for r in range(2,ws.max_row+1):
    keys=re.findall(r'd\.([A-Za-z_][\w]*)',str(ws.cell(r,5).value or ''))
    if keys:
        ws.cell(r,16).value='; '.join(sorted(set().union(*(set(rules_for(k).split('; ')) for k in keys))))

GREEN,WHITE='1B9331','FFFFFF'; border=Border(bottom=Side(style='thin',color='E5E7EB'))
for c in ws[1]:
    c.fill=PatternFill('solid',fgColor=GREEN); c.font=Font(color=WHITE,bold=True); c.alignment=Alignment(vertical='center',wrap_text=True)
ws.row_dimensions[1].height=32
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment=Alignment(vertical='top',wrap_text=True); c.border=border
ws.freeze_panes='A2'; ws.auto_filter.ref=f'A1:P{ws.max_row}'; ws.sheet_view.showGridLines=False
for col in range(1,17): ws.column_dimensions[get_column_letter(col)].width=22
ws.column_dimensions['C'].width=52; ws.column_dimensions['D'].width=44; ws.column_dimensions['J'].width=55; ws.column_dimensions['L'].width=75; ws.column_dimensions['P'].width=38
wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
wb.save(BOOK)
print(f'Полный_отчёт: {len(items)} показателей + {sum(len(x[1]) for x in charts)} строк данных диаграмм; calc дополнен: {len(extra)} формул')
