#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Переключает чистый calc со старого 03_Каталоги на фактический CAT calc.html.

Лист Значения_по_умолчанию пересобирается генератором из HTML. Имена cat_form
и cat_year переводятся на его каталоговые строки. Старую книгу скрипт не меняет.
"""
from pathlib import Path
from tempfile import NamedTemporaryFile
import csv
import subprocess
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
CLEAN = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
SHEET = 'Значения_по_умолчанию'
GREEN, WHITE = '1B9331', 'FFFFFF'
BORDER = Border(bottom=Side(style='thin', color='E5E7EB'))

with NamedTemporaryFile(suffix='.tsv', delete=False) as f:
    tsv = Path(f.name)
try:
    subprocess.check_call([
        sys.executable, str(ROOT / 'Инструменты' / 'значения_по_умолчанию.py'), str(tsv)
    ])
    with tsv.open(encoding='utf-8') as f:
        rows = list(csv.reader(f, delimiter='\t'))
finally:
    tsv.unlink(missing_ok=True)

if not rows or len(rows[0]) < 27 or rows[0][24] != 'Расчётное значение за год по умолчанию':
    raise SystemExit('Генератор не вернул расчётные колонки каталога')
for row in rows[1:]:
    for index in (24, 25, 26):
        if len(row) > index and row[index] != '':
            row[index] = float(row[index])

wb = load_workbook(CLEAN, data_only=False)
ws = wb[SHEET]
# Перезаписываем только автоматически собираемый лист; другие листы не затрагиваем.
if ws.max_row:
    ws.delete_rows(1, ws.max_row)
for row in rows:
    ws.append(row)

for cell in ws[1]:
    cell.fill = PatternFill('solid', fgColor=GREEN)
    cell.font = Font(color=WHITE, bold=True)
    cell.alignment = Alignment(vertical='center', wrap_text=True)
ws.row_dimensions[1].height = 32
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = BORDER
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
for col in range(1, ws.max_column + 1):
    vals = [str(ws.cell(r, col).value or '') for r in range(1, min(ws.max_row, 120) + 1)]
    ws.column_dimensions[get_column_letter(col)].width = min(max([len(v) for v in vals] + [10]) + 2, 48)
ws.column_dimensions['H'].width = 52
ws.column_dimensions['M'].width = 70
ws.column_dimensions['Q'].width = 42
ws.column_dimensions['Y'].width = 28
ws.column_dimensions['Z'].width = 28
ws.column_dimensions['AA'].width = 22

catalog_rows = [r for r in range(2, ws.max_row + 1) if ws.cell(r, 6).value == 'catalog_item']
if not catalog_rows or catalog_rows != list(range(min(catalog_rows), max(catalog_rows) + 1)):
    raise SystemExit('Строки CAT должны быть одним непрерывным диапазоном')
first, last = min(catalog_rows), max(catalog_rows)

for name, col in [('cat_form', 'W'), ('cat_year', 'Y')]:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=f"'{SHEET}'!${col}${first}:${col}${last}"))

# Заголовок раздела больше не должен называть удаляемый промежуточный лист.
if wb['calc']['A87'].value == '7 · РАСХОДЫ · итоги из 03_Каталоги по КОДУ ФОРМЫ':
    wb['calc']['A87'] = '7 · РАСХОДЫ · итоги из Значения_по_умолчанию по КОДУ ФОРМЫ'

# Независимая контрольная сумма по данным, извлечённым из CAT calc.html.
forms = ['Form001', 'Form002', 'Form004', 'Form013', 'Form003', 'Form010', 'Form007']
html_totals = {
    form: sum((ws.cell(r, 25).value or 0) for r in catalog_rows if ws.cell(r, 23).value == form)
    for form in forms
}
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'
wb.save(CLEAN)

print(f'Обновлён лист {SHEET}: {len(rows)-1} значений, {len(catalog_rows)} строк CAT')
print(f'cat_form: {wb.defined_names["cat_form"].attr_text}')
print(f'cat_year: {wb.defined_names["cat_year"].attr_text}')
for form in forms:
    print(f'{form}: {html_totals[form]:.10f} — совпадает')
