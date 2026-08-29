#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Добавляет в чистый calc сценарное финансовое ядро по настоящему calc.html.

Ядро считает выбранную выручку, налог, эквайринг и безубыточность с теми же
ограничениями и делением пополам, что функция calc(). Старую книгу не меняет.
"""
from copy import copy
from pathlib import Path
import json
import subprocess
import sys

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
HARNESS = ROOT / 'Инструменты' / 'харнесс.js'

# Сценарии проверяют не только исходный экран, но и новые ветки HTML.
SCENARIOS = [
    ('НПД 5%', {'поля': {'regime': 'npd', 'npd_who': 'mix'}}),
    ('НПД 4%', {'поля': {'regime': 'npd', 'npd_who': 'phys'}}),
    ('НПД 6%', {'поля': {'regime': 'npd', 'npd_who': 'jur'}}),
    ('УСН 6%', {'поля': {'regime': 'usn6'}}),
    ('УСН 15%', {'поля': {'regime': 'usn15'}}),
    ('АУСН 8%', {'поля': {'regime': 'ausn8'}}),
    ('АУСН 20%', {'поля': {'regime': 'ausn20'}}),
    ('без налогов', {'поля': {'tax_off': True}}),
    ('фонды включены', {'поля': {'fund_on': True, 'disc_on': True}}),
    ('сайт самостоятельно', {'радио': {'site_mode': 'self'}}),
    ('дополнительная комиссия', {'допКомиссии': 1.5}),
    ('фонды и свой сайт', {'поля': {'fund_on': True, 'disc_on': True}, 'радио': {'site_mode': 'self'}}),
]

wb = load_workbook(BOOK, data_only=False)
ws = wb['calc']

# Значения констант из чистого calc для независимого Python-зеркала формулы Excel.
vals = {ws.cell(r, 1).value: ws.cell(r, 3).value for r in range(1, ws.max_row + 1)}
FIX, EXTRA, THR, CAP = (float(vals[k]) for k in ('fixed_contrib', 'extra_contrib', 'contrib_threshold', 'contrib_cap'))
M15, M20 = float(vals['min_tax_usn']), float(vals['min_tax_ausn'])
rates = {'npd5': .05, 'npd4': .04, 'npd6': .06, 'usn6': .06, 'usn15': .15, 'ausn8': .08, 'ausn20': .20}

def tax_of(r, rg, a, C, no_tax=False):
    if no_tax or r <= 0:
        return 0.0
    if rg.startswith('npd'):
        return r * rates[rg]
    if rg == 'usn6':
        contrib = FIX + min(max(0.0, r - THR) * EXTRA, CAP)
        return max(r * .06, contrib)
    gross = max(r * (1 - a) - C, 0.0)
    if rg == 'usn15':
        # Точная неподвижная точка 12 проходов calc.html.
        add = min(max(0.0, (gross - FIX - THR) * EXTRA / (1 + EXTRA)), CAP)
        contrib = FIX + add
        base = max(gross - contrib, 0.0)
        return max(base * .15, r * M15) + contrib
    if rg == 'ausn8':
        return r * .08
    return max(gross * .20, r * M20)

def revenue_for(goal, d, rg, a, C, no_tax=False):
    lo, hi = 0.0, 1e9
    for _ in range(90):
        mid = (lo + hi) / 2
        if mid * d - C - tax_of(mid, rg, a, C, no_tax) < goal:
            lo = mid
        else:
            hi = mid
    return lo

# Проверяем математическое зеркало по фактическому результату настоящего calc().
for title, override in SCENARIOS:
    raw = subprocess.check_output(['node', str(HARNESS), str(ROOT), json.dumps(override, ensure_ascii=False)], text=True)
    d = json.loads(raw)
    R = float(d['R']); Rb = float(d['Rb']); C = float(d['C']); Ny = float(d['Ny'])
    a = float(d['aq'] / R) if R else (float(d['aqB'] / Rb) if Rb else 0.0)
    sdiv = float(d['goalSelfSiteCost'] / R) if R else 0.0
    den = max(.40, 1 - a - sdiv - float(d['fundP']) - float(d['discP']))
    no_tax = bool(override.get('поля', {}).get('tax_off'))
    py_R = revenue_for(Ny, den, d['regimeCode'], a, C, no_tax) if float(d['sh']) > 0 else 0.0
    py_Rb = revenue_for(0, 1 - a - sdiv, d['regimeCode'], a, C, no_tax)
    if abs(py_R - R) > 1e-5 or abs(py_Rb - Rb) > 1e-5:
        raise SystemExit(f'{title}: Python-зеркало не совпало: R Δ={py_R-R}, Rb Δ={py_Rb-Rb}')
    print(f'{title}: R и Rb совпадают с calc.html')

# Формула налога используется и отдельно, и как LAMBDA внутри деления пополам.
TAX_BODY = (
    'IF(tax_off=1,0,IF(rev<=0,0,CHOOSE(regime_no,'
    'rev*rate_npd_avg,rev*rate_npd_fiz,rev*rate_npd_ur,'
    'MAX(rev*rate_usn_inc,fixed_contrib+MIN(MAX(0,rev-contrib_threshold)*extra_contrib,contrib_cap)),'
    'LET(gross,MAX(rev*(1-effective_acquiring)-total_costs,0),contrib,fixed_contrib+MIN(MAX(0,(gross-fixed_contrib-contrib_threshold)*extra_contrib/(1+extra_contrib)),contrib_cap),base,MAX(gross-contrib,0),MAX(base*rate_usn_prof,rev*min_tax_usn)+contrib),'
    'rev*rate_ausn_inc,'
    'LET(gross,MAX(rev*(1-effective_acquiring)-total_costs,0),MAX(gross*rate_ausn_prof,rev*min_tax_ausn)))))'
)

def bisection_formula(goal, denominator, zero_if_no_shooting=False):
    core = (
        f'LET(goal,{goal},den,{denominator},taxFn,LAMBDA(rev,{TAX_BODY}),' 
        'bounds,REDUCE(HSTACK(0,1000000000),SEQUENCE(90),LAMBDA(z,i,'
        'LET(lo,INDEX(z,1,1),hi,INDEX(z,1,2),m,(lo+hi)/2,'
        'IF(m*den-total_costs-taxFn(m)<goal,HSTACK(m,hi),HSTACK(lo,m))))),INDEX(bounds,1,1))'
    )
    if zero_if_no_shooting:
        core = f'IF(shooting_time<=0,0,{core})'
    # Новые функции сохраняются с маркером совместимости OOXML.
    return '=' + core

rows = [
    ('tax_off', 'Налоги и Страховые взносы / не учитывать', 0, '0/1', 'Флажок tax_off', 'default', '', 'Веб/calc.html'),
    ('fund_on', 'Резерв на развитие / включение', 0, '0/1', 'Флажок fund_on', 'default', '', 'Веб/calc.html'),
    ('fund_pct_model', 'Резерв на развитие / доля', .10, '%', 'fund_pct / 100', 'default', 'fund_on', 'Веб/calc.html'),
    ('disc_on', 'Резерв на программу лояльности / включение', 0, '0/1', 'Флажок disc_on', 'default', '', 'Веб/calc.html'),
    ('disc_pct_model', 'Резерв на программу лояльности / доля', .15, '%', 'disc_pct / 100', 'default', 'disc_on', 'Веб/calc.html'),
    ('extra_bank_rate', 'Эквайринг и дополнительные банковские комиссии / дополнительная доля', 0, '%', 'суммаДопКомиссий() / 100', 'default', '', 'Веб/calc.html'),
    ('effective_acquiring', 'Эквайринг и дополнительные банковские комиссии / эффективная доля', '=MIN(MAX(acquiring+extra_bank_rate,0),0.1)', '%', 'ограничение 0–10%', 'calc', 'acquiring, extra_bank_rate', 'Веб/calc.html'),
    ('effective_fund', 'Резерв на развитие / эффективная доля', '=IF(fund_on=1,MIN(MAX(fund_pct_model,0),0.2),0)', '%', 'выключено → 0; предел 20%', 'calc', 'fund_on, fund_pct_model', 'Веб/calc.html'),
    ('effective_discount', 'Резерв на программу лояльности / эффективная доля', '=IF(disc_on=1,MIN(MAX(disc_pct_model,0),0.15),0)', '%', 'выключено → 0; предел 15%', 'calc', 'disc_on, disc_pct_model', 'Веб/calc.html'),
    ('revenue_denominator', 'Доля выручки после процентных отчислений', '=MAX(0.4,1-effective_acquiring-site_divisor-effective_fund-effective_discount)', '%', 'не меньше 40%', 'calc', 'эффективные доли', 'Веб/calc.html'),
    ('revenue_target', 'Выручка', bisection_formula('target_income', 'revenue_denominator', True), '₽/год', '90 шагов деления пополам', 'calc', 'target_income, total_costs, налог, доли', 'Веб/calc.html'),
    ('tax_target', 'Налоги и Страховые взносы', '=LET(rev,revenue_target,' + TAX_BODY + ')', '₽/год', 'taxOf(revenue_target)', 'calc', 'revenue_target, regime_no', 'Веб/calc.html'),
    ('acquiring_target', 'Эквайринг и дополнительные банковские комиссии', '=revenue_target*effective_acquiring', '₽/год', 'R × эффективная доля', 'calc', 'revenue_target, effective_acquiring', 'Веб/calc.html'),
    ('total_expenses_model', 'Всего расходов', '=total_costs+tax_target+acquiring_target', '₽/год', 'Финансовые расходы + налоги + эквайринг', 'calc', 'total_costs, tax_target, acquiring_target', 'Веб/calc.html'),
    ('revenue_break_even', 'Точка безубыточности', bisection_formula('0', '1-effective_acquiring-site_divisor'), '₽/год', '90 шагов; без резервных фондов', 'calc', 'total_costs, налог, обязательные доли', 'Веб/calc.html'),
    ('tax_break_even', 'Налоги и Страховые взносы в Точке безубыточности', '=LET(rev,revenue_break_even,' + TAX_BODY + ')', '₽/год', 'taxOf(revenue_break_even)', 'calc', 'revenue_break_even, regime_no', 'Веб/calc.html'),
    ('acquiring_break_even', 'Эквайринг и дополнительные банковские комиссии в Точке безубыточности', '=revenue_break_even*effective_acquiring', '₽/год', 'Rb × эффективная доля', 'calc', 'revenue_break_even, effective_acquiring', 'Веб/calc.html'),
    ('rate_zero_model', 'Ставка в ноль', '=IF(shooting_time=0,0,revenue_break_even/shooting_time)', '₽/ч', 'Точка безубыточности / съёмочное время', 'calc', 'revenue_break_even, shooting_time', 'Веб/calc.html'),
]
start = 154
ws['A152'] = '12 · СЦЕНАРНОЕ ФИНАНСОВОЕ ЯДРО · эталон calc.html'
for c, value in enumerate(['ID', 'Название', 'Значение', 'Ед.', 'Формула / как получено', 'Тип', 'Зависит от', 'Источник / форма', None], 1):
    ws.cell(153, c).value = value
for c in range(1, 10):
    for sr, tr in [(128, 152), (129, 153)]:
        if ws.cell(sr, c).has_style:
            ws.cell(tr, c)._style = copy(ws.cell(sr, c)._style)

for i, item in enumerate(rows, start):
    for c, value in enumerate(item + (None,), 1):
        ws.cell(i, c).value = value
        if ws.cell(113, c).has_style:
            ws.cell(i, c)._style = copy(ws.cell(113, c)._style)
    ident = item[0]
    if ident in wb.defined_names:
        del wb.defined_names[ident]
    wb.defined_names.add(DefinedName(ident, attr_text=f"'calc'!$C${i}"))

# Главная ставка теперь питается актуальной выбранной выручкой, а семь старых
# веток остаются рядом только как прозрачный контроль базового состояния.
ws['C132'] = '=IF(net_time=0,0,revenue_target/net_time)'
ws['G132'] = 'revenue_target, net_time'
ws['H132'] = 'Веб/calc.html · выбранный сценарий'
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'
wb.save(BOOK)
# Компактная формула выше является читаемой спецификацией алгоритма. Для самой
# книги сразу разворачиваем её в совместимые с Excel 2010+ обычные формулы.
subprocess.check_call([sys.executable, str(ROOT/'Инструменты'/'перевести_финансовое_ядро_на_совместимые_формулы.py')])
print(f'Финансовое ядро записано: {len(rows)} строк; совместимые формулы развёрнуты')
