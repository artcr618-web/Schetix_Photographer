#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Собирает CSS-декларации фактических calc.html и report.html."""
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
import tinycss2,re
ROOT=Path(__file__).resolve().parents[1];BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа.xlsx'
PAGES=['Веб/calc.html','Веб/report.html'];HEAD=['Категория','Компонент / токен','Селектор','Свойство','Значение','Страница','Контекст / адаптив','Печать','Источник','Статус']
def clean(s):return re.sub(r'\s+',' ',str(s or '')).strip()
def category(prop,ctx):
 if prop.startswith('--'):return 'CSS-токен'
 if 'keyframes' in ctx or prop in {'animation','animation-name','transition','transform'}:return 'Движение'
 if prop.startswith(('font','line-height','letter-spacing','text-')):return 'Типографика'
 if prop in {'color','background','background-color','opacity'} or prop.startswith(('border','fill','stroke')):return 'Цвет и границы'
 if prop.startswith(('margin','padding')) or prop in {'gap','row-gap','column-gap'}:return 'Отступы'
 if prop in {'display','position','top','right','bottom','left','z-index','float','clear','overflow','visibility'} or prop.startswith(('grid','flex','align','justify')):return 'Компоновка'
 if prop.startswith(('width','height','min-','max-')):return 'Размеры'
 if prop in {'border-radius','box-shadow','filter','clip-path'}:return 'Форма и эффекты'
 return 'Прочее'
def component(sel,prop):
 if prop.startswith('--'):return prop
 ids=re.findall(r'#[A-Za-zА-Яа-яЁё0-9_-]+',sel);cls=re.findall(r'\.[A-Za-zА-Яа-яЁё0-9_-]+',sel)
 return ids[-1] if ids else cls[-1] if cls else clean(sel)[:80]
rows=[];style_count=0
def parse_rules(rules,page,ctx=''):
 for rule in rules:
  if rule.type=='qualified-rule':
   sel=clean(tinycss2.serialize(rule.prelude))
   for d in tinycss2.parse_declaration_list(rule.content,skip_whitespace=True,skip_comments=True):
    if d.type!='declaration':continue
    prop=d.lower_name;value=clean(tinycss2.serialize(d.value));rows.append([category(prop,ctx),component(sel,prop),sel,prop,value,page,ctx,'да' if 'print' in ctx.lower() else 'нет','<style> '+page,'используется'])
  elif rule.type=='at-rule':
   head='@'+rule.lower_at_keyword;pre=clean(tinycss2.serialize(rule.prelude));newctx=clean((ctx+' → ' if ctx else '')+head+(' '+pre if pre else ''))
   if rule.content:
    nested=tinycss2.parse_rule_list(rule.content,skip_whitespace=True,skip_comments=True)
    if nested:parse_rules(nested,page,newctx)
    else:rows.append(['Директива',head,head,head,clean(tinycss2.serialize(rule.content)),page,newctx,'да' if 'print' in newctx.lower() else 'нет','<style> '+page,'используется'])
   else:rows.append(['Директива',head,head,head,pre,page,newctx,'нет','<style> '+page,'используется'])
for page in PAGES:
 soup=BeautifulSoup((ROOT/page).read_text(encoding='utf-8'),'html.parser');styles=soup.find_all('style');style_count+=len(styles)
 for style in styles:parse_rules(tinycss2.parse_stylesheet(style.get_text(),skip_whitespace=True,skip_comments=True),page)
wb=load_workbook(BOOK);ws=wb['CSS_и_компоненты'];ws.delete_rows(1,ws.max_row);ws.append(HEAD)
for r in rows:ws.append(r)
for c in ws[1]:c.fill=PatternFill('solid',fgColor='1B9331');c.font=Font(color='FFFFFF',bold=True);c.alignment=Alignment(vertical='center',wrap_text=True)
ws.row_dimensions[1].height=34;border=Border(bottom=Side(style='thin',color='E5E7EB'));colors={'CSS-токен':'E8F5EA','Типографика':'EAF3FA','Движение':'F3E8FF','Отступы':'FFF4CC'}
for row in ws.iter_rows(min_row=2):
 fill=PatternFill('solid',fgColor=colors[row[0].value]) if row[0].value in colors else None
 for c in row:
  c.alignment=Alignment(vertical='top',wrap_text=True);c.border=border
  if fill:c.fill=fill
ws.freeze_panes='A2';ws.auto_filter.ref=f'A1:J{ws.max_row}';ws.sheet_view.showGridLines=False
for i,w in enumerate([20,32,65,28,60,20,42,12,28,18],1):ws.column_dimensions[get_column_letter(i)].width=w
readme=wb['00_Читать']
for row in readme.iter_rows():
 if row[0].value=='CSS_и_компоненты':row[2].value=f'calc.html + report.html: {len(rows)} CSS-деклараций';break
wb.save(BOOK)
print(f'CSS calc.html + report.html: {len(rows)} деклараций, токенов {sum(r[0]=="CSS-токен" for r in rows)}, style-блоков {style_count}')
