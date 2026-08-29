#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Синхронизирует расчёт Обучения чистой книги с Form006 calc.html."""
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
wb = load_workbook(BOOK, data_only=False)
ws = wb['calc']

# В том же разделе эффективных нулей, сразу после Рабочего места.
values = ['education_excluded','Обучение / не учитывать в расчёте',0,'0/1','EXC:Form006','default',None,'Веб/calc.html',None]
for c,value in enumerate(values,1):
    ws.cell(220,c).value=value
    if ws.cell(219,c).has_style: ws.cell(220,c)._style=copy(ws.cell(219,c)._style)
if 'education_excluded' in wb.defined_names: del wb.defined_names['education_excluded']
wb.defined_names.add(DefinedName('education_excluded',attr_text="'calc'!$C$220"))

form_range="'Значения_по_умолчанию'!$W$57:$W$94"
value_range="'Значения_по_умолчанию'!$Z$57:$Z$94"
term_range="'Значения_по_умолчанию'!$AA$57:$AA$94"
ws['C75']=f'=SUMIF({form_range},"Form006",{value_range})'
ws['E75']='Сумма стоимости всех строк Form006'
ws['F75']='calc'; ws['G75']='Form006'; ws['H75']='JavaScript CAT в calc.html'
ws['C76']=f'=SUMIF({form_range},"Form006",{term_range})'
ws['E76']='Сумма месяцев без дохода по всем строкам Form006'
ws['F76']='calc'; ws['G76']='Form006'; ws['H76']='JavaScript CAT в calc.html'
ws['C79']='=IF(education_excluded=1,0,edu_months*income_month)'
ws['E79']='Исключено → 0; иначе период без дохода × Желаемый доход за месяц'
ws['G79']='education_excluded, edu_months, income_month'
ws['C80']='=IF(education_excluded=1,0,IF(edu_life=0,0,(edu_cost+edu_opportunity)/edu_life))'
ws['E80']='Исключено или срок 0 → 0; иначе стоимость и упущенный доход / срок'
ws['G80']='education_excluded, edu_cost, edu_opportunity, edu_life'
ws['H80']='Зеркало eduY в calc.html'

wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
wb.save(BOOK)
print('Обучение синхронизировано с Form006: стоимость, период без дохода, исключение')
