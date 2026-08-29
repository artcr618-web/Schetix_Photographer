#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Собирает пользовательские тексты из фактических calc.html и report.html."""
from pathlib import Path
from collections import Counter,defaultdict
from bs4 import BeautifulSoup,Tag
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side
from openpyxl.utils import get_column_letter
import re
ROOT=Path(__file__).resolve().parents[1];BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа.xlsx'
PAGES=[('Веб/calc.html','phc-root','CALC'),('Веб/report.html','phr-root','REPORT')]
HEADERS=['Страница','№ блока','Блок ID','Название блока','Порядок текста','Ключ текста','data-t','DOM-ID','Тег / тип','Текст','Условие показа','Источник','Статус']
def clean(s):return re.sub(r'\s+',' ',str(s or '')).strip()
def css_path(el,stop):
 parts=[];p=el
 while isinstance(p,Tag) and p is not stop and len(parts)<6:
  if p.get('id'):parts.append('#'+p.get('id'));break
  siblings=[x for x in p.parent.find_all(p.name,recursive=False)] if p.parent else []
  parts.append(f'{p.name}:nth-of-type({siblings.index(p)+1 if p in siblings else 1})');p=p.parent
 return ' > '.join(reversed(parts))
def info(el,prefix):
 block=el.find_parent(attrs={'data-block-id':True})
 if block:
  bn=block.find(class_='bn');no=clean(bn.get_text(' ',strip=True)) if bn else '—'
  h=block.find(['h1','h2','h3','h4']);title=block.get('data-block-name') or (clean(h.get_text(' ',strip=True)) if h else '')
  return no,block.get('data-block-id'),title
 return '—',prefix+'-GLOBAL','Глобальные элементы'
def condition(el):
 result=[]
 if el.has_attr('hidden'):result.append('hidden')
 if 'display:none' in el.get('style','').replace(' ',''):result.append('display:none')
 branch=el.find_parent(attrs={'data-branch-id':True})
 if branch:result.append('ветка '+branch.get('data-branch-id',''))
 return '; '.join(result)
# Считаем повторения data-t во всех страницах как мультимножество.
all_keys=[];soups=[]
for rel,rootid,prefix in PAGES:
 soup=BeautifulSoup((ROOT/rel).read_text(encoding='utf-8'),'html.parser');root=soup.select_one('#'+rootid) or soup
 soups.append((rel,prefix,soup,root));all_keys += [e.get('data-t') for e in soup.select('[data-t]')]
counts=Counter(all_keys);seen=Counter();rows=[];orders=defaultdict(int)
for page_no,(rel,prefix,soup,root) in enumerate(soups):
 marked=set()
 for el in soup.select('[data-t]'):
  text=clean(el.get_text(' ',strip=True))
  if not text:continue
  no,bid,title=info(el,prefix);orders[(rel,bid)]+=1;key=el.get('data-t');seen[key]+=1
  occurrence=f'#{seen[key]}' if counts[key]>1 else ''
  rows.append([rel,no,bid,title,orders[(rel,bid)],key+occurrence,key,el.get('id',''),el.name,text,condition(el),rel,'data-t; повторяется' if counts[key]>1 else 'data-t; используется']);marked.add(id(el))
 selector='h1,h2,h3,h4,p,label,button,a,th,option,.q,.qd,.hint,.fnote,.sm,.tp,.workmore-q,.logi,.wnote'
 for el in soup.select(selector):
  if id(el) in marked or el.get('data-t'):continue
  children=el.select('[data-t]');own=clean(el.get_text(' ',strip=True))
  if children and own==clean(' '.join(x.get_text(' ',strip=True) for x in children)):continue
  if not own or own in {'×','—'}:continue
  no,bid,title=info(el,prefix);orders[(rel,bid)]+=1
  key=('ID:'+el.get('id')) if el.get('id') else 'STATIC:'+css_path(el,el.find_parent(attrs={'data-block-id':True}) or root)
  rows.append([rel,no,bid,title,orders[(rel,bid)],key,'',el.get('id',''),el.name,own,condition(el),rel,'статический текст без data-t'])
page_order={p[0]:i for i,p in enumerate(PAGES)}
rows.sort(key=lambda r:(page_order[r[0]],r[2],r[4],r[5]))
wb=load_workbook(BOOK);ws=wb['Тексты'];ws.delete_rows(1,ws.max_row);ws.append(HEADERS)
for row in rows:ws.append(row)
for c in ws[1]:c.fill=PatternFill('solid',fgColor='1B9331');c.font=Font(color='FFFFFF',bold=True);c.alignment=Alignment(vertical='center',wrap_text=True)
ws.row_dimensions[1].height=34;border=Border(bottom=Side(style='thin',color='E5E7EB'))
for row in ws.iter_rows(min_row=2):
 for c in row:c.alignment=Alignment(vertical='top',wrap_text=True);c.border=border
 if row[12].value=='статический текст без data-t':
  for c in row:c.fill=PatternFill('solid',fgColor='FFF4CC')
ws.freeze_panes='A2';ws.auto_filter.ref=f'A1:M{ws.max_row}';ws.sheet_view.showGridLines=False
for i,w in enumerate([20,10,22,40,12,40,20,24,14,100,34,24,28],1):ws.column_dimensions[get_column_letter(i)].width=w
readme=wb['00_Читать']
for row in readme.iter_rows():
 if row[0].value=='Тексты':row[2].value=f'calc.html + report.html: {len(rows)} строк, {sum(bool(r[6]) for r in rows)} data-t';break
wb.save(BOOK)
print(f'Тексты calc.html + report.html: {len(rows)} строк; data-t {sum(bool(r[6]) for r in rows)}; статические {sum(not r[6] for r in rows)}')
