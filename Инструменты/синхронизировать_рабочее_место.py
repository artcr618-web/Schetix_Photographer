#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Синхронизирует блок «Рабочее место» чистого calc с calc.html.

Добавляет эффективный ноль полного исключения Form009b и сохраняет отдельные
ветки: дом, собственное жильё без ипотеки, отдельное помещение.
"""
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
wb = load_workbook(BOOK, data_only=False)
ws = wb['calc']

# Новую строку добавляем в хвост, не сдвигая вручную собранный лист и уже
# закреплённые определённые имена.
ws['A217'] = '15 · ЭФФЕКТИВНЫЕ НУЛИ БЛОКОВ'
for c, value in enumerate(['ID','Название','Значение','Ед.','Формула / как получено','Тип','Зависит от','Источник / форма',None], 1):
    ws.cell(218, c).value = value
for c in range(1, 10):
    for sr, tr in [(128,217),(129,218)]:
        if ws.cell(sr,c).has_style:
            ws.cell(tr,c)._style = copy(ws.cell(sr,c)._style)
values = [
    'workspace_excluded', 'Рабочее место / не учитывать в расчёте', 0,
    '0/1', 'EXC:Form009b', 'default', None, 'Веб/calc.html', None,
]
for c, value in enumerate(values, 1):
    ws.cell(219, c).value = value
    if ws.cell(64, c).has_style:
        ws.cell(219, c)._style = copy(ws.cell(64, c)._style)
if 'workspace_excluded' in wb.defined_names:
    del wb.defined_names['workspace_excluded']
wb.defined_names.add(DefinedName('workspace_excluded', attr_text="'calc'!$C$219"))

old_ws = '=IF(ws_mode="Работаю из дома",(IF(own_home="да",0,home_rent)+home_util)*cab_share*months,(office_rent+office_util)*months)'
new_ws = '=IF(workspace_excluded=1,0,IF(ws_mode="Работаю из дома",(IF(own_home="да",0,home_rent)+home_util)*cab_share*months,(office_rent+office_util)*months))'
if ws['C72'].value not in (old_ws, new_ws):
    raise SystemExit(f'Неожиданная формула C72: {ws["C72"].value!r}')
ws['C72'] = new_ws
ws['G72'] = 'workspace_excluded, ws_mode, own_home, home_rent, home_util, cab_share, office_rent, office_util, months'
ws['H72'] = 'Исключение блока → 0; собственное жильё → стоимость жилья 0'

old_dep = '=SUMIF(cat_form,"Form013",cat_year)'
new_dep = '=IF(workspace_excluded=1,0,SUMIF(cat_form,"Form013",cat_year))'
if ws['C92'].value not in (old_dep, new_dep):
    raise SystemExit(f'Неожиданная формула C92: {ws["C92"].value!r}')
ws['C92'] = new_dep
ws['G92'] = 'workspace_excluded, Form013'
ws['H92'] = 'Обустройство рабочего места получает эффективный ноль вместе со всем блоком'

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'
wb.save(BOOK)
print('Рабочее место синхронизировано: дом, своё жильё, помещение, полное исключение')
