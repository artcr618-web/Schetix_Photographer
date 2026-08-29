#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Заменяет LET/LAMBDA/REDUCE на обычные формулы Excel 2010+.

90 шагов деления пополам разворачиваются в скрытую техническую таблицу
листa calc. Математика остаётся той же, но книгу могут пересчитать старые Excel
и LibreOffice без динамических функций Microsoft 365.
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font,PatternFill
from openpyxl.utils import get_column_letter

ROOT=Path(__file__).resolve().parents[1]
BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа.xlsx'
wb=load_workbook(BOOK,data_only=False); ws=wb['calc']
START=230; LAST=319; FINAL=320; CURRENT=321

# K:P — целевая Выручка; Q:V — Точка безубыточности.
headers=['Низ target','Верх target','Середина target','База target','Взносы target','Налог target',
         'Низ zero','Верх zero','Середина zero','База zero','Взносы zero','Налог zero']
for c,title in enumerate(headers,11): ws.cell(229,c).value=title

def contrib(gross):
    return f'=fixed_contrib+MIN(MAX(0,({gross}-fixed_contrib-contrib_threshold)*extra_contrib/(1+extra_contrib)),contrib_cap)'
def tax(rev,gross,con):
    return (f'=IF(tax_off=1,0,IF({rev}<=0,0,CHOOSE(regime_no,'
            f'{rev}*rate_npd_avg,{rev}*rate_npd_fiz,{rev}*rate_npd_ur,'
            f'MAX({rev}*rate_usn_inc,fixed_contrib+MIN(MAX(0,{rev}-contrib_threshold)*extra_contrib,contrib_cap)),'
            f'MAX(MAX({gross}-{con},0)*rate_usn_prof,{rev}*min_tax_usn)+{con},'
            f'{rev}*rate_ausn_inc,MAX({gross}*rate_ausn_prof,{rev}*min_tax_ausn))))')

for r in range(START,LAST+1):
    if r==START:
        ws[f'K{r}']=0; ws[f'L{r}']=1000000000
        ws[f'Q{r}']=0; ws[f'R{r}']=1000000000
    else:
        p=r-1
        cond=f'M{p}*revenue_denominator-total_costs-P{p}<target_income'
        ws[f'K{r}']=f'=IF({cond},M{p},K{p})'
        ws[f'L{r}']=f'=IF({cond},L{p},M{p})'
        cond0=f'S{p}*(1-effective_acquiring-site_divisor)-total_costs-V{p}<0'
        ws[f'Q{r}']=f'=IF({cond0},S{p},Q{p})'
        ws[f'R{r}']=f'=IF({cond0},R{p},S{p})'
    ws[f'M{r}']=f'=(K{r}+L{r})/2'
    ws[f'N{r}']=f'=MAX(M{r}*(1-effective_acquiring)-total_costs,0)'
    ws[f'O{r}']=contrib(f'N{r}')
    ws[f'P{r}']=tax(f'M{r}',f'N{r}',f'O{r}')
    ws[f'S{r}']=f'=(Q{r}+R{r})/2'
    ws[f'T{r}']=f'=MAX(S{r}*(1-effective_acquiring)-total_costs,0)'
    ws[f'U{r}']=contrib(f'T{r}')
    ws[f'V{r}']=tax(f'S{r}',f'T{r}',f'U{r}')

# Финальный low после 90-го сравнения и налог ровно от финальной величины.
ws[f'K{FINAL}']=f'=IF(M{LAST}*revenue_denominator-total_costs-P{LAST}<target_income,M{LAST},K{LAST})'
ws[f'N{FINAL}']=f'=MAX(K{FINAL}*(1-effective_acquiring)-total_costs,0)'
ws[f'O{FINAL}']=contrib(f'N{FINAL}')
ws[f'P{FINAL}']=tax(f'K{FINAL}',f'N{FINAL}',f'O{FINAL}')
ws[f'Q{FINAL}']=f'=IF(S{LAST}*(1-effective_acquiring-site_divisor)-total_costs-V{LAST}<0,S{LAST},Q{LAST})'
ws[f'T{FINAL}']=f'=MAX(Q{FINAL}*(1-effective_acquiring)-total_costs,0)'
ws[f'U{FINAL}']=contrib(f'T{FINAL}')
ws[f'V{FINAL}']=tax(f'Q{FINAL}',f'T{FINAL}',f'U{FINAL}')
# Налог текущего сценария.
ws[f'M{CURRENT}']='=revenue_current'
ws[f'N{CURRENT}']=f'=MAX(M{CURRENT}*(1-effective_acquiring)-total_costs,0)'
ws[f'O{CURRENT}']=contrib(f'N{CURRENT}')
ws[f'P{CURRENT}']=tax(f'M{CURRENT}',f'N{CURRENT}',f'O{CURRENT}')

# Публичные строки calc ссылаются на итог скрытого алгоритма.
ws['C164']=f'=IF(shooting_time<=0,0,K{FINAL})'
ws['E164']='90 шагов деления пополам; совместимые вспомогательные формулы K:V'
ws['C165']=f'=P{FINAL}'
ws['E165']='taxOf(revenue_target), рассчитан в скрытой таблице'
ws['C168']=f'=Q{FINAL}'
ws['E168']='90 шагов деления пополам без добровольных резервов'
ws['C169']=f'=V{FINAL}'
ws['E169']='taxOf(revenue_break_even), рассчитан в скрытой таблице'
ws['C180']=f'=P{CURRENT}'
ws['E180']='taxOf(revenue_current), рассчитан в скрытой таблице'

# Служебный блок доступен для аудита, но не мешает основному листу.
for c in range(11,23):
    ws.column_dimensions[get_column_letter(c)].hidden=True
    ws.column_dimensions[get_column_letter(c)].width=14
for r in range(229,CURRENT+1): ws.row_dimensions[r].hidden=True
for c in range(11,23):
    ws.cell(229,c).font=Font(bold=True,color='FFFFFF')
    ws.cell(229,c).fill=PatternFill('solid',fgColor='555555')

wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
wb.save(BOOK)
print(f'Финансовое ядро переведено на совместимые формулы: {LAST-START+1} итераций × 2 сценария')
