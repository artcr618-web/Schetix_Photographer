#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Публикует пользовательскую часть `01_Глоссарий` в report и PDF.

Единственный источник терминов и определений — лист `01_Глоссарий` основной
книги. В справочник автоматически попадают все строки с Назначением `user` и
непустым определением, а также ранее утверждённое пользовательское meta-понятие
«Значение по умолчанию». Скрипт обновляет оба массива справочника в report.html и
его рабочем каркасе. Карта связей строк детализации сохраняется отдельно: её
нельзя угадывать по совпадению слов, поэтому она остаётся явным реестром.

Обычный запуск не нужен: скрипт входит в `Инструменты/карта.py`.
"""
from pathlib import Path
import json
import re

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'


PUBLIC_META = {'Значение по умолчанию'}


def glossary_user_rows():
    wb = openpyxl.load_workbook(BOOK, data_only=True, read_only=True)
    ws = wb['01_Глоссарий']
    rows = []
    seen = set()
    # A № · B Наименование · D Назначение · F Определение.
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 2).value or '').strip()
        purpose = str(ws.cell(r, 4).value or '').strip()
        definition = re.sub(r'\s+', ' ', str(ws.cell(r, 6).value or '')).strip()
        if purpose != 'user' and name not in PUBLIC_META:
            continue
        if not name or not definition:
            raise SystemExit(f'01_Глоссарий, строка {r}: user-понятие без имени или определения')
        key = name.casefold().replace('ё', 'е')
        if key in seen:
            raise SystemExit(f'01_Глоссарий, строка {r}: повтор user-понятия «{name}»')
        seen.add(key)
        rows.append((name, definition))
    return sorted(rows, key=lambda item: item[0].casefold().replace('ё', 'е'))


rows = glossary_user_rows()
refd = 'var REFD=[\n' + ',\n'.join(
    '   ' + json.dumps([name, definition], ensure_ascii=False)
    for name, definition in rows
) + '\n  ];'
spr = 'var СПР = ' + json.dumps(rows, ensure_ascii=False) + ';'

for relative in ('Веб/report.html', 'Веб/Части/каркас.html'):
    path = ROOT / relative
    text = path.read_text(encoding='utf-8')
    text, n1 = re.subn(r'var REFD=\[.*?\n *\];', refd, text, count=1, flags=re.S)
    text, n2 = re.subn(r'var СПР = \[.*?\];\n', spr + '\n', text, count=1, flags=re.S)
    if n1 != 1 or n2 != 1:
        raise SystemExit(f'{relative}: не найдены массивы REFD/СПР ({n1}/{n2})')
    path.write_text(text, encoding='utf-8')
    print(f'{relative}: опубликовано user-понятий {len(rows)}')
