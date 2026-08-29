#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Синхронизирует две ветки Сайта и эффективный ноль Form014."""
from copy import copy
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

ROOT=Path(__file__).resolve().parents[1]
BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа.xlsx'
wb=load_workbook(BOOK,data_only=False); ws=wb['calc']

values=['site_excluded','Сайт / не учитывать в расчёте',0,'0/1','EXC:Form014','default',None,'Веб/calc.html',None]
for c,value in enumerate(values,1):
    ws.cell(221,c).value=value
    if ws.cell(220,c).has_style: ws.cell(221,c)._style=copy(ws.cell(220,c)._style)
if 'site_excluded' in wb.defined_names: del wb.defined_names['site_excluded']
wb.defined_names.add(DefinedName('site_excluded',attr_text="'calc'!$C$221"))

ws['C85']='=IF(site_excluded=1,0,IF(site_mode="Нанимал (а) специалиста",IF(site_life=0,0,site_cost/site_life),0))'
ws['E85']='Исключено или создано самостоятельно → 0; подрядчик: стоимость / срок'
ws['G85']='site_excluded, site_mode, site_cost, site_life'; ws['H85']='Зеркало siteY в calc.html'
ws['C86']='=IF(site_excluded=1,0,IF(site_mode="Нанимал (а) специалиста",0,IF(site_life*net_time=0,0,site_hours/(site_life*net_time))))'
ws['E86']='Доля Выручки для компенсации Инвестиционного времени самостоятельного создания'
ws['G86']='site_excluded, site_mode, site_hours, site_life, net_time'; ws['H86']='Зеркало sdiv в calc.html'

wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
wb.save(BOOK)
print('Сайт синхронизирован: подрядчик, самостоятельное создание, полное исключение')
