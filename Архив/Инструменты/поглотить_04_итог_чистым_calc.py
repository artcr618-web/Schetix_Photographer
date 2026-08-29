#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Переносит уникальные результаты 04_Итог в чистый calc по контракту calc().

Сводный старый лист не копируется: расчёты становятся именованными строками
calc, а самопроверки остаются в штатном проверить.py.
"""
from copy import copy
from pathlib import Path
import json
import subprocess

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
CLEAN = ROOT / 'Книга' / 'Калькулятор_ставки_часа_чистая.xlsx'
OLD = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
HARNESS = ROOT / 'Инструменты' / 'харнесс.js'

d = json.loads(subprocess.check_output(['node', str(HARNESS), str(ROOT)], text=True))
old = load_workbook(OLD, data_only=True, read_only=True)['04_Итог']
# Сначала подтверждаем, какие старые показатели действительно совпадают с HTML.
checks = {
    'B5': d['rateHour'], 'B6': d['R'] / d['py'] if d['py'] else 0,
    'B7': d['rateWorkFull'], 'B8': d['R'], 'B9': d['R'] / 12,
    'B10': d['taxAll'], 'B11': d['taxAll'] / d['R'] if d['R'] else 0,
    'B12': d['Rb'], 'B41': d['currentResult'],
}
for cell, html in checks.items():
    book = float(old[cell].value)
    if abs(book - float(html)) > 1e-6:
        raise SystemExit(f'04_Итог!{cell}: старая книга={book}, calc.html={html}')

wb = load_workbook(CLEAN, data_only=False)
ws = wb['calc']
# Берём ту же taxOf-формулу, которая уже проверена финансовым ядром.
tax_target_formula = ws['C165'].value
if not isinstance(tax_target_formula, str) or not tax_target_formula.startswith('=LET(rev,revenue_target,'):
    raise SystemExit('Не найдена ожидаемая формула tax_target в C165')
tax_current_formula = tax_target_formula.replace('=LET(rev,revenue_target,', '=LET(rev,revenue_current,', 1)

rows = [
    ('rate_hour_model', 'Желаемая ставка', '=IF(shooting_time=0,0,revenue_target/shooting_time)', '₽/ч', 'Выручка / съёмочное время', 'calc', 'revenue_target, shooting_time', 'd.rateHour'),
    ('project_price_model', 'Стоимость съёмочного проекта', '=IF(projects_year=0,0,revenue_target/projects_year)', '₽/проект', 'Выручка / количество проектов', 'calc', 'revenue_target, projects_year', 'calc.html · projPrice'),
    ('revenue_month_model', 'Выручка в месяц', '=revenue_target/months', '₽/мес', 'Выручка / 12', 'calc', 'revenue_target, months', 'd.R / 12'),
    ('effective_tax_load', 'Эффективная налоговая нагрузка', '=IF(revenue_target=0,0,tax_target/revenue_target)', '%', 'Налоги / Выручка', 'calc', 'tax_target, revenue_target', 'calc.html'),
    ('revenue_current', 'Выручка при Текущей ставке', '=current_rate*shooting_time', '₽/год', 'Текущая ставка × съёмочное время', 'calc', 'current_rate, shooting_time', 'd.Rc'),
    ('tax_current', 'Налоги и Страховые взносы при Текущей ставке', tax_current_formula, '₽/год', 'taxOf(Rc)', 'calc', 'revenue_current, regime_no', 'd.taxC'),
    ('acquiring_current', 'Эквайринг и дополнительные банковские комиссии при Текущей ставке', '=revenue_current*effective_acquiring', '₽/год', 'Rc × эффективная доля', 'calc', 'revenue_current, effective_acquiring', 'd.aqC'),
    ('current_fund_model', 'Резерв на развитие при Текущей ставке', '=revenue_current*effective_fund', '₽/год', 'Rc × эффективная доля', 'calc', 'revenue_current, effective_fund', 'd.currentFund'),
    ('current_discount_model', 'Резерв на программу лояльности при Текущей ставке', '=revenue_current*effective_discount', '₽/год', 'Rc × эффективная доля', 'calc', 'revenue_current, effective_discount', 'd.currentDiscountReserve'),
    ('current_self_site_model', 'Сайт / компенсация Инвестиционного времени при Текущей ставке', '=revenue_current*site_divisor', '₽/год', 'Rc × поправка сайта', 'calc', 'revenue_current, site_divisor', 'd.currentSelfSiteCost'),
    ('current_costs_total_model', 'Расходы при Текущей ставке / полный состав', '=total_costs+tax_current+acquiring_current+current_fund_model+current_discount_model+current_self_site_model', '₽/год', 'Полный состав текущего сценария', 'calc', 'текущие затраты', 'd.currentCostsTotal'),
    ('current_result_model', 'Текущий доход', '=revenue_current-current_costs_total_model', '₽/год', 'Выручка минус все затраты', 'calc', 'revenue_current, current_costs_total_model', 'd.currentResult'),
    ('current_income_model', 'Текущий доход / неотрицательная часть', '=MAX(current_result_model,0)', '₽/год', 'MAX(результат; 0)', 'calc', 'current_result_model', 'd.currentIncome'),
    ('current_loss_model', 'Текущий доход / убыток', '=MIN(current_result_model,0)', '₽/год', 'MIN(результат; 0)', 'calc', 'current_result_model', 'd.currentLoss'),
    ('current_is_loss_model', 'Текущий доход / признак убытка', '=current_result_model<0', 'да/нет', 'Результат меньше нуля', 'calc', 'current_result_model', 'd.currentIsLoss'),
    ('rate_difference_model', 'Желаемая ставка / разница с Текущей ставкой', '=rate_hour_model-current_rate', '₽/ч', 'Желаемая ставка − текущая ставка', 'calc', 'rate_hour_model, current_rate', 'calc.html'),
    ('income_gap_model', 'Желаемый доход / разница с Текущим доходом', '=target_income-current_result_model', '₽/год', 'Желаемый доход − текущий результат', 'calc', 'target_income, current_result_model', 'calc.html'),
]

ws['A173'] = '13 · ИТОГОВЫЕ ПОКАЗАТЕЛИ · поглощён 04_Итог'
headers = ['ID', 'Название', 'Значение', 'Ед.', 'Формула / как получено', 'Тип', 'Зависит от', 'Источник / поле d', None]
for c, value in enumerate(headers, 1):
    ws.cell(174, c).value = value
for c in range(1, 10):
    for sr, tr in [(128, 173), (129, 174)]:
        if ws.cell(sr, c).has_style:
            ws.cell(tr, c)._style = copy(ws.cell(sr, c)._style)

for row_no, item in enumerate(rows, 175):
    for c, value in enumerate(item + (None,), 1):
        ws.cell(row_no, c).value = value
        if ws.cell(113, c).has_style:
            ws.cell(row_no, c)._style = copy(ws.cell(113, c)._style)
    ident = item[0]
    if ident in wb.defined_names:
        del wb.defined_names[ident]
    wb.defined_names.add(DefinedName(ident, attr_text=f"'calc'!$C${row_no}"))

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'
wb.save(CLEAN)
print(f'04_Итог поглощён: {len(rows)} итоговых показателей перенесено в calc')
print('9 контрольных значений старого листа совпали с calc.html')
