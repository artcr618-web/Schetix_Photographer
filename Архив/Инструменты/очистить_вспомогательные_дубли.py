#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Удаляет подтверждённые остатки старых режимов из чистого calc."""
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа_чистая.xlsx'
wb=load_workbook(BOOK,data_only=False); ws=wb['calc']
remove={
    'week_target','shoot_auto','fm_week','projects_month','projects_week',
    'shooting_month','shooting_week','ops_per_shoot','frames_total','equip_only',
}
for r in range(1,ws.max_row+1):
    if ws.cell(r,1).value in remove:
        for c in range(1,10): ws.cell(r,c).value=None
for ident in remove:
    if ident in wb.defined_names: del wb.defined_names[ident]
left=[]
for row in ws.iter_rows():
    for c in row:
        if isinstance(c.value,str) and c.value.startswith('='):
            for ident in remove:
                if ident in c.value: left.append(f'{c.coordinate}:{ident}')
if left: raise SystemExit('Остались зависимости: '+', '.join(left))
wb.save(BOOK)
print('Удалён старый авторежим:',', '.join(sorted(remove)))
