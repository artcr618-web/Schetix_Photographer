#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Удаляет старые налоговые дубли после перехода на сценарное ядро calc.html."""
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
BOOK=ROOT/'Книга'/'Калькулятор_ставки_часа_чистая.xlsx'
wb=load_workbook(BOOK,data_only=False); ws=wb['calc']

# Семь статических контрольных веток совпадали только в базовом состоянии.
# Их заменяет единое сценарное revenue_target с taxOf() и делением пополам.
remove_ids={
 'revenue_base','rate_current','threshold_credit','injury_contrib',
 'revenue_npd_avg','revenue_npd_fiz','revenue_npd_ur','revenue_usn_inc',
 'revenue_usn_prof','revenue_ausn_inc','revenue_ausn_prof',
}
for r in range(1,ws.max_row+1):
    if ws.cell(r,1).value in remove_ids:
        for c in range(1,10): ws.cell(r,c).value=None
# Заголовок и шапка удалённого контрольного раздела.
for r in (142,143):
    for c in range(1,10): ws.cell(r,c).value=None
for ident in remove_ids:
    if ident in wb.defined_names: del wb.defined_names[ident]

# Ни одна оставшаяся формула не должна читать удалённые имена.
left=[]
for row in ws.iter_rows():
    for c in row:
        if isinstance(c.value,str) and c.value.startswith('='):
            for ident in remove_ids:
                if ident in c.value: left.append(f'{c.coordinate}:{ident}')
if left: raise SystemExit('Остались зависимости: '+', '.join(left))

wb.calculation.fullCalcOnLoad=True; wb.calculation.forceFullCalc=True; wb.calculation.calcMode='auto'
wb.save(BOOK)
print('Удалены старые налоговые дубли:',len(remove_ids))
