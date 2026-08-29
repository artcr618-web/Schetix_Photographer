#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Удаляет неиспользуемый агрегат Профильное время из чистой книги.

Старую контрольную книгу не меняет: там core_time ещё участвует в старом
04_Итог. В активной чистой архитектуре строка и определённое имя удаляются.
"""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BOOK = ROOT / 'Книга' / 'Калькулятор_ставки_часа_чистая.xlsx'
wb = load_workbook(BOOK, data_only=False)
ws = wb['calc']

rows = [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == 'core_time']
if rows not in ([53], []):
    raise SystemExit(f'Неожиданные строки core_time: {rows}')
if rows:
    # Строку физически не сдвигаем: последующие 160+ определённых имён уже
    # адресуют конкретные строки. Пустая строка не является активной записью.
    for c in range(1, 10):
        ws.cell(rows[0], c).value = None
if 'core_time' in wb.defined_names:
    del wb.defined_names['core_time']

# Защита: в активных листах чистой книги не должно остаться понятия или ID.
left = []
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for cell in row:
            text = str(cell.value or '')
            if 'core_time' in text or 'ПРОФИЛЬНОЕ время' in text or 'Профильное время' in text:
                left.append(f'{sh.title}!{cell.coordinate}')
if left:
    raise SystemExit('Остались ссылки: ' + ', '.join(left))

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'
wb.save(BOOK)
print('Профильное время удалено из чистого calc; строка 53 оставлена пустой без сдвига адресов')
