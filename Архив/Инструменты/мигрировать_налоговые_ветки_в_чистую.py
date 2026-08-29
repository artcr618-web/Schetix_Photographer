#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Переносит семь формул целевой выручки из 05_Расчёт внутрь чистого calc.

Это устраняет внешнюю ссылку чистого calc. Перед записью контрольные значения
старого листа по всем семи режимам сверяются с настоящим calc() из calc.html.
Старая книга не изменяется.
"""
from copy import copy
from pathlib import Path
import json
import subprocess

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName

ROOT = Path(__file__).resolve().parents[1]
CLEAN = ROOT / 'Книга' / 'Калькулятор_ставки_часа_чистая.xlsx'
CONTROL = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
HARNESS = ROOT / 'Инструменты' / 'харнесс.js'

branches = [
    ('revenue_npd_avg', 'Выручка · НПД, смешанные заказчики 5%', 'B', {'поля': {'regime': 'npd', 'npd_who': 'mix'}}),
    ('revenue_npd_fiz', 'Выручка · НПД, частные лица 4%', 'C', {'поля': {'regime': 'npd', 'npd_who': 'phys'}}),
    ('revenue_npd_ur', 'Выручка · НПД, компании и ИП 6%', 'D', {'поля': {'regime': 'npd', 'npd_who': 'jur'}}),
    ('revenue_usn_inc', 'Выручка · УСН «Доходы» 6%', 'E', {'поля': {'regime': 'usn6'}}),
    ('revenue_usn_prof', 'Выручка · УСН «Доходы минус расходы» 15%', 'F', {'поля': {'regime': 'usn15'}}),
    ('revenue_ausn_inc', 'Выручка · АУСН «Доходы» 8%', 'G', {'поля': {'regime': 'ausn8'}}),
    ('revenue_ausn_prof', 'Выручка · АУСН «Доходы минус расходы» 20%', 'H', {'поля': {'regime': 'ausn20'}}),
]

control_formula = load_workbook(CONTROL, data_only=False, read_only=True)
control_values = load_workbook(CONTROL, data_only=True, read_only=True)
oldf, oldv = control_formula['05_Расчёт'], control_values['05_Расчёт']

# Сначала доказываем равенство действующего контрольного набора настоящему HTML.
for ident, title, col, override in branches:
    raw = subprocess.check_output(
        ['node', str(HARNESS), str(ROOT), json.dumps(override, ensure_ascii=False)], text=True
    )
    html = float(json.loads(raw)['R'])
    old = float(oldv[f'{col}4'].value)
    if abs(html - old) > 1e-6:
        raise SystemExit(f'{title}: calc.html={html}, 05_Расчёт={old}')

wb = load_workbook(CLEAN, data_only=False)
ws = wb['calc']
start = 144
# Секция добавляется в хвост и потому не сдвигает вручную собранные 141 строку.
ws['A142'] = '11 · СЕМЬ НАЛОГОВЫХ ВЕТОК · контрольная формула выручки'
headers = ['ID', 'Название', 'Значение', 'Ед.', 'Формула / как получено', 'Тип', 'Зависит от', 'Источник / форма', None]
for c, value in enumerate(headers, 1):
    ws.cell(143, c).value = value
# Оформление наследуем от существующих строк того же типа.
for c in range(1, 10):
    for source_row, target_row in [(128, 142), (129, 143)]:
        src, dst = ws.cell(source_row, c), ws.cell(target_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format

for offset, (ident, title, col, override) in enumerate(branches):
    row = start + offset
    formula = oldf[f'{col}4'].value
    values = [ident, title, formula, '₽/год', formula[1:] if isinstance(formula, str) and formula.startswith('=') else formula,
              'calc', 'target_income, total_costs, налоговые параметры', 'Веб/calc.html · старая книга — контроль', None]
    for c, value in enumerate(values, 1):
        ws.cell(row, c).value = value
        src, dst = ws.cell(113, c), ws.cell(row, c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
    if ident in wb.defined_names:
        del wb.defined_names[ident]
    wb.defined_names.add(DefinedName(ident, attr_text=f"'calc'!$C${row}"))

ws['C132'] = '=IF(net_time=0,0,CHOOSE(regime_no,revenue_npd_avg,revenue_npd_fiz,revenue_npd_ur,revenue_usn_inc,revenue_usn_prof,revenue_ausn_inc,revenue_ausn_prof)/net_time)'
ws['G132'] = 'regime_no, revenue_* , net_time'
ws['H132'] = 'Веб/calc.html · семь веток внутри calc'
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = 'auto'
wb.save(CLEAN)

print('Перенесено налоговых веток:', len(branches))
for ident, title, col, override in branches:
    print(f'{ident}: {float(oldv[f"{col}4"].value):.10f} — совпадает с calc.html')
print('Внешних ссылок C132 на 05_Расчёт больше нет')
