#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Создаёт чистую книгу Счётикса рядом с действующей старой книгой.

Копируются только уже проверенные источники: глоссарий, Состав и автоматически
собранные значения по умолчанию. Остальные целевые листы создаются с чистой
структурой и статусом «ожидает сборки».

Скрипт не изменяет действующую книгу и не перезаписывает уже созданную чистую.
"""
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import csv
import subprocess
import sys
import tempfile

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
TARGET = ROOT / 'Книга' / 'Калькулятор_ставки_часа_чистая.xlsx'

if TARGET.exists() and '--пересоздать' not in sys.argv:
    raise SystemExit(f'Чистая книга уже существует: {TARGET}\nДля полного пересоздания: --пересоздать')

src = load_workbook(SOURCE, data_only=False, read_only=True)

# Актуальные дефолты всегда извлекаются из фактического HTML.
with tempfile.NamedTemporaryFile(suffix='.tsv', delete=False) as f:
    defaults_tsv = Path(f.name)
subprocess.check_call([
    sys.executable, str(ROOT / 'Инструменты' / 'значения_по_умолчанию.py'),
    str(defaults_tsv)
])
with defaults_tsv.open(encoding='utf-8') as f:
    defaults = list(csv.reader(f, delimiter='\t'))
# Расчётные величины должны быть числами: их суммируют формулы calc.
for row in defaults[1:]:
    for index in (24, 25, 26):
        if len(row) > index and row[index] != '':
            row[index] = float(row[index])
defaults_tsv.unlink(missing_ok=True)

wb = Workbook()
wb.remove(wb.active)

GREEN = '1B9331'
GREEN_LIGHT = 'E8F5EA'
BLUE = '2F80C4'
BLUE_LIGHT = 'EAF3FA'
GRAY = 'F3F4F6'
DARK = '1A1A1A'
WHITE = 'FFFFFF'
BORDER = Border(bottom=Side(style='thin', color='E5E7EB'))


def sheet(name, headers=None):
    ws = wb.create_sheet(name)
    if headers:
        ws.append(headers)
    return ws


def style_table(ws, header_row=1, freeze=None):
    if ws.max_row >= header_row:
        for c in ws[header_row]:
            if c.value is not None:
                c.fill = PatternFill('solid', fgColor=GREEN)
                c.font = Font(color=WHITE, bold=True)
                c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[header_row].height = 32
        ws.auto_filter.ref = f'A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}'
    for row in ws.iter_rows(min_row=header_row + 1):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER
    if freeze:
        ws.freeze_panes = freeze
    for col in range(1, ws.max_column + 1):
        values = [str(ws.cell(r, col).value or '') for r in range(1, min(ws.max_row, 120) + 1)]
        width = min(max([len(v) for v in values] + [10]) + 2, 48)
        ws.column_dimensions[get_column_letter(col)].width = width


# 00 · карта новой книги
ws = sheet('00_Читать')
ws.append(['СЧЁТИКС · ЧИСТАЯ КНИГА'])
ws.append(['Статус', 'Собирается с нуля. Старая книга остаётся действующим контрольным контуром до завершения миграции.'])
ws.append([])
ws.append(['Лист', 'Назначение', 'Состояние'])
rows = [
    ('00_Читать', 'Карта чистой книги и порядок работы', 'пересобрать последним'),
    ('01_Глоссарий', 'Единственный источник понятий и определений', 'перенесён проверенный лист'),
    ('calc', 'Все параметры, константы и формулы модели', 'ожидает полной сборки из calc.html'),
    ('Состав', 'Группирующие понятия и их состав', 'перенесён проверенный лист'),
    ('Значения_по_умолчанию', 'Все фактические дефолты анкеты и каталогов', 'собран автоматически из calc.html'),
    ('Интерфейс', 'Страницы, блоки и все элементы интерфейса', 'ожидает автоматической сборки'),
    ('Тексты', 'Только используемые тексты с привязкой к элементам', 'ожидает автоматической сборки'),
    ('CSS_и_компоненты', 'Токены, компоненты, адаптив, печать и логотип', 'ожидает автоматической сборки'),
    ('Программа_лояльности', 'Внутренняя модель распределения фонда', 'ожидает переноса проверенной механики'),
    ('Полный_отчёт', 'Все результаты d, форматы, блоки и паспорта', 'ожидает сборки из контракта и паспортов'),
]
for r in rows:
    ws.append(r)
ws['A1'].font = Font(size=18, bold=True, color=GREEN)
for c in ws[4]:
    c.fill = PatternFill('solid', fgColor=GREEN)
    c.font = Font(color=WHITE, bold=True)
for col, width in {'A':28, 'B':72, 'C':42}.items():
    ws.column_dimensions[col].width = width
ws.freeze_panes = 'A5'
for row in ws.iter_rows():
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)

# 01 · проверенный глоссарий. Пустые удалённые строки не переносим.
ws = sheet('01_Глоссарий', ['№', 'Наименование', 'ID', 'Назначение', 'Тип', 'Определение'])
for row in src['01_Глоссарий'].iter_rows(min_row=2, max_col=6, values_only=True):
    vals = list(row)
    if vals[1] is None:
        continue
    ws.append(vals)
style_table(ws, freeze='A2')
ws.column_dimensions['F'].width = 100

# calc · чистая структура, формулы будут переноситься по одной цепочке.
headers = ['Раздел', 'ID', 'Наименование', 'Значение', 'Единица', 'Формула / алгоритм',
           'Тип', 'Зависит от', 'Источник', 'Поле d', 'Паспорт', 'Статус сверки']
ws = sheet('calc', headers)
ws.append(['СЛУЖЕБНОЕ', '', 'Лист будет собран из фактической функции calc() после паспортизации расчётов.',
           '', '', '', '', '', 'Веб/calc.html', '', '', 'ожидает сборки'])
style_table(ws, freeze='A2')

# Состав · проверенный лист, компактно переносим без старых пустых строк.
ws = sheet('Состав')
for row in src['Состав'].iter_rows(max_col=12, values_only=True):
    vals = list(row)
    if any(v is not None for v in vals):
        ws.append(vals)
ws['A1'].font = Font(size=16, bold=True, color=GREEN)
for c in ws[4]:
    c.fill = PatternFill('solid', fgColor=GREEN)
    c.font = Font(color=WHITE, bold=True)
    c.alignment = Alignment(wrap_text=True)
ws.freeze_panes = 'A5'
for col in range(1, 13):
    ws.column_dimensions[get_column_letter(col)].width = 30 if col > 1 else 18
for row in ws.iter_rows():
    for c in row:
        c.alignment = Alignment(vertical='top', wrap_text=True)

# Дефолты · полностью из HTML.
ws = sheet('Значения_по_умолчанию')
for row in defaults:
    ws.append(row)
style_table(ws, freeze='A2')
ws.column_dimensions['H'].width = 52
ws.column_dimensions['M'].width = 70
ws.column_dimensions['Q'].width = 42

# Новые структурные листы.
ws = sheet('Интерфейс', [
    'Страница', 'Порядок страницы', 'Блок ID', '№ блока', 'Название блока',
    'Порядок элемента', 'Тип элемента', 'DOM-ID', 'name', 'data-t',
    'Наименование', 'Параметр calc', 'Поле d', 'Условие показа', 'Источник', 'Статус'])
style_table(ws, freeze='A2')

ws = sheet('Тексты', [
    'Страница', 'Блок ID', 'Элемент ID', 'data-t', 'Тип текста', 'Текст',
    'Условие показа', 'Используется в HTML', 'Статус'])
style_table(ws, freeze='A2')
ws.column_dimensions['F'].width = 100

ws = sheet('CSS_и_компоненты', [
    'Категория', 'Компонент / токен', 'Селектор', 'Свойство', 'Значение',
    'Страница', 'Адаптив', 'Печать', 'Источник', 'Статус'])
style_table(ws, freeze='A2')

ws = sheet('Программа_лояльности', [
    'Раздел', 'ID', 'Наименование', 'Значение', 'Единица', 'Формула / правило',
    'Зависит от', 'Поле d', 'Источник', 'Статус'])
ws.append(['СЛУЖЕБНОЕ', '', 'Внутренняя механика будет перенесена из calc(), тестов, паспорта и старого 13b.',
           '', '', '', '', '', '', 'ожидает сборки'])
style_table(ws, freeze='A2')

ws = sheet('Полный_отчёт', [
    'Блок отчёта', 'Порядок', 'Наименование показателя', 'Понятие глоссария',
    'Поле d', 'Параметр calc', 'Единица', 'Формат показа', 'Условие показа',
    'Паспорт', 'HTML ID / место', 'Зеркальная формула Excel', 'Контрольное значение',
    'Сходится', 'Статус', 'Правило ID'])
ws.append(['СЛУЖЕБНОЕ', '', 'Лист будет собран из контракта calc(), фактического report.html и паспортов.',
           '', '', '', '', '', '', '', '', '', '', '', 'ожидает сборки', ''])
style_table(ws, freeze='A2')

# Общие свойства.
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

TARGET.parent.mkdir(parents=True, exist_ok=True)
wb.save(TARGET)
print(f'Создана чистая книга: {TARGET}')
print(f'Листов: {len(wb.sheetnames)}; глоссарий: {wb["01_Глоссарий"].max_row - 1}; дефолтов: {len(defaults) - 1}')
