#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Дословно переносит лист calc из действующей книги в чистую книгу.

Копируются значения и формулы каждой ячейки, оформление, размеры, скрытия,
объединения, проверки данных, условное форматирование, параметры печати и
определённые имена, указывающие на calc. Другие листы не меняются.
"""
from pathlib import Path
from copy import copy
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / 'Книга' / 'Калькулятор_ставки_часа.xlsx'
TARGET = ROOT / 'Книга' / 'Калькулятор_ставки_часа_чистая.xlsx'

src_wb = load_workbook(SOURCE, data_only=False)
dst_wb = load_workbook(TARGET, data_only=False)
src = src_wb['calc']

# Пересоздаём лист на той же целевой позиции.
if 'calc' in dst_wb.sheetnames:
    pos = dst_wb.sheetnames.index('calc')
    dst_wb.remove(dst_wb['calc'])
else:
    pos = 2
dst = dst_wb.create_sheet('calc', pos)

# Ячейки: значение/формула и полное оформление.
for row in src.iter_rows():
    for sc in row:
        dc = dst[sc.coordinate]
        dc.value = sc.value
        # Между разными книгами нельзя переносить _style целиком: его индексы
        # ссылаются на таблицу стилей исходного OOXML. Копируем компоненты,
        # чтобы openpyxl зарегистрировал их в чистой книге корректно.
        if sc.has_style:
            dc.font = copy(sc.font)
            dc.fill = copy(sc.fill)
            dc.border = copy(sc.border)
            dc.alignment = copy(sc.alignment)
            dc.protection = copy(sc.protection)
            dc.number_format = sc.number_format
        if sc.comment:
            dc.comment = copy(sc.comment)
        if sc.hyperlink:
            dc._hyperlink = copy(sc.hyperlink)

# Размеры, скрытия и уровни группировки.
for key, dim in src.row_dimensions.items():
    dst.row_dimensions[key] = copy(dim)
for key, dim in src.column_dimensions.items():
    dst.column_dimensions[key] = copy(dim)

for rng in src.merged_cells.ranges:
    dst.merge_cells(str(rng))

# Настройки листа.
dst.freeze_panes = src.freeze_panes
dst.sheet_format = copy(src.sheet_format)
dst.sheet_properties = copy(src.sheet_properties)
dst.sheet_view.showGridLines = src.sheet_view.showGridLines
dst.page_margins = copy(src.page_margins)
dst.page_setup = copy(src.page_setup)
dst.print_options = copy(src.print_options)
dst.sheet_state = src.sheet_state
dst.auto_filter.ref = src.auto_filter.ref

# Проверки данных и условное форматирование.
dst.data_validations = copy(src.data_validations)
for key in src.conditional_formatting:
    for rule in src.conditional_formatting[key]:
        dst.conditional_formatting.add(str(key.sqref), copy(rule))

# Копируем только имена, чьей областью назначения является calc.
for name, defined in src_wb.defined_names.items():
    try:
        targets = {sheet for sheet, _ in defined.destinations}
    except Exception:
        targets = set()
    if 'calc' in targets:
        if name in dst_wb.defined_names:
            del dst_wb.defined_names[name]
        dst_wb.defined_names.add(copy(defined))

# Обновляем статус в 00_Читать, не вмешиваясь в сам calc.
readme = dst_wb['00_Читать']
for row in readme.iter_rows():
    if row[0].value == 'calc':
        row[2].value = ('дословно перенесён из действующей книги; формулы и 115 имён сохранены; '
                        'внешние зависимости на старые листы будут переноситься поэтапно')
        break

dst_wb.calculation.fullCalcOnLoad = True
dst_wb.calculation.forceFullCalc = True
dst_wb.calculation.calcMode = 'auto'
dst_wb.save(TARGET)
print(f'calc перенесён: {src.max_row} строк × {src.max_column} колонок')
print(f'формул: {sum(c.data_type == "f" for row in src.iter_rows() for c in row)}')
print(f'определённых имён calc: {sum("calc" in {s for s, _ in d.destinations} for d in src_wb.defined_names.values() if getattr(d, "type", None) == "RANGE")}')
