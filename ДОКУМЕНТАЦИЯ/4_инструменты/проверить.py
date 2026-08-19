#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ПРОВЕРКА ЦЕЛОСТНОСТИ ПРОЕКТА СЧЁТИКС.
Запускать после любой правки:  python3 проверить.py [корень]
Численные проверки выполняют НАСТОЯЩИЙ код calc()/parts(), вырезанный из HTML.
Код возврата: 0 — всё чисто, 1 — есть падения."""
import json, subprocess, sys, os, re, io, collections

КОРЕНЬ = sys.argv[1] if len(sys.argv) > 1 else '/home/user/schetix'
ЗДЕСЬ  = os.path.dirname(os.path.abspath(__file__))
ХАРНЕСС = os.path.join(ЗДЕСЬ, 'харнесс.js')
ok, fail, warn = [], [], []

def проверка(имя, условие, факт=''):
    (ok if условие else fail).append((имя, факт))
def замечание(имя, факт=''):
    warn.append((имя, факт))

def расчёт(**переопр):
    r = subprocess.run(['node', ХАРНЕСС, КОРЕНЬ, json.dumps(переопр)],
                       capture_output=True, text=True)
    if r.returncode or not r.stdout.strip():
        raise RuntimeError((r.stderr or 'пустой вывод').strip()[:200])
    return json.loads(r.stdout)

РЕЖИМЫ = [('npd', 'phys', 'НПД 4 %'), ('npd', 'mix', 'НПД 5 %'), ('npd', 'jur', 'НПД 6 %'),
          ('usn6', 'mix', 'УСН 6 %'), ('usn15', 'mix', 'УСН 15 %'),
          ('ausn8', 'mix', 'АУСН 8 %'), ('ausn20', 'mix', 'АУСН 20 %')]

# ══════════════════════════════ 1. МОДЕЛЬ
print('▶ модель (настоящий calc() из calc.html)')
try:
    базовый = расчёт()
except RuntimeError as e:
    print('  ✗ харнесс не запустился:', e); sys.exit(2)

for код, кто, имя in РЕЖИМЫ:
    d = расчёт(поля={'regime': код, 'npd_who': кто})
    остаток = d['R'] - d['C'] - d['aq'] - d['taxAll'] - d['fundY'] - d['discY']
    проверка(f'на руки = цель · {имя}', abs(остаток - d['Ny']) < 1,
             f"{остаток:,.0f} против {d['Ny']:,.0f}")

d = расчёт(поля={'fund_on': True, 'disc_on': True})
проверка('на руки = цель · с фондом и резервом скидки',
         abs(d['R'] - d['C'] - d['aq'] - d['taxAll'] - d['fundY'] - d['discY'] - d['Ny']) < 1)

d = расчёт(поля={'tax_off': True})
проверка('на руки = цель · без оформления',
         abs(d['R'] - d['C'] - d['aq'] - d['Ny']) < 1)

d = базовый
проверка('иерархия: операционное + резерв + проектное = чистое',
         abs(d['side'] + d['fmT'] + d['pool'] - d['NT']) < 0.01)
проверка('иерархия: профильное + клиентское = проектное',
         abs(d['core'] + d['clT'] - d['pool']) < 0.01)
проверка('иерархия: съёмка + обработка = профильное',
         abs(d['sh'] + d['post'] - d['core']) < 0.01)
проверка('сумма 13 сегментов кольца = выручка',
         abs(sum(d['__parts']) - d['R']) < 1, f"Δ {sum(d['__parts']) - d['R']:.4f}")
проверка('сумма 6 доходных сегментов = доход',
         abs(sum(d['__parts'][:6]) - d['Ny']) < 1)

# инвариант сетки скидок: выручка не зависит от длительности съёмки
выручки = [расчёт(поля={'shoot_manual': str(S)})['R'] for S in (1, 2, 4, 6)]
проверка('сетка скидок: выручка одинакова при любой длительности',
         max(выручки) - min(выручки) < 1, f'разброс {max(выручки)-min(выручки):.2f} ₽')

# ставка падает с ростом длительности
ставки = [расчёт(поля={'shoot_manual': str(S)}) for S in (1, 2, 4, 6)]
ставки = [x['R'] / x['sh'] for x in ставки]
проверка('сетка скидок: ставка часа убывает с длительностью',
         all(ставки[i] > ставки[i+1] for i in range(3)),
         ' → '.join(f'{r:,.0f}' for r in ставки))

# ══════════════════════════════ 2. КРАЕВЫЕ СЛУЧАИ
print('▶ краевые случаи')
d = расчёт(поля={'promo_per_day': '7.5'})
проверка('продвижение 7,5 ч/день не даёт отрицательных часов', d['pool'] > 0,
         f"pool {d['pool']:.0f} ч, ставка {d['R']/d['sh'] if d['sh'] else 0:,.0f} ₽/ч")
d = расчёт(поля={'acq_rate': '30', 'fund_on': True, 'fund_pct': '30',
                 'disc_on': True, 'disc_pct': '15'})
проверка('экстремальные проценты не взрывают выручку', d['R'] < 5_000_000,
         f"R {d['R']:,.0f} при D {1-0.30-0.30-0.15:.2f}")
d = расчёт(поля={'shoot_manual': '0'})
проверка('нулевая длительность съёмки обрабатывается', d['sh'] > 0 or d['R'] == 0,
         f"sh {d['sh']}, R {d['R']:,.0f}")

# точка безубыточности из анкеты должна давать ноль «на руки»
for код, кто, имя in [('npd', 'mix', 'НПД 5 %'), ('usn6', 'mix', 'УСН 6 %'),
                      ('usn15', 'mix', 'УСН 15 %'), ('ausn20', 'mix', 'АУСН 20 %')]:
    d = расчёт(поля={'regime': код, 'npd_who': кто})
    остаток = d['Rb'] - d['C'] - d['aqB'] - d['taxB']
    проверка(f'безубыточность анкеты сходится в ноль · {имя}', abs(остаток) < 1,
             f'{остаток:+,.0f} ₽')

# ══════════════════════════════ 2b. ПРОГРАММА ЛОЯЛЬНОСТИ (блок 03 отчёта)
print('▶ программа лояльности')
import math
def лояльность(d):
    """Повторяет арифметику блока 03 из report.html, чтобы проверить
       заявленную книгой сходимость: объём + постоянным + сертификаты = фонд."""
    if not d['discP']: return None
    hp = d['R']/d['pool'] if d['pool'] else 0
    hourAt = lambda S: (S*(1+d['K'])+d['cl'])*hp/S
    S = max(round(d['S'] or 2), 1)
    rRate = lambda x: math.ceil(x/100)*100
    чек = rRate(hourAt(S))*S
    фонд = round(d['discY'])
    пост = 0; заказы = 0; людей = 0
    бюдж = фонд*0.45; цел = d['py']*0.30
    остБ, остЗ = бюдж, цел
    ступени = ((0.05,2,0.40),(0.07,4,0.40),(0.10,12,0.20))
    for p_, s_, w in ступени:
        ценаКл = s_*чек*p_
        поБ = min(бюдж*w, остБ)/ценаКл if ценаКл else 0
        поЗ = min(цел*w, остЗ)/s_
        n = max(0, math.floor(min(поБ, поЗ)))
        сум = round(n*s_*чек*p_)
        остБ -= сум; остЗ -= n*s_
        пост += сум; заказы += n*s_; людей += n
    if людей == 0:
        p_, s_, w = ступени[0]
        цена = round(s_*чек*p_)
        if цена <= бюдж and s_ <= цел:
            пост, заказы = цена, s_
    r500 = lambda n: max(round(n/500)*500, 500)
    бюджСпец = min(фонд*0.30, max(фонд-пост, 0))
    v0, v1, v2 = r500(чек*0.85), r500(чек*0.40), r500(чек*0.25)
    ост = бюджСпец
    n0 = 1 if ост >= v0 else 0; ост -= v0*n0
    n1 = max(0, math.floor(ост*0.45/v1)); ост -= v1*n1
    n2 = max(0, math.floor(ост/v2));      ост -= v2*n2
    спец = v0*n0 + v1*n1 + v2*n2
    объём = max(фонд-пост-спец, 0)
    return dict(фонд=фонд, пост=пост, спец=спец, объём=объём,
                сумма=пост+спец+объём, заказы=заказы, py=d['py'])

for имя, поля in (('резерв 15 %', {'disc_on': True, 'disc_pct': '15'}),
                  ('резерв 10 %', {'disc_on': True, 'disc_pct': '10'}),
                  ('резерв 5 %',  {'disc_on': True, 'disc_pct': '5'}),
                  ('резерв 5 %, съёмка 6 ч', {'disc_on': True, 'disc_pct': '5', 'shoot_manual': '6'}),
                  ('резерв 5 %, съёмка 8 ч', {'disc_on': True, 'disc_pct': '5', 'shoot_manual': '8'})):
    л = лояльность(расчёт(поля=поля))
    проверка(f'фонд лояльности сходится · {имя}', abs(л['сумма']-л['фонд']) <= 3,
             f"сумма направлений {л['сумма']:,} против фонда {л['фонд']:,}")
    проверка(f'постоянные ≤ 30 % потока заказов · {имя}', л['заказы'] <= л['py']*0.31,
             f"{л['заказы']:.0f} из {л['py']:.0f} = {л['заказы']/л['py']*100:.0f} %")

# ══════════════════════════════ 3. СОГЛАСОВАННОСТЬ ФАЙЛОВ
print('▶ согласованность файлов')
def читать(п): return io.open(os.path.join(КОРЕНЬ, п), encoding='utf-8').read()
calc, rep, karkas = читать('web/calc.html'), читать('web/report.html'), читать('web/части/каркас.html')

проверка('report.html совпадает с каркасом', rep == karkas,
         f'{sum(1 for a, b in zip(rep.splitlines(), karkas.splitlines()) if a != b)} строк расходятся')
# Отчёт больше не хранит свою ставку эквайринга: он берёт фактическую
# из анкеты (переменная AQФАКТ). Проверяем, что зашитого числа не осталось.
aq_rep = re.search(r'AQ\s*:\s*(0\.\d+)', rep)
aq_form = re.search(r'id="acq_rate"[^>]*value="([\d.]+)"', calc)
проверка('ставка эквайринга берётся из анкеты, а не зашита в отчёт',
         aq_rep is None and 'AQФАКТ' in rep,
         f'в отчёте осталось зашитое {aq_rep.group(1) if aq_rep else "—"}'
         + (f' · в анкете {aq_form.group(1)} %' if aq_form else ''))

for имя, файл in (('calc', calc), ('report', rep)):
    css = '\n'.join(re.findall(r'<style[^>]*>([\s\S]*?)</style>', файл))
    проверка(f'{имя}: баланс CSS-скобок', css.count('{') == css.count('}'),
             f"{css.count('{') - css.count('}'):+d}")
    исп = set(re.findall(r'var\((--[\w-]+)\)', файл))
    объ = set(re.findall(r'(--[\w-]+)\s*:', файл))
    проверка(f'{имя}: все CSS-переменные объявлены', not (исп - объ), ', '.join(sorted(исп - объ)))

# баланс HTML-тегов во всех трёх файлах
VOID = {'br','img','input','meta','link','hr','source','area','col','embed','track','wbr',
        'circle','path','rect','polygon','line','ellipse','stop','use','polyline','animate'}
def баланс_тегов(s):
    s = re.sub(r'(<script[^>]*>)([\s\S]*?)(</script>)',
               lambda m: m.group(1) + re.sub(r'[^\n]', ' ', m.group(2)) + m.group(3), s)
    s = re.sub(r'(<style[^>]*>)([\s\S]*?)(</style>)',
               lambda m: m.group(1) + re.sub(r'[^\n]', ' ', m.group(2)) + m.group(3), s)
    s = re.sub(r'<!--[\s\S]*?-->', lambda m: re.sub(r'[^\n]', ' ', m.group(0)), s)
    стек, беды = [], []
    for m in re.finditer(r'<(/?)([a-zA-Z][\w-]*)([^>]*?)(/?)>', s):
        зак, имя, само = m.group(1), m.group(2).lower(), m.group(4)
        if имя in VOID or само == '/': continue
        if зак:
            if стек and стек[-1][0] == имя: стек.pop()
            else: беды.append('лишний </%s> поз.%d' % (имя, m.start()))
        else: стек.append((имя, m.start()))
    return беды + ['не закрыт <%s> поз.%d' % (и, п) for и, п in стек]

for имя, файл in (('calc', calc), ('report', rep), ('index', читать('web/index.html'))):
    б = баланс_тегов(файл)
    проверка(f'{имя}: баланс HTML-тегов', not б, '; '.join(б[:3]))

# 24 прямых потомка .wp
def детей(s):
    s = re.sub(r'<style[^>]*>[\s\S]*?</style>|<script[^>]*>[\s\S]*?</script>|<!--[\s\S]*?-->', '', s)
    VOID = {'br','img','input','meta','link','hr','source','area','col','embed','track','wbr',
            'circle','path','rect','polygon','line','ellipse','stop','use','polyline'}
    н = s.index('<div class="wp">') + len('<div class="wp">')
    гл, дети, тек = 0, 0, None
    for m in re.finditer(r'<(/?)([a-zA-Z][\w-]*)([^>]*?)(/?)>', s[н:]):
        зак, имя_, само = m.group(1), m.group(2).lower(), m.group(4)
        if зак:
            гл -= 1
            if гл == 0 and тек is not None: дети += 1; тек = None
            if гл < 0: break
        else:
            if имя_ in VOID or само == '/': continue
            if гл == 0: тек = m.start()
            гл += 1
    return дети
n = детей(rep)
проверка("в .wp ровно 21 блок (как ждёт собрать.py)", n == 21, f"{n}")

# логотип: 8 правил книги
for имя, файл in (('calc', calc), ('report', rep)):
    знаки = re.findall(r'<svg class="(?:bx|lx)".*?</svg>', файл, re.S)
    проверка(f'{имя}: зелёный ромб — правый (x=19.81)',
             all(re.search(r'x="19\.81"[^/]*?fill="var\(--c-g-500\)"', b) for b in знаки))
    проверка(f'{имя}: ровно один зелёный ромб на знак',
             all(len(re.findall(r'--c-g-500', b)) == 1 for b in знаки))
    проверка(f'{имя}: геометрия знака 10.10 / rx 1.90',
             all('width="10.10"' in b and 'rx="1.90"' in b for b in знаки))
    проверка(f'{имя}: нет следов анимации знака',
             not re.search(r'u-(inc|exp|res|prof)', файл))
    проверка(f'{имя}: написание «счётикс» с буквой ё',
             'счетикс' not in файл.lower() or 'счётикс' in файл.lower(),
             'найдено «счетикс» без ё' if 'счетикс' in файл.lower() else '')
проверка('каркас: написание «счётикс» с буквой ё', 'счетикс' not in karkas.lower())

# JS обоих файлов синтаксически валиден
for имя, путь in (('calc', 'web/calc.html'), ('report', 'web/report.html'), ('index', 'web/index.html')):
    s = читать(путь)
    код = '\n'.join(re.findall(r'<script[^>]*>([\s\S]*?)</script>', s))
    врем = '/tmp/_js_%s.js' % имя
    io.open(врем, 'w', encoding='utf-8').write(код)
    p = subprocess.run(['node', '-e',
        "new Function(require('fs').readFileSync(process.argv[1],'utf8'))", врем],
        capture_output=True, text=True)
    os.remove(врем)
    проверка(f'{имя}: JS компилируется', p.returncode == 0, (p.stderr or '').strip().split('\n')[0][:120])

# ══════════════════════════════ 4. КНИГА ↔ КОД
print('▶ книга ↔ код')
try:
    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(КОРЕНЬ, 'Калькулятор_ставки_часа.xlsx'))
    ws = wb['02_Параметры']
    СВЯЗЬ = {'income_month': 'income_month', 'current_rate': 'current_rate', 'frames_out': 'frames_out',
             'shoot_duration': 'shoot_manual', 'post_ratio': 'post_ratio', 'client_time': 'client_time',
             'promo_per_day': 'promo_per_day', 'acc_per_quarter': 'acc_per_quarter',
             'acc_cost_month': 'acc_cost_month', 'home_rent': 'home_rent', 'home_util': 'home_util',
             'home_area': 'home_area', 'cab_area': 'cab_area', 'office_rent': 'office_rent',
             'office_util': 'office_util', 'edu_life': 'edu_life', 'site_cost': 'site_cost',
             'site_hours': 'site_hours', 'site_life': 'site_life', 'fm_pct': 'fm_pct'}
    расх = []
    for r in range(1, ws.max_row + 1):
        i = ws.cell(r, 1).value
        if not isinstance(i, str) or i.strip() not in СВЯЗЬ: continue
        i = i.strip(); v = ws.cell(r, 3).value
        m = re.search(r'id="%s"[^>]*value="([^"]*)"' % СВЯЗЬ[i], calc)
        if not m: continue
        try:
            bv = float(v); hv = float(m.group(1).replace(' ', ''))
            if i in ('fm_pct',) and bv < 1: bv *= 100
            if abs(bv - hv) > 1e-9: расх.append(f'{i}: книга {bv:g} ≠ форма {hv:g}')
        except Exception: pass
    проверка('значения по умолчанию совпадают с книгой', not расх, '; '.join(расх))

    # ── переключатели книги против анкеты. Раньше не сверялись, и книга
    # считала с включённым резервом на форс-мажоры, а анкета — без него:
    # ставка расходилась на 496 ₽ (9 926 против 9 430).
    ПЕРЕКЛ = {'fm_on':   ('Заложить резерв',        r'id="fm_on"[^>]*checked'),
              'acc_mode':('Веду самостоятельно',    r'name="acc_mode" value="self" checked'),
              'ws_mode': ('Работаю из дома',        r'name="ws_mode" value="home" checked'),
              'site_mode':('Нанимал (а) специалиста', r'name="site_mode" value="hired" checked')}
    плохо = []
    for r in range(1, ws.max_row + 1):
        и = ws.cell(r, 1).value
        if not isinstance(и, str) or и.strip() not in ПЕРЕКЛ: continue
        и = и.strip(); знач = str(ws.cell(r, 3).value or '').strip()
        вкл_текст, шаблон = ПЕРЕКЛ[и]
        книга_вкл = (знач == вкл_текст)
        форма_вкл = bool(re.search(шаблон, calc))
        if книга_вкл != форма_вкл:
            плохо.append(f'{и}: книга «{знач}», в анкете {"включено" if форма_вкл else "выключено"}')
    проверка('переключатели книги совпадают с анкетой', not плохо, '; '.join(плохо))

    # Пересчитана ли книга: у формул должны быть сохранённые значения.
    # Без них проверки не могут сверять расчёт кода с расчётом книги.
    wbv = openpyxl.load_workbook(os.path.join(КОРЕНЬ, 'Калькулятор_ставки_часа.xlsx'), data_only=True)
    формул = пусто = 0
    for имя_л in wb.sheetnames:
        л, лв = wb[имя_л], wbv[имя_л]
        for стр in л.iter_rows():
            for c in стр:
                if isinstance(c.value, str) and c.value.startswith('='):
                    формул += 1
                    if лв[c.coordinate].value is None: пусто += 1
    проверка('книга пересчитана (у формул есть значения)', пусто == 0,
             f'{пусто} из {формул} формул без результата — см. Первый_запуск/02_ПЕРЕСЧЁТ_КНИГИ.md')

    # налоговые константы
    конст = {}
    for r in range(1, ws.max_row + 1):
        i = ws.cell(r, 1).value
        if isinstance(i, str): конст[i.strip()] = ws.cell(r, 3).value
    пары = [('fixed_contrib', r'FIX=(\d+)', 1), ('contrib_threshold', r'THR=(\d+)', 1),
            ('contrib_cap', r'CAP=(\d+)', 1), ('min_tax_usn', r'M15=\.?(\d+)', 0.01),
            ('min_tax_ausn', r'M20=\.?(\d+)', 0.01), ('limit_npd', r'npd:\s*\[(\d+)', 1),
            ('limit_ausn', r'ausn8:\s*\[(\d+)', 1)]
    плохо = []
    for ид, шаб, множ in пары:
        m = re.search(шаб, calc)
        if not m or ид not in конст: continue
        зн = float(m.group(1)) * (множ if множ != 0.01 else 0.01)
        if abs(float(конст[ид]) - зн) > 1e-9: плохо.append(f'{ид}: книга {конст[ид]} ≠ код {зн:g}')
    проверка('налоговые константы совпадают с книгой', not плохо, '; '.join(плохо))
except ImportError:
    замечание('openpyxl не установлен — сверка с книгой пропущена')

# ── демо-набор отчёта должен совпадать с расчётом по умолчанию
# Иначе отчёт, открытый без данных, показывает устаревшие числа
# (так «ставка в ноль» разошлась: 3 331 в демо против 2 814 в расчёте).
m = re.search(r'var DEMO=\{(.*?)\};', rep, re.S)
if not m:
    fail.append(('демо-набор найден в report.html', ''))
else:
    демо = {}
    for k, v in re.findall(r'(\w+):(-?[\d.]+)(?=[,}])', m.group(1)):
        демо[k] = float(v)
    расх = []
    for k, v in демо.items():
        если_есть = базовый.get(k)
        if если_есть is None: continue
        доп = max(1e-4, abs(если_есть) * 1e-5)
        if abs(если_есть - v) > доп:
            расх.append(f'{k}: демо {v:g} ≠ расчёт {если_есть:g}')
    проверка('демо-набор совпадает с расчётом по умолчанию', not расх,
             '; '.join(расх[:4]) + (f' и ещё {len(расх)-4}' if len(расх) > 4 else ''))


# ── справочник ↔ таблицы: ссылки не должны вести в никуда,
# а записи не должны висеть без единой ссылки.
try:
    спр = json.loads(re.search(r'var СПР = (\[.*?\]);\n', rep, re.S).group(1))
    имена = {т for т, _ in спр}

    def объекты(текст, начало):
        """Куски вида «начало{...}» с учётом вложенных скобок."""
        куски = []
        for m in re.finditer(начало, текст):
            i = текст.index('{', m.end() - 1)
            гл, j = 0, i
            while j < len(текст):
                if текст[j] == '{': гл += 1
                elif текст[j] == '}':
                    гл -= 1
                    if гл == 0: break
                j += 1
            куски.append(текст[i:j + 1])
        return куски

    куски = объекты(rep, r'СВЯЗЬ\w* = ') + объекты(rep, r'связь:\s*') + объекты(rep, r'связь=')
    цели, битые = set(), []
    for кусок in куски:
        пары = (re.findall(r"'([^']+)'\s*:\s*'([^']+)'", кусок)
                + re.findall(r'"([^"]+)"\s*:\s*"([^"]+)"', кусок))   # JSON-кавычки
        for ключ, цель in пары:
            if not re.match(r'^[А-ЯЁ]', цель): continue      # записи начинаются с заглавной
            цели.add(цель)
            if цель not in имена: битые.append(f'{ключ} → {цель}')
    for цель in re.findall(r"ссылка(?:_на)?\('[^']+',\s*'([^']+)'\)", rep):
        цели.add(цель)
        if цель not in имена: битые.append(f'заголовок → {цель}')
    проверка('ссылки таблиц ведут на существующие записи справочника', not битые,
             '; '.join(sorted(set(битые))[:4]))
    сироты = sorted(имена - цели)
    if сироты:
        замечание(f'записей справочника без ссылок: {len(сироты)}',
                  ', '.join(сироты[:6]) + ('…' if len(сироты) > 6 else ''))
except Exception as e:
    замечание('связи справочника не разобраны', str(e)[:80])

# ══════════════════════════════ 5. ПУБЛИКАЦИОННЫЙ КОНТУР (замечания, не падения)
print('▶ готовность к публикации')
# DEV_NO_BLUR остался в архивной версии с платной стеной, в рабочих файлах
# его нет и быть не должно. Следим только за счётчиком и адресом оплаты.
for имя, шаб in (('YM_ID', r'YM_ID'), ('PAY_URL', r'PAY_URL')):
    if not re.search(шаб, calc): замечание(f'{имя} отсутствует в calc.html')
# THANKS_URL больше не используется: сердечки открывают окно благодарности.
if "'#plan'" in rep or 'href="#plan"' in rep: замечание('ссылка «30 постов» — заглушка #plan')

# ══════════════════════════════ ИТОГ
print()
for имя, факт in ok:   print(f'  ✓ {имя}' + (f'  — {факт}' if факт else ''))
for имя, факт in fail: print(f'  ✗ {имя}' + (f'  — {факт}' if факт else ''))
for имя, факт in warn: print(f'  ! {имя}' + (f'  — {факт}' if факт else ''))
print(f'\nитого: {len(ok)} прошло · {len(fail)} упало · {len(warn)} замечаний')
sys.exit(1 if fail else 0)
